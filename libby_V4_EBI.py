import tkinter as tk
import threading
import os
import requests
import chromadb
import openpyxl
import fitz
import pandas as pd
from sentence_transformers import SentenceTransformer

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
BASE_PATH       = "C:\\RAG Test"
DB_PATH         = "C:\\RAG Test\\rag_db"
COLLECTION_NAME = "knowledge_v4"
TOP_K           = 5
MIN_SCORE       = 1.4
OLLAMA_URL      = "http://localhost:11434/api/generate"
OLLAMA_MODEL    = "phi3:mini"

SYSTEM_PROMPT = """You are Libby V4, a professional enterprise business intelligence assistant.
You have access to business data including sales, inventory, purchase orders, sales orders, quotes and pricing.

Rules:
- Answer ONLY from the provided context. Never use outside knowledge.
- When numerical data is provided, perform the requested calculations accurately.
- Always show your working for calculations (e.g. "Total: $1,200 + $800 = $2,000")
- Format currency with $ and two decimal places e.g. $1,234.56
- Format percentages with % e.g. 12.5%
- Use bullet points "•" for lists
- Use numbered steps for processes
- When comparing periods, clearly label each (e.g. "Q1 2024 vs Q1 2025")
- If data is not in the context, say exactly: "Data not found in knowledge base."
- Never guess or estimate. Never use outside knowledge.
- Be concise and professional."""

# ─────────────────────────────────────────────
# THEMES — Dark Navy & Silver
# ─────────────────────────────────────────────
THEMES = {
    "dark": {
        "BG_DARK":       "#0d1117",   # GitHub-dark near black
        "BG_PANEL":      "#161b22",   # Deep navy panel
        "BG_INPUT":      "#0d1117",   # Input bg
        "BG_BUBBLE_U":   "#1c2128",   # User bubble
        "BG_BUBBLE_L":   "#161b22",   # Libby bubble
        "BG_SIDEBAR":    "#0d1117",   # Sidebar
        "ACCENT":        "#58a6ff",   # Bright blue accent
        "ACCENT_DIM":    "#1f6feb",   # Dimmer blue
        "TEXT_MAIN":     "#c9d1d9",   # Silver-white text
        "TEXT_DIM":      "#484f58",   # Muted gray
        "TEXT_SOURCE":   "#58a6ff",   # Blue source text
        "SUCCESS":       "#3fb950",   # Green success
        "WARNING":       "#d29922",   # Amber warning
        "BORDER":        "#30363d",   # Subtle border
        "BTN_CLEAR":     "#21262d",
        "BTN_CLEAR_FG":  "#f85149",
        "BG_CALC":       "#0d2137",   # Calculation result bg
        "TEXT_CALC":     "#79c0ff",   # Calculation result text
    },
    "light": {
        "BG_DARK":       "#f6f8fa",
        "BG_PANEL":      "#ffffff",
        "BG_INPUT":      "#f6f8fa",
        "BG_BUBBLE_U":   "#ddf4ff",
        "BG_BUBBLE_L":   "#f6f8fa",
        "BG_SIDEBAR":    "#f6f8fa",
        "ACCENT":        "#0969da",
        "ACCENT_DIM":    "#0550ae",
        "TEXT_MAIN":     "#24292f",
        "TEXT_DIM":      "#6e7781",
        "TEXT_SOURCE":   "#0969da",
        "SUCCESS":       "#1a7f37",
        "WARNING":       "#9a6700",
        "BORDER":        "#d0d7de",
        "BTN_CLEAR":     "#ffebe9",
        "BTN_CLEAR_FG":  "#cf222e",
        "BG_CALC":       "#ddf4ff",
        "TEXT_CALC":     "#0969da",
    }
}

# ─────────────────────────────────────────────
# PANDAS DATA STORE
# Stores all Excel files as dataframes for calculations
# ─────────────────────────────────────────────
dataframes = {}   # filename → pandas DataFrame dict (one per sheet)

def load_excel_as_dataframe(filepath):
    """Load all sheets of an Excel file into pandas DataFrames."""
    filename = os.path.basename(filepath)
    try:
        xl = pd.ExcelFile(filepath)
        sheets = {}
        for sheet in xl.sheet_names:
            df = pd.read_excel(filepath, sheet_name=sheet)
            # Clean column names — strip whitespace
            df.columns = [str(c).strip() for c in df.columns]
            sheets[sheet] = df
        dataframes[filename] = sheets
        print(f"  → Loaded {filename} into pandas ({len(sheets)} sheet(s))")
    except Exception as e:
        print(f"  → Could not load {filename} into pandas: {e}")

def perform_calculation(question, filename, sheet_name, df):
    """
    Try to perform calculations on a DataFrame based on the question.
    Returns a calculation result string or None if not applicable.
    """
    q = question.lower()
    result_lines = []

    try:
        # Get numeric columns only
        numeric_cols = df.select_dtypes(include="number").columns.tolist()
        if not numeric_cols:
            return None

        # ── TOTAL ──────────────────────────────
        if any(word in q for word in ["total", "sum", "how much"]):
            for col in numeric_cols:
                total = df[col].sum()
                if total != 0:
                    # Format as currency if col name suggests money
                    if any(w in col.lower() for w in ["price", "cost", "amount", "revenue", "sales", "value", "total"]):
                        result_lines.append(f"• Total {col}: ${total:,.2f}")
                    else:
                        result_lines.append(f"• Total {col}: {total:,.2f}")

        # ── AVERAGE ────────────────────────────
        if any(word in q for word in ["average", "avg", "mean"]):
            for col in numeric_cols:
                avg = df[col].mean()
                if any(w in col.lower() for w in ["price", "cost", "amount", "revenue", "sales", "value"]):
                    result_lines.append(f"• Average {col}: ${avg:,.2f}")
                else:
                    result_lines.append(f"• Average {col}: {avg:,.2f}")

        # ── HIGHEST ────────────────────────────
        if any(word in q for word in ["highest", "maximum", "max", "most", "top"]):
            for col in numeric_cols:
                max_val = df[col].max()
                max_idx = df[col].idxmax()
                row_info = ""
                # Try to find an identifier column
                id_cols = [c for c in df.columns if any(w in c.lower() for w in ["name", "sku", "id", "product", "item", "description"])]
                if id_cols:
                    row_info = f" ({df[id_cols[0]].iloc[max_idx]})"
                if any(w in col.lower() for w in ["price", "cost", "amount", "revenue", "sales", "value"]):
                    result_lines.append(f"• Highest {col}: ${max_val:,.2f}{row_info}")
                else:
                    result_lines.append(f"• Highest {col}: {max_val:,.2f}{row_info}")

        # ── LOWEST ─────────────────────────────
        if any(word in q for word in ["lowest", "minimum", "min", "least"]):
            for col in numeric_cols:
                min_val = df[col].min()
                min_idx = df[col].idxmin()
                row_info = ""
                id_cols = [c for c in df.columns if any(w in c.lower() for w in ["name", "sku", "id", "product", "item", "description"])]
                if id_cols:
                    row_info = f" ({df[id_cols[0]].iloc[min_idx]})"
                if any(w in col.lower() for w in ["price", "cost", "amount", "revenue", "sales", "value"]):
                    result_lines.append(f"• Lowest {col}: ${min_val:,.2f}{row_info}")
                else:
                    result_lines.append(f"• Lowest {col}: {min_val:,.2f}{row_info}")

        # ── COUNT ──────────────────────────────
        if any(word in q for word in ["how many", "count", "number of"]):
            result_lines.append(f"• Total records: {len(df):,}")

        # ── PERCENTAGE ─────────────────────────
        if any(word in q for word in ["percent", "percentage", "%"]):
            for col in numeric_cols:
                total = df[col].sum()
                if total > 0:
                    result_lines.append(f"• {col} breakdown available ({total:,.2f} total)")

    except Exception as e:
        print(f"Calculation error: {e}")
        return None

    if result_lines:
        header = f"📊 Calculated from: {filename} — {sheet_name}"
        return header + "\n" + "\n".join(result_lines)
    return None

# ─────────────────────────────────────────────
# FILE READERS
# ─────────────────────────────────────────────
def read_excel_text(filepath):
    """Convert Excel to text for RAG indexing."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    all_text = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_lines = [f"Sheet: {sheet_name}"]
        for row in ws.iter_rows(values_only=True):
            cleaned = [str(cell).strip() if cell is not None else "" for cell in row]
            if any(cleaned):
                sheet_lines.append(" | ".join(cleaned))
        if len(sheet_lines) > 1:
            all_text.append("\n".join(sheet_lines))
    wb.close()
    return "\n\n".join(all_text)

def read_pdf(filepath):
    doc = fitz.open(filepath)
    all_text = []
    for page_num in range(len(doc)):
        page = doc[page_num]
        text = page.get_text().strip()
        if text:
            all_text.append(f"[Page {page_num + 1}]\n{text}")
    doc.close()
    return "\n\n".join(all_text)

def chunk_text(text, chunk_size=500, overlap=100):
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunks.append(text[start:end].strip())
        start += chunk_size - overlap
    return [c for c in chunks if len(c) > 50]

# ─────────────────────────────────────────────
# RAG BACKEND
# ─────────────────────────────────────────────
rag_model   = None
collection  = None
known_files = []

def load_backend():
    global rag_model, collection, known_files

    rag_model  = SentenceTransformer("all-MiniLM-L6-v2")
    client     = chromadb.PersistentClient(path=DB_PATH)
    collection = client.get_or_create_collection(name=COLLECTION_NAME)

    existing_ids = set(collection.get()["ids"])
    new_chunks, new_embeddings, new_ids, new_metadata = [], [], [], []

    for root, dirs, files in os.walk(BASE_PATH):
        dirs[:] = [d for d in dirs if "rag_db" not in d]
        for file in files:
            filepath = os.path.join(root, file)

            if file.endswith(".txt"):
                try:
                    with open(filepath, "r", encoding="utf-8") as f:
                        content = f.read()
                    file_type = "txt"
                except Exception as e:
                    print(f"Error reading {file}: {e}")
                    continue

            elif file.endswith(".xlsx"):
                try:
                    content = read_excel_text(filepath)
                    file_type = "xlsx"
                    # Also load into pandas for calculations
                    load_excel_as_dataframe(filepath)
                except Exception as e:
                    print(f"Error reading {file}: {e}")
                    continue

            elif file.endswith(".pdf"):
                try:
                    content = read_pdf(filepath)
                    file_type = "pdf"
                    if not content.strip():
                        continue
                except Exception as e:
                    print(f"Error reading {file}: {e}")
                    continue

            else:
                continue

            chunks = chunk_text(content)
            print(f"  → {len(chunks)} chunks from {file}")

            for i, chunk in enumerate(chunks):
                chunk_id = f"{filepath}::chunk{i}"
                if chunk_id in existing_ids:
                    continue
                new_chunks.append(chunk)
                new_ids.append(chunk_id)
                new_metadata.append({
                    "source":      filepath,
                    "chunk_index": i,
                    "filename":    file,
                    "file_type":   file_type
                })

    if new_chunks:
        new_embeddings = rag_model.encode(new_chunks).tolist()
        collection.add(
            documents=new_chunks,
            embeddings=new_embeddings,
            ids=new_ids,
            metadatas=new_metadata
        )

    all_meta = collection.get(include=["metadatas"])["metadatas"]
    seen = set()
    for m in all_meta:
        src = m.get("source", "")
        if src and src not in seen:
            seen.add(src)
            known_files.append({
                "source":    src,
                "filename":  m.get("filename", os.path.basename(src)),
                "file_type": m.get("file_type", "txt")
            })

def retrieve(question):
    query_embedding = rag_model.encode([question]).tolist()
    results = collection.query(
        query_embeddings=query_embedding,
        n_results=TOP_K,
        include=["documents", "metadatas", "distances"]
    )
    filtered = []
    for doc, meta, dist in zip(
        results["documents"][0],
        results["metadatas"][0],
        results["distances"][0]
    ):
        if dist <= MIN_SCORE:
            filtered.append({
                "text":      doc.strip(),
                "source":    meta.get("source", "Unknown"),
                "filename":  meta.get("filename", "Unknown"),
                "file_type": meta.get("file_type", "txt")
            })
    return filtered

def get_calculations(question):
    """Run pandas calculations across all loaded Excel files."""
    results = []
    for filename, sheets in dataframes.items():
        for sheet_name, df in sheets.items():
            calc = perform_calculation(question, filename, sheet_name, df)
            if calc:
                results.append(calc)
    return results

def ask_ollama(question, chunks, calc_results):
    if not chunks and not calc_results:
        return "Data not found in knowledge base.", []

    context_parts = []

    # Add calculation results first — they are precise
    if calc_results:
        context_parts.append("=== CALCULATED DATA ===")
        context_parts.extend(calc_results)

    # Add retrieved text chunks
    if chunks:
        context_parts.append("=== RETRIEVED DOCUMENTS ===")
        for c in chunks:
            context_parts.append(f"[Source: {c['filename']}]\n{c['text']}")

    context = "\n\n".join(context_parts)
    sources = list(set([c["source"] for c in chunks])) if chunks else []

    prompt = f"""{SYSTEM_PROMPT}

--- BUSINESS DATA CONTEXT ---
{context}
--- END CONTEXT ---

Question: {question}
Answer:"""

    try:
        response = requests.post(
            OLLAMA_URL,
            json={
                "model":   OLLAMA_MODEL,
                "prompt":  prompt,
                "stream":  False,
                "options": {
                    "temperature": 0.05,  # Very low — maximum factual accuracy
                    "top_p":       0.9,
                    "num_predict": 500    # More tokens for detailed business answers
                }
            },
            timeout=120
        )
        response.raise_for_status()
        return response.json().get("response", "").strip(), sources
    except requests.exceptions.ConnectionError:
        return "⚠️  Ollama not running. Open CMD and type: ollama serve", []
    except Exception as e:
        return f"⚠️  Error: {str(e)}", []


# ─────────────────────────────────────────────
# MAIN APP
# ─────────────────────────────────────────────
class LibbyApp:
    def __init__(self, root):
        self.root          = root
        self.ready         = False
        self.current_theme = "dark"
        self.T             = THEMES["dark"]
        self.dot_count     = 0
        self.animating     = False
        self._chat_widgets = []

        self.root.title("Libby V4 — Enterprise Knowledge System")
        self.root.geometry("1200x780")
        self.root.minsize(800, 600)
        self.root.configure(bg=self.T["BG_DARK"])

        self._build_ui()
        self._start_backend()

    def _build_ui(self):
        T = self.T

        # ── Header ──────────────────────────
        self.header = tk.Frame(self.root, bg=T["BG_PANEL"], pady=14)
        self.header.pack(fill="x")

        # Left side — logo and title
        left_header = tk.Frame(self.header, bg=T["BG_PANEL"])
        left_header.pack(side="left", padx=16)

        self.avatar_canvas = tk.Canvas(
            left_header, width=52, height=52,
            bg=T["BG_PANEL"], highlightthickness=0
        )
        self.avatar_canvas.pack(side="left", padx=(0, 12))
        self._draw_avatar()

        self.title_frame = tk.Frame(left_header, bg=T["BG_PANEL"])
        self.title_frame.pack(side="left")

        self.title_label = tk.Label(
            self.title_frame, text="LIBBY  V4",
            bg=T["BG_PANEL"], fg=T["ACCENT"],
            font=("Courier New", 22, "bold")
        )
        self.title_label.pack(anchor="w")

        self.subtitle_label = tk.Label(
            self.title_frame,
            text="Enterprise Business Intelligence  •  Offline  •  TXT / XLSX / PDF",
            bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
            font=("Courier New", 9)
        )
        self.subtitle_label.pack(anchor="w")

        # Right side — controls
        self.right_frame = tk.Frame(self.header, bg=T["BG_PANEL"])
        self.right_frame.pack(side="right", padx=16)

        self.theme_btn = tk.Button(
            self.right_frame, text="☀️  Light Mode",
            bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
            activebackground=T["BG_INPUT"],
            font=("Courier New", 9),
            relief="flat", cursor="hand2",
            command=self._toggle_theme
        )
        self.theme_btn.pack(anchor="e", pady=(0, 4))

        self.status_label = tk.Label(
            self.right_frame, text="⏳  Loading...",
            bg=T["BG_PANEL"], fg=T["WARNING"],
            font=("Courier New", 9)
        )
        self.status_label.pack(anchor="e")

        self.top_divider = tk.Frame(self.root, bg=T["BORDER"], height=1)
        self.top_divider.pack(fill="x")

        # ── Body ────────────────────────────
        self.body = tk.Frame(self.root, bg=T["BG_DARK"])
        self.body.pack(fill="both", expand=True)

        # ── Sidebar ─────────────────────────
        self.sidebar = tk.Frame(self.body, bg=T["BG_SIDEBAR"], width=240)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)
        self._build_sidebar_contents()

        self.side_divider = tk.Frame(self.body, bg=T["BORDER"], width=1)
        self.side_divider.pack(side="left", fill="y")

        # ── Chat ────────────────────────────
        self.chat_outer = tk.Frame(self.body, bg=T["BG_PANEL"])
        self.chat_outer.pack(side="left", fill="both", expand=True)

        self.chat_canvas = tk.Canvas(
            self.chat_outer, bg=T["BG_PANEL"],
            highlightthickness=0, bd=0
        )
        scrollbar = tk.Scrollbar(self.chat_outer, command=self.chat_canvas.yview)
        self.chat_canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.chat_canvas.pack(side="left", fill="both", expand=True)

        self.chat_inner = tk.Frame(self.chat_canvas, bg=T["BG_PANEL"])
        self.chat_window = self.chat_canvas.create_window(
            (0, 0), window=self.chat_inner, anchor="nw"
        )
        self.chat_inner.bind("<Configure>", self._on_frame_configure)
        self.chat_canvas.bind("<Configure>", self._on_canvas_configure)
        self.chat_canvas.bind_all("<MouseWheel>", self._on_mousewheel)

        self.bottom_divider = tk.Frame(self.root, bg=T["BORDER"], height=1)
        self.bottom_divider.pack(fill="x")

        # ── Input ────────────────────────────
        self.input_frame = tk.Frame(self.root, bg=T["BG_PANEL"], pady=10)
        self.input_frame.pack(fill="x", padx=12)

        self.input_box = tk.Text(
            self.input_frame, height=3,
            bg=T["BG_INPUT"], fg=T["TEXT_MAIN"],
            insertbackground=T["ACCENT"],
            font=("Courier New", 11),
            relief="flat", bd=0,
            padx=12, pady=10, wrap="word"
        )
        self.input_box.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self.input_box.bind("<Return>", self._on_enter)
        self.input_box.insert("1.0", "Ask a business question...")
        self.input_box.config(fg=T["TEXT_DIM"])
        self.input_box.bind("<FocusIn>", self._clear_placeholder)
        self.input_box.bind("<FocusOut>", self._restore_placeholder)

        self.btn_frame = tk.Frame(self.input_frame, bg=T["BG_PANEL"])
        self.btn_frame.pack(side="right")

        self.send_btn = tk.Button(
            self.btn_frame, text="QUERY",
            bg=T["ACCENT_DIM"], fg=T["TEXT_MAIN"],
            activebackground=T["ACCENT"],
            font=("Courier New", 11, "bold"),
            relief="flat", padx=16, pady=6,
            cursor="hand2", command=self._send_question
        )
        self.send_btn.pack(pady=(0, 4))

        self.clear_btn = tk.Button(
            self.btn_frame, text="🗑  Clear",
            bg=T["BTN_CLEAR"], fg=T["BTN_CLEAR_FG"],
            activebackground=T["BTN_CLEAR"],
            font=("Courier New", 9),
            relief="flat", padx=16, pady=4,
            cursor="hand2", command=self._clear_chat
        )
        self.clear_btn.pack()

        # ── Footer ───────────────────────────
        self.footer = tk.Frame(self.root, bg=T["BG_PANEL"], pady=5)
        self.footer.pack(fill="x")
        self.footer_label = tk.Label(
            self.footer,
            text="🔒  Air-gapped  •  Enterprise Mode  •  Sales / Inventory / PO / SO / Pricing",
            bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
            font=("Courier New", 8)
        )
        self.footer_label.pack()

        self._add_libby_message(
            "LIBBY V4 — Enterprise Business Intelligence.\n"
            "I can answer questions about sales, inventory, purchase orders,\n"
            "sales orders, quotes and pricing — and perform calculations.\n"
            "Ask me anything from your business data.",
            source=None, file_type=None, is_calc=False
        )

    # ─────────────────────────────────────────
    # SIDEBAR
    # ─────────────────────────────────────────
    def _build_sidebar_contents(self):
        T = self.T

        self.sidebar_title = tk.Label(
            self.sidebar, text="📁  DATA SOURCES",
            bg=T["BG_SIDEBAR"], fg=T["ACCENT"],
            font=("Courier New", 9, "bold"), pady=12
        )
        self.sidebar_title.pack(anchor="w", padx=12)

        tk.Frame(self.sidebar, bg=T["BORDER"], height=1).pack(fill="x", padx=8)

        # Pandas status section
        self.pandas_label = tk.Label(
            self.sidebar, text="📊  CALCULATION ENGINE",
            bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"],
            font=("Courier New", 7, "bold"), pady=6
        )
        self.pandas_label.pack(anchor="w", padx=12)

        self.pandas_status = tk.Label(
            self.sidebar, text="⏳  Loading...",
            bg=T["BG_SIDEBAR"], fg=T["WARNING"],
            font=("Courier New", 7)
        )
        self.pandas_status.pack(anchor="w", padx=16)

        tk.Frame(self.sidebar, bg=T["BORDER"], height=1).pack(fill="x", padx=8, pady=4)

        self.sidebar_list = tk.Frame(self.sidebar, bg=T["BG_SIDEBAR"])
        self.sidebar_list.pack(fill="both", expand=True, padx=8, pady=4)

        self.sidebar_footer_label = tk.Label(
            self.sidebar,
            text="📄 TXT   📊 XLSX   📕 PDF",
            bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"],
            font=("Courier New", 7), pady=8
        )
        self.sidebar_footer_label.pack(side="bottom")

    def _refresh_sidebar(self):
        T = self.T
        for widget in self.sidebar_list.winfo_children():
            widget.destroy()

        # Update pandas status
        xlsx_count = len(dataframes)
        self.pandas_status.config(
            text=f"✅  {xlsx_count} Excel file(s) ready for calculation",
            fg=T["SUCCESS"]
        )

        for entry in known_files:
            filename  = entry["filename"]
            folder    = os.path.basename(os.path.dirname(entry["source"]))
            file_type = entry.get("file_type", "txt")
            icon = "📊" if file_type == "xlsx" else ("📕" if file_type == "pdf" else "📄")

            # Show if file is calculation-enabled
            calc_badge = " ⚡" if file_type == "xlsx" else ""

            row = tk.Frame(self.sidebar_list, bg=T["BG_SIDEBAR"])
            row.pack(anchor="w", fill="x", pady=2)
            tk.Label(
                row, text=f"{icon} {filename}{calc_badge}",
                bg=T["BG_SIDEBAR"], fg=T["TEXT_MAIN"],
                font=("Courier New", 8),
                wraplength=200, justify="left", anchor="w"
            ).pack(anchor="w")
            tk.Label(
                row, text=f"   📁 {folder}",
                bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"],
                font=("Courier New", 7), anchor="w"
            ).pack(anchor="w", pady=(0, 4))

    # ─────────────────────────────────────────
    # AVATAR
    # ─────────────────────────────────────────
    def _draw_avatar(self):
        c = self.avatar_canvas
        T = self.T
        c.delete("all")
        # Corporate look — clean geometric design
        c.create_rectangle(4, 4, 48, 48, fill=T["BG_PANEL"], outline=T["ACCENT"], width=2)
        c.create_rectangle(8, 8, 44, 26, fill=T["ACCENT_DIM"], outline="", width=0)
        # Bar chart icon
        c.create_rectangle(10, 32, 17, 44, fill=T["ACCENT"], outline="")
        c.create_rectangle(20, 26, 27, 44, fill=T["ACCENT"], outline="")
        c.create_rectangle(30, 30, 37, 44, fill=T["ACCENT"], outline="")
        c.create_rectangle(40, 22, 44, 44, fill=T["SUCCESS"], outline="")
        tk.Label(c, text="V4", bg=T["BG_PANEL"], fg=T["TEXT_MAIN"],
                 font=("Courier New", 7, "bold")).place(x=11, y=10)

    # ─────────────────────────────────────────
    # THEME
    # ─────────────────────────────────────────
    def _toggle_theme(self):
        self.current_theme = "light" if self.current_theme == "dark" else "dark"
        self.T = THEMES[self.current_theme]
        self.theme_btn.config(
            text="🌙  Dark Mode" if self.current_theme == "light" else "☀️  Light Mode"
        )
        self._apply_theme()

    def _apply_theme(self):
        T = self.T
        self.root.configure(bg=T["BG_DARK"])
        self.header.configure(bg=T["BG_PANEL"])
        self.avatar_canvas.configure(bg=T["BG_PANEL"])
        self._draw_avatar()
        self.title_frame.configure(bg=T["BG_PANEL"])
        self.title_label.configure(bg=T["BG_PANEL"], fg=T["ACCENT"])
        self.subtitle_label.configure(bg=T["BG_PANEL"], fg=T["TEXT_DIM"])
        self.right_frame.configure(bg=T["BG_PANEL"])
        self.theme_btn.configure(bg=T["BG_PANEL"], fg=T["TEXT_DIM"])
        self.status_label.configure(bg=T["BG_PANEL"])
        self.top_divider.configure(bg=T["BORDER"])
        self.body.configure(bg=T["BG_DARK"])
        self.sidebar.configure(bg=T["BG_SIDEBAR"])
        self.sidebar_title.configure(bg=T["BG_SIDEBAR"], fg=T["ACCENT"])
        self.pandas_label.configure(bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"])
        self.pandas_status.configure(bg=T["BG_SIDEBAR"])
        self.sidebar_list.configure(bg=T["BG_SIDEBAR"])
        self.sidebar_footer_label.configure(bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"])
        self.side_divider.configure(bg=T["BORDER"])
        self.chat_outer.configure(bg=T["BG_PANEL"])
        self.chat_canvas.configure(bg=T["BG_PANEL"])
        self.chat_inner.configure(bg=T["BG_PANEL"])
        self.bottom_divider.configure(bg=T["BORDER"])
        self.input_frame.configure(bg=T["BG_PANEL"])
        self.btn_frame.configure(bg=T["BG_PANEL"])
        self.input_box.configure(bg=T["BG_INPUT"], fg=T["TEXT_MAIN"], insertbackground=T["ACCENT"])
        self.send_btn.configure(bg=T["ACCENT_DIM"], fg=T["TEXT_MAIN"])
        self.clear_btn.configure(bg=T["BTN_CLEAR"], fg=T["BTN_CLEAR_FG"])
        self.footer.configure(bg=T["BG_PANEL"])
        self.footer_label.configure(bg=T["BG_PANEL"], fg=T["TEXT_DIM"])
        self.status_label.configure(
            fg=T["SUCCESS"] if self.current_theme == "light" else T["WARNING"]
        )
        self._refresh_sidebar()
        self._refresh_chat_bubbles()

    # ─────────────────────────────────────────
    # CHAT BUBBLES
    # ─────────────────────────────────────────
    def _add_user_message(self, text):
        T     = self.T
        outer = tk.Frame(self.chat_inner, bg=T["BG_PANEL"], pady=4)
        outer.pack(fill="x", padx=10)
        self._chat_widgets.append(("user", text, None, None, False))
        tk.Label(outer, text="QUERY", bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
                 font=("Courier New", 8, "bold")).pack(anchor="e", padx=6)
        bubble = tk.Frame(outer, bg=T["BG_BUBBLE_U"], padx=14, pady=8)
        bubble.pack(anchor="e", padx=6)
        tk.Label(bubble, text=text, bg=T["BG_BUBBLE_U"], fg=T["TEXT_MAIN"],
                 font=("Courier New", 11), wraplength=680, justify="left").pack(anchor="w")

    def _add_libby_message(self, text, source=None, file_type=None, is_calc=False):
        T     = self.T
        outer = tk.Frame(self.chat_inner, bg=T["BG_PANEL"], pady=4)
        outer.pack(fill="x", padx=10)
        self._chat_widgets.append(("libby", text, source, file_type, is_calc))

        # Label — show calculator icon for calculation results
        label_text = "LIBBY V4  📊" if is_calc else "LIBBY V4"
        tk.Label(outer, text=label_text, bg=T["BG_PANEL"], fg=T["ACCENT"],
                 font=("Courier New", 8, "bold")).pack(anchor="w", padx=6)

        # Use different bubble color for calculations
        bubble_bg = T["BG_CALC"] if is_calc else T["BG_BUBBLE_L"]
        text_color = T["TEXT_CALC"] if is_calc else T["TEXT_MAIN"]

        bubble = tk.Frame(outer, bg=bubble_bg, padx=14, pady=8)
        bubble.pack(anchor="w", padx=6)
        tk.Label(bubble, text=text, bg=bubble_bg, fg=text_color,
                 font=("Courier New", 11), wraplength=680, justify="left").pack(anchor="w")

        if source:
            icon = "📊" if file_type == "xlsx" else ("📕" if file_type == "pdf" else "📄")
            tk.Label(
                bubble, text=f"{icon}  {source}",
                bg=bubble_bg, fg=T["TEXT_SOURCE"],
                font=("Courier New", 8), wraplength=680, justify="left"
            ).pack(anchor="w", pady=(4, 0))

    def _refresh_chat_bubbles(self):
        saved = list(self._chat_widgets)
        for widget in self.chat_inner.winfo_children():
            widget.destroy()
        self._chat_widgets = []
        for kind, text, source, file_type, is_calc in saved:
            if kind == "user":
                self._add_user_message(text)
            else:
                self._add_libby_message(text, source, file_type, is_calc)

    # ─────────────────────────────────────────
    # THINKING ANIMATION
    # ─────────────────────────────────────────
    def _add_thinking_frame(self):
        T     = self.T
        frame = tk.Frame(self.chat_inner, bg=T["BG_PANEL"], pady=4)
        frame.pack(fill="x", padx=10)
        tk.Label(frame, text="LIBBY V4", bg=T["BG_PANEL"], fg=T["ACCENT"],
                 font=("Courier New", 8, "bold")).pack(anchor="w", padx=6)
        self.thinking_label = tk.Label(
            frame, text="📊  Analyzing data",
            bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
            font=("Courier New", 10, "italic")
        )
        self.thinking_label.pack(anchor="w", padx=10)
        self.animating = True
        self._animate_dots()
        self._scroll_to_bottom()
        return frame

    def _animate_dots(self):
        if not self.animating:
            return
        dots = "." * (self.dot_count % 4)
        self.thinking_label.config(text=f"📊  Analyzing data{dots}")
        self.dot_count += 1
        self.root.after(300, self._animate_dots)

    def _stop_animation(self, frame):
        self.animating = False
        frame.destroy()

    # ─────────────────────────────────────────
    # CLEAR
    # ─────────────────────────────────────────
    def _clear_chat(self):
        for widget in self.chat_inner.winfo_children():
            widget.destroy()
        self._chat_widgets = []
        self._add_libby_message(
            "Session cleared. Ready for queries.",
            source=None, file_type=None, is_calc=False
        )

    # ─────────────────────────────────────────
    # SCROLL
    # ─────────────────────────────────────────
    def _on_frame_configure(self, event):
        self.chat_canvas.configure(scrollregion=self.chat_canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.chat_canvas.itemconfig(self.chat_window, width=event.width)

    def _on_mousewheel(self, event):
        self.chat_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _scroll_to_bottom(self):
        self.root.after(50, lambda: self.chat_canvas.yview_moveto(1.0))

    # ─────────────────────────────────────────
    # INPUT
    # ─────────────────────────────────────────
    def _clear_placeholder(self, event):
        if self.input_box.get("1.0", "end-1c") == "Ask a business question...":
            self.input_box.delete("1.0", "end")
            self.input_box.config(fg=self.T["TEXT_MAIN"])

    def _restore_placeholder(self, event):
        if not self.input_box.get("1.0", "end-1c").strip():
            self.input_box.insert("1.0", "Ask a business question...")
            self.input_box.config(fg=self.T["TEXT_DIM"])

    def _on_enter(self, event):
        if not event.state & 0x1:
            self._send_question()
            return "break"

    # ─────────────────────────────────────────
    # SEND & QUERY
    # ─────────────────────────────────────────
    def _send_question(self):
        if not self.ready:
            return
        question = self.input_box.get("1.0", "end-1c").strip()
        if not question or question == "Ask a business question...":
            return
        self.input_box.delete("1.0", "end")
        self.input_box.config(fg=self.T["TEXT_MAIN"])
        self._add_user_message(question)
        thinking_frame = self._add_thinking_frame()
        self.send_btn.config(state="disabled", bg=self.T["BG_INPUT"])
        self.input_box.config(state="disabled")

        def run_query():
            # Run both retrieval and calculations in parallel
            chunks     = retrieve(question)
            calc_results = get_calculations(question)
            answer, sources = ask_ollama(question, chunks, calc_results)
            source_str = ", ".join(sources) if sources else None
            file_type  = chunks[0]["file_type"] if chunks else None
            is_calc    = len(calc_results) > 0
            self.root.after(0, lambda: self._show_answer(
                answer, source_str, file_type, is_calc, thinking_frame
            ))

        threading.Thread(target=run_query, daemon=True).start()

    def _show_answer(self, answer, source, file_type, is_calc, thinking_frame):
        self._stop_animation(thinking_frame)
        self._add_libby_message(answer, source=source, file_type=file_type, is_calc=is_calc)
        self._scroll_to_bottom()
        self.send_btn.config(state="normal", bg=self.T["ACCENT_DIM"])
        self.input_box.config(state="normal")
        self.input_box.focus()
        self.root.update_idletasks()
        self.root.update()

    # ─────────────────────────────────────────
    # BACKEND
    # ─────────────────────────────────────────
    def _start_backend(self):
        def load():
            load_backend()
            self.ready = True
            count = collection.count()
            self.root.after(0, lambda: self.status_label.config(
                text=f"✅  {count} chunks  •  phi3:mini  •  Offline",
                fg=self.T["SUCCESS"]
            ))
            self.root.after(0, self._refresh_sidebar)

        threading.Thread(target=load, daemon=True).start()


# ─────────────────────────────────────────────
# LAUNCH
# ─────────────────────────────────────────────
if __name__ == "__main__":
    root = tk.Tk()
    app  = LibbyApp(root)
    root.mainloop()