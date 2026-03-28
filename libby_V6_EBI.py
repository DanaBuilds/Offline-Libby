import tkinter as tk
from tkinter import filedialog, messagebox
import threading
import os
import requests
import chromadb
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import fitz
import pandas as pd
from sentence_transformers import SentenceTransformer
from datetime import datetime

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
BASE_PATH       = "C:\\RAG Test"
DB_PATH         = "C:\\RAG Test\\rag_db"
COLLECTION_NAME = "knowledge_v4"
TOP_K           = 5
MIN_SCORE       = 1.4
OLLAMA_URL      = "http://localhost:11434/api/generate"
OLLAMA_MODEL    = "phi3:mini"
COMPANY_NAME    = "Libby Enterprise Intelligence"

SYSTEM_PROMPT = """You are Libby V6, an ultra-concise enterprise data retrieval system.

STRICT OUTPUT RULES — no exceptions:
- Return numbers and dollar amounts ONLY. No narrative. No explanation.
- Maximum 3 lines per answer.
- Show formula if calculating: e.g. $6,750 + $4,800 = $11,550
- Format currency: $1,234.56
- Format percentages: 12.5%
- If data is not found: "Not found."
- Never use full sentences. Never explain. Never pad.
- NEVER use markdown. No backticks. No code blocks. No asterisks. Plain text only."""

# ─────────────────────────────────────────────
# THEMES
# ─────────────────────────────────────────────
THEMES = {
    "dark": {
        "BG_DARK":       "#1c1c1e",
        "BG_PANEL":      "#2c2c2e",
        "BG_INPUT":      "#1c1c1e",
        "BG_BUBBLE_U":   "#3a3a3c",
        "BG_BUBBLE_L":   "#2c2c2e",
        "BG_SIDEBAR":    "#1c1c1e",
        "ACCENT":        "#e5e5ea",
        "ACCENT_DIM":    "#636366",
        "TEXT_MAIN":     "#f2f2f7",
        "TEXT_DIM":      "#636366",
        "TEXT_SOURCE":   "#aeaeb2",
        "SUCCESS":       "#30d158",
        "WARNING":       "#ff9f0a",
        "BORDER":        "#38383a",
        "BTN_CLEAR":     "#2c2c2e",
        "BTN_CLEAR_FG":  "#ff453a",
        "BG_CALC":       "#3a3a3c",
        "TEXT_CALC":     "#f2f2f7",
        "TEXT_BUBBLE_U": "#f2f2f7",
        "TEXT_BUBBLE_L": "#f2f2f7",
        "BTN_EXPORT":    "#1a3a1a",
        "BTN_EXPORT_FG": "#30d158",
    },
    "light": {
        "BG_DARK":       "#f2f2f7",
        "BG_PANEL":      "#ffffff",
        "BG_INPUT":      "#f2f2f7",
        "BG_BUBBLE_U":   "#e5e5ea",
        "BG_BUBBLE_L":   "#f2f2f7",
        "BG_SIDEBAR":    "#f2f2f7",
        "ACCENT":        "#1c1c1e",
        "ACCENT_DIM":    "#3a3a3c",
        "TEXT_MAIN":     "#1c1c1e",
        "TEXT_DIM":      "#8e8e93",
        "TEXT_SOURCE":   "#636366",
        "SUCCESS":       "#34c759",
        "WARNING":       "#ff9500",
        "BORDER":        "#d1d1d6",
        "BTN_CLEAR":     "#ffe5e5",
        "BTN_CLEAR_FG":  "#ff3b30",
        "BG_CALC":       "#e5e5ea",
        "TEXT_CALC":     "#1c1c1e",
        "TEXT_BUBBLE_U": "#1c1c1e",
        "TEXT_BUBBLE_L": "#1c1c1e",
        "BTN_EXPORT":    "#e5f5ea",
        "BTN_EXPORT_FG": "#1a7f37",
    }
}

# ─────────────────────────────────────────────
# PANDAS DATA STORE
# ─────────────────────────────────────────────
dataframes = {}

def load_excel_as_dataframe(filepath):
    filename = os.path.basename(filepath)
    try:
        xl = pd.ExcelFile(filepath)
        sheets = {}
        for sheet in xl.sheet_names:
            df = pd.read_excel(filepath, sheet_name=sheet)
            df.columns = [str(c).strip() for c in df.columns]
            sheets[sheet] = df
        dataframes[filename] = sheets
        print(f"  → Loaded {filename} into pandas ({len(sheets)} sheet(s))")
    except Exception as e:
        print(f"  → Could not load {filename}: {e}")

# ─────────────────────────────────────────────
# SMART REPORT ENGINE — fixed column detection
# ─────────────────────────────────────────────
def find_column(df, candidates):
    """
    Find the best matching column name in a DataFrame.
    Tries exact match first, then case-insensitive partial match.
    Returns the actual column name or None.
    """
    cols_lower = {c.lower(): c for c in df.columns}
    for candidate in candidates:
        c = candidate.lower()
        # Exact match
        if c in cols_lower:
            return cols_lower[c]
        # Partial match
        for col_lower, col_actual in cols_lower.items():
            if c in col_lower or col_lower in c:
                return col_actual
    return None

def detect_report_intent(request, df):
    """
    Analyze the user's report request and return intent dict.
    Now uses the actual DataFrame columns for accurate matching.
    """
    r = request.lower()
    intent = {
        "group_by":       None,
        "sort_col":       None,
        "sort_asc":       False,
        "filter_col":     None,
        "filter_val":     None,
        "include_totals": True,
        "title":          "Business Intelligence Report"
    }

    # ── Detect group by ─────────────────────────
    group_candidates = {
        "region":    ["Region", "Area", "Territory"],
        "sales rep": ["Sales Rep", "Rep", "Salesperson", "Account Manager"],
        "supplier":  ["Supplier", "Vendor", "Distributor"],
        "customer":  ["Customer Name", "Customer", "Client"],
        "category":  ["Category", "Type", "Product Type"],
        "month":     ["Month", "Period", "Date"],
        "status":    ["Status", "State"],
        "product":   ["Product Name", "Product", "Description", "SKU"],
        "sku":       ["SKU", "Part Number", "Item Code"],
    }
    for keyword, candidates in group_candidates.items():
        if keyword in r:
            col = find_column(df, candidates)
            if col:
                intent["group_by"] = col
                break

    # ── Detect sort column ───────────────────────
    sort_candidates = {
        "revenue":        ["Revenue", "Sales", "Total Revenue"],
        "stock":          ["On Hand", "Stock", "Quantity", "Qty"],
        "stock value":    ["Stock Value", "Value", "Total Value"],
        "balance":        ["Outstanding Balance", "Balance"],
        "ytd":            ["YTD Revenue", "YTD", "Revenue"],
        "cost":           ["Total Cost", "Cost", "PO Total"],
        "price":          ["Unit Price", "Price"],
        "variance":       ["Variance"],
        "units":          ["Units Sold", "Quantity", "Units"],
        "reorder":        ["Reorder Qty", "Reorder Point"],
    }
    for keyword, candidates in sort_candidates.items():
        if keyword in r:
            col = find_column(df, candidates)
            if col:
                intent["sort_col"] = col
                break

    # If no sort column specified, find best numeric column
    if not intent["sort_col"]:
        numeric_cols = df.select_dtypes(include="number").columns.tolist()
        priority = ["Revenue", "Total", "Value", "Amount", "Cost", "On Hand", "Balance"]
        for p in priority:
            col = find_column(df, [p])
            if col and col in numeric_cols:
                intent["sort_col"] = col
                break
        if not intent["sort_col"] and numeric_cols:
            intent["sort_col"] = numeric_cols[0]

    # ── Detect sort direction ───────────────────
    if any(w in r for w in ["lowest to highest", "ascending", "smallest", "least to most"]):
        intent["sort_asc"] = True

    # ── Detect filters ──────────────────────────
    filter_map = {
        "active only":    ("Status", "Active"),
        "on hold":        ("Status", "On Hold"),
        "pending only":   ("Status", "Pending"),
        "received only":  ("Status", "Received"),
        "partial only":   ("Status", "Partial"),
        "reorder only":   ("Status", "Reorder"),
        "status ok":      ("Status", "OK"),
        "north only":     ("Region", "North"),
        "south only":     ("Region", "South"),
        "east only":      ("Region", "East"),
        "west only":      ("Region", "West"),
        "valves only":    ("Category", "Valves"),
        "fittings only":  ("Category", "Fittings"),
        "gauges only":    ("Category", "Gauges"),
    }
    for keyword, (col_hint, val) in filter_map.items():
        if keyword in r:
            col = find_column(df, [col_hint])
            if col:
                intent["filter_col"] = col
                intent["filter_val"] = val
                break

    # ── Build title ──────────────────────────────
    parts = []
    if intent["group_by"]:   parts.append(f"By {intent['group_by']}")
    if intent["filter_val"]: parts.append(f"— {intent['filter_val']} Only")
    if intent["sort_col"]:   parts.append(f"Sorted by {intent['sort_col']}")
    direction = "Highest to Lowest" if not intent["sort_asc"] else "Lowest to Highest"
    parts.append(f"({direction})")
    intent["title"] = "Business Intelligence Report — " + " ".join(parts)

    return intent

def find_best_dataframe(request):
    """Find the most relevant DataFrame based on keywords in the request."""
    r = request.lower()
    dataset_hints = {
        "monthly_sales":         ["sales", "revenue", "monthly", "rep", "region", "variance", "target", "units sold"],
        "inventory_status":      ["inventory", "stock", "reorder", "on hand", "sku", "category"],
        "customer_accounts":     ["customer", "account", "balance", "credit", "outstanding", "ytd"],
        "open_purchase_orders":  ["purchase", "po", "supplier", "order", "received"],
        "quotes":                ["quote", "quoting"],
        "sales_orders":          ["sales order", "so-"],
        "pricing":               ["price", "pricing", "unit price", "margin"],
    }

    best_file = None
    best_score = 0

    for filename, sheets in dataframes.items():
        fn_lower = filename.lower().replace(" ", "_").replace("-", "_")
        score = 0
        for hint_key, keywords in dataset_hints.items():
            if hint_key in fn_lower:
                score += sum(2 for k in keywords if k in r)
            score += sum(1 for k in keywords if k in r and k in fn_lower)
        if score > best_score:
            best_score = score
            best_file  = filename

    # Fallback to first file
    if not best_file and dataframes:
        best_file = list(dataframes.keys())[0]

    if best_file:
        sheet_name = list(dataframes[best_file].keys())[0]
        return best_file, sheet_name, dataframes[best_file][sheet_name].copy()

    return None, None, None

def build_report_dataframe(df, intent):
    """Apply filtering, grouping and sorting to the DataFrame."""

    # ── Apply filter ────────────────────────────
    if intent["filter_col"] and intent["filter_val"]:
        fc = intent["filter_col"]
        if fc in df.columns:
            df = df[df[fc].astype(str).str.lower() == intent["filter_val"].lower()]

    # ── Apply grouping ──────────────────────────
    numeric_cols = df.select_dtypes(include="number").columns.tolist()
    if intent["group_by"] and intent["group_by"] in df.columns:
        gc = intent["group_by"]
        if numeric_cols:
            df = df.groupby(gc)[numeric_cols].sum().reset_index()
            numeric_cols = df.select_dtypes(include="number").columns.tolist()

    # ── Apply sorting ───────────────────────────
    sort_col = intent.get("sort_col")
    if sort_col and sort_col in df.columns:
        df = df.sort_values(by=sort_col, ascending=intent["sort_asc"])
    elif numeric_cols:
        df = df.sort_values(by=numeric_cols[0], ascending=intent["sort_asc"])

    return df.reset_index(drop=True)

def generate_excel_report(request, filepath):
    """Main report generator."""
    filename, sheet_name, df = find_best_dataframe(request)

    if df is None:
        return "No data available to generate report.", False

    intent = detect_report_intent(request, df)
    df     = build_report_dataframe(df, intent)

    if df.empty:
        return "No matching data found for the requested filter.", False

    # ── Build Excel ──────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    NAVY     = "0d1b2a"
    SILVER   = "c9d1d9"
    WHITE    = "ffffff"
    LIGHT    = "f6f8fa"
    BLUE     = "1f6feb"
    TOTAL_BG = "ddf4ff"

    def thin_border():
        s = Side(style="thin", color="d0d7de")
        return Border(left=s, right=s, top=s, bottom=s)

    def thick_border():
        s = Side(style="medium", color="0d1b2a")
        return Border(left=s, right=s, top=s, bottom=s)

    def is_currency(col):
        return any(w in col.lower() for w in [
            "price", "cost", "amount", "revenue", "sales",
            "value", "total", "balance", "limit", "variance",
            "quote", "ytd", "target"
        ])

    num_cols = len(df.columns)
    last_col = get_column_letter(num_cols)

    # ── Banner ──────────────────────────────────
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = COMPANY_NAME
    ws["A1"].font      = Font(name="Calibri", bold=True, size=16, color=WHITE)
    ws["A1"].fill      = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = intent["title"]
    ws["A2"].font      = Font(name="Calibri", bold=True, size=12, color=WHITE)
    ws["A2"].fill      = PatternFill("solid", fgColor=NAVY)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"] = f"Generated: {datetime.now().strftime('%B %d, %Y  %H:%M')}    Source: {filename}"
    ws["A3"].font      = Font(name="Calibri", size=9, italic=True, color=SILVER)
    ws["A3"].fill      = PatternFill("solid", fgColor=NAVY)
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 6

    # ── Column Headers ───────────────────────────
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=5, column=col_idx, value=col_name)
        cell.font      = Font(name="Calibri", bold=True, size=11, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
    ws.row_dimensions[5].height = 22

    # ── Data Rows ────────────────────────────────
    numeric_col_names = df.select_dtypes(include="number").columns.tolist()
    last_data_row = 5

    for row_idx, (_, row) in enumerate(df.iterrows(), 6):
        bg = LIGHT if row_idx % 2 == 0 else WHITE
        for col_idx, (col_name, val) in enumerate(row.items(), 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_name in numeric_col_names and pd.notna(val):
                cell.value = val
                if is_currency(col_name):
                    cell.number_format = '$#,##0.00'
                else:
                    cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.value     = val
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.font   = Font(name="Calibri", size=10)
            cell.fill   = PatternFill("solid", fgColor=bg)
            cell.border = thin_border()
        ws.row_dimensions[row_idx].height = 18
        last_data_row = row_idx

    # ── Totals Row ───────────────────────────────
    if intent["include_totals"] and numeric_col_names:
        total_row = last_data_row + 1
        ws.row_dimensions[total_row].height = 22
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=total_row, column=col_idx)
            if col_name in numeric_col_names:
                col_total = df[col_name].sum()
                cell.value = col_total
                if is_currency(col_name):
                    cell.number_format = '$#,##0.00'
                else:
                    cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.value     = "TOTAL" if col_idx == 1 else ""
                cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font   = Font(name="Calibri", bold=True, size=11, color=NAVY)
            cell.fill   = PatternFill("solid", fgColor=TOTAL_BG)
            cell.border = thick_border()
    else:
        total_row = last_data_row

    # ── Footer ───────────────────────────────────
    footer_row = total_row + 2
    ws.merge_cells(f"A{footer_row}:{last_col}{footer_row}")
    fc = ws.cell(row=footer_row, column=1,
                 value=f"🔒  Confidential  •  {COMPANY_NAME}  •  {datetime.now().strftime('%Y-%m-%d')}  •  Air-gapped Report")
    fc.font      = Font(name="Calibri", size=9, italic=True, color=SILVER)
    fc.fill      = PatternFill("solid", fgColor=NAVY)
    fc.alignment = Alignment(horizontal="center")
    ws.row_dimensions[footer_row].height = 16

    # ── Auto column widths ───────────────────────
    for col_idx, col_name in enumerate(df.columns, 1):
        try:
            max_len = max(
                len(str(col_name)),
                df[col_name].astype(str).str.len().max() if not df.empty else 0
            )
        except Exception:
            max_len = len(str(col_name))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 4, 12), 40)

    wb.save(filepath)

    # ── Summary for chat ─────────────────────────
    rows     = len(df)
    sort_dir = "highest to lowest" if not intent["sort_asc"] else "lowest to highest"
    summary  = f"✅  Report generated\n• Rows: {rows:,}\n• Sorted by: {intent['sort_col']} ({sort_dir})"
    if intent["filter_val"]:
        summary += f"\n• Filter: {intent['filter_val']} only"
    if intent["group_by"]:
        summary += f"\n• Grouped by: {intent['group_by']}"

    return summary, True


# ─────────────────────────────────────────────
# FILE READERS
# ─────────────────────────────────────────────
def read_excel_text(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    all_text = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_lines = [f"Sheet: {sheet_name}"]
        for row in ws.iter_rows(values_only=True):
            cleaned = [str(cell).strip() if cell is not None else "" for cell in row]
            if any(cleaned):
                sheet_lines.append(" | ".join(cleaned))
        if len(sheet_lines) > 1:
            all_text.append("\n".join(sheet_lines))
    wb.close()
    return "\n\n".join(all_text)

def read_pdf(filepath):
    doc = fitz.open(filepath)
    all_text = []
    for page_num in range(len(doc)):
        page = doc[page_num]
        text = page.get_text().strip()
        if text:
            all_text.append(f"[Page {page_num + 1}]\n{text}")
    doc.close()
    return "\n\n".join(all_text)

def chunk_text(text, chunk_size=500, overlap=100):
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunks.append(text[start:end].strip())
        start += chunk_size - overlap
    return [c for c in chunks if len(c) > 50]

# ─────────────────────────────────────────────
# RAG BACKEND
# ─────────────────────────────────────────────
rag_model   = None
collection  = None
known_files = []

def load_backend():
    global rag_model, collection, known_files
    rag_model  = SentenceTransformer("all-MiniLM-L6-v2")
    client     = chromadb.PersistentClient(path=DB_PATH)
    collection = client.get_or_create_collection(name=COLLECTION_NAME)

    existing_ids = set(collection.get()["ids"])
    new_chunks, new_embeddings, new_ids, new_metadata = [], [], [], []

    for root, dirs, files in os.walk(BASE_PATH):
        dirs[:] = [d for d in dirs if "rag_db" not in d]
        for file in files:
            filepath = os.path.join(root, file)
            if file.endswith(".txt"):
                try:
                    with open(filepath, "r", encoding="utf-8") as f:
                        content = f.read()
                    file_type = "txt"
                except Exception as e:
                    print(f"Error: {e}"); continue
            elif file.endswith(".xlsx"):
                try:
                    content = read_excel_text(filepath)
                    file_type = "xlsx"
                    load_excel_as_dataframe(filepath)
                except Exception as e:
                    print(f"Error: {e}"); continue
            elif file.endswith(".pdf"):
                try:
                    content = read_pdf(filepath)
                    file_type = "pdf"
                    if not content.strip(): continue
                except Exception as e:
                    print(f"Error: {e}"); continue
            else:
                continue

            chunks = chunk_text(content)
            for i, chunk in enumerate(chunks):
                chunk_id = f"{filepath}::chunk{i}"
                if chunk_id in existing_ids: continue
                new_chunks.append(chunk)
                new_ids.append(chunk_id)
                new_metadata.append({
                    "source": filepath, "chunk_index": i,
                    "filename": file, "file_type": file_type
                })

    if new_chunks:
        new_embeddings = rag_model.encode(new_chunks).tolist()
        collection.add(documents=new_chunks, embeddings=new_embeddings,
                       ids=new_ids, metadatas=new_metadata)

    all_meta = collection.get(include=["metadatas"])["metadatas"]
    seen = set()
    for m in all_meta:
        src = m.get("source", "")
        if src and src not in seen:
            seen.add(src)
            known_files.append({
                "source": src,
                "filename": m.get("filename", os.path.basename(src)),
                "file_type": m.get("file_type", "txt")
            })

def retrieve(question):
    query_embedding = rag_model.encode([question]).tolist()
    results = collection.query(
        query_embeddings=query_embedding, n_results=TOP_K,
        include=["documents", "metadatas", "distances"]
    )
    filtered = []
    for doc, meta, dist in zip(results["documents"][0],
                               results["metadatas"][0],
                               results["distances"][0]):
        if dist <= MIN_SCORE:
            filtered.append({
                "text": doc.strip(),
                "source": meta.get("source", "Unknown"),
                "filename": meta.get("filename", "Unknown"),
                "file_type": meta.get("file_type", "txt")
            })
    return filtered

def get_calculations(question):
    results = []
    for filename, sheets in dataframes.items():
        for sheet_name, df in sheets.items():
            q = question.lower()
            numeric_cols = df.select_dtypes(include="number").columns.tolist()
            if not numeric_cols: continue
            result_lines = []

            def is_currency(col):
                return any(w in col.lower() for w in ["price","cost","amount","revenue","sales","value","total","balance","limit","variance","quote","ytd","target"])
            def fmt(val, col): return f"${val:,.2f}" if is_currency(col) else f"{val:,.2f}"
            def get_id(df, idx):
                id_cols = [c for c in df.columns if any(w in c.lower() for w in ["name","sku","id","product","item","description","number","customer","rep","month"])]
                return f" ({df[id_cols[0]].iloc[idx]})" if id_cols else ""

            try:
                if any(w in q for w in ["total","sum","how much"]):
                    for col in numeric_cols:
                        total = df[col].sum()
                        if total != 0: result_lines.append(f"{col}: {fmt(total, col)}")
                if any(w in q for w in ["average","avg","mean"]):
                    for col in numeric_cols:
                        result_lines.append(f"Avg {col}: {fmt(df[col].mean(), col)}")
                if any(w in q for w in ["highest","maximum","max","most","top"]):
                    for col in numeric_cols:
                        idx = df[col].idxmax()
                        result_lines.append(f"Max {col}: {fmt(df[col].max(), col)}{get_id(df, idx)}")
                if any(w in q for w in ["lowest","minimum","min","least"]):
                    for col in numeric_cols:
                        idx = df[col].idxmin()
                        result_lines.append(f"Min {col}: {fmt(df[col].min(), col)}{get_id(df, idx)}")
                if any(w in q for w in ["how many","count","number of"]):
                    result_lines.append(f"Count: {len(df):,}")
            except Exception as e:
                print(f"Calc error: {e}")

            if result_lines:
                results.append(f"[{filename} — {sheet_name}]\n" + "\n".join(result_lines))
    return results

def lookup_full_row(question, history):
    """
    When a follow-up question references a specific record found in history
    (e.g. 'that quote', 'that customer'), find and return the full row data
    from the matching pandas DataFrame so all columns are available.
    """
    if not history:
        return ""

    # Find the last Libby answer that contained a dollar amount or identifier
    last_answer = ""
    last_user_q = ""
    for role, msg in reversed(history):
        if role == "Libby" and not last_answer:
            last_answer = msg
        if role == "User" and not last_user_q:
            last_user_q = msg
        if last_answer and last_user_q:
            break

    if not last_answer or not last_user_q:
        return ""

    # Build a combined search term from both previous question and answer
    search_terms = (last_user_q + " " + last_answer).lower()

    # Search all dataframes for a row that matches
    result_lines = []
    for filename, sheets in dataframes.items():
        for sheet_name, df in sheets.items():
            for _, row in df.iterrows():
                row_str = " ".join([str(v) for v in row.values]).lower()
                # Check if any significant part of previous answer appears in this row
                # Extract numbers/values from last answer to match against
                import re
                amounts = re.findall(r'\$?[\d,]+\.?\d*', last_answer)
                matches = sum(1 for a in amounts if a.replace('$','').replace(',','') in row_str.replace(',',''))
                if matches > 0:
                    result_lines.append(f"\n[Full record from {filename} — {sheet_name}]")
                    for col, val in row.items():
                        result_lines.append(f"{col}: {val}")
                    break  # Found matching row in this sheet
            if result_lines:
                break

    return "\n".join(result_lines)

def ask_ollama(question, chunks, calc_results, history=None):
    if not chunks and not calc_results:
        return "Not found.", []
    context_parts = []
    if calc_results:
        context_parts.append("=== CALCULATED DATA ===")
        context_parts.extend(calc_results)
    if chunks:
        context_parts.append("=== SOURCE DOCUMENTS ===")
        for c in chunks:
            context_parts.append(f"[{c['filename']}]\n{c['text']}")

    # For follow-up questions inject full matching row so all columns are available
    vague_words = ["that", "it", "those", "the same", "its", "their", "this", "them"]
    if any(w in question.lower().split() for w in vague_words) and history:
        full_row = lookup_full_row(question, history)
        if full_row:
            context_parts.append("=== FULL RECORD (follow-up context) ===")
            context_parts.append(full_row)

    context = "\n\n".join(context_parts)
    sources = list(set([c["source"] for c in chunks])) if chunks else []

    # Build conversation history so Libby can connect follow-up questions
    history_block = ""
    if history:
        history_block = "\n--- RECENT CONVERSATION ---\n"
        for role, msg in history[-6:]:   # Last 3 exchanges
            history_block += f"{role}: {msg}\n"
        history_block += "--- END CONVERSATION ---\n"

    prompt = f"""{SYSTEM_PROMPT}
{history_block}
--- DATA ---
{context}
--- END ---

Question: {question}
Answer:"""
    try:
        response = requests.post(
            OLLAMA_URL,
            json={"model": OLLAMA_MODEL, "prompt": prompt, "stream": False,
                  "options": {"temperature": 0.0, "top_p": 0.9, "num_predict": 150}},
            timeout=120
        )
        response.raise_for_status()
        raw = response.json().get("response", "").strip()
        # Strip any markdown code blocks Mistral sneaks in
        raw = raw.replace("```plaintext", "").replace("```python", "").replace("```", "").strip()
        return raw, sources
    except requests.exceptions.ConnectionError:
        return "⚠️  Ollama not running. Run: ollama serve", []
    except Exception as e:
        return f"⚠️  Error: {str(e)}", []


# ─────────────────────────────────────────────
# MAIN APP
# ─────────────────────────────────────────────
class LibbyApp:
    def __init__(self, root):
        self.root          = root
        self.ready         = False
        self.current_theme = "light"
        self.T             = THEMES["light"]
        self.dot_count     = 0
        self.animating     = False
        self._anim_msg     = "⚡  Calculating"
        self._chat_widgets = []
        self._history      = []   # Stores (role, message) tuples for context

        self.root.title("Libby V6 — Enterprise Intelligence + Smart Reports")
        self.root.geometry("1200x800")
        self.root.minsize(800, 600)
        self.root.configure(bg=self.T["BG_DARK"])

        self._build_ui()
        self._start_backend()

    def _build_ui(self):
        T = self.T

        # ── Header ──────────────────────────
        self.header = tk.Frame(self.root, bg=T["BG_PANEL"], pady=14)
        self.header.pack(fill="x")

        left_header = tk.Frame(self.header, bg=T["BG_PANEL"])
        left_header.pack(side="left", padx=16)

        self.avatar_canvas = tk.Canvas(
            left_header, width=52, height=52,
            bg=T["BG_PANEL"], highlightthickness=0
        )
        self.avatar_canvas.pack(side="left", padx=(0, 12))
        self._draw_avatar()

        self.title_frame = tk.Frame(left_header, bg=T["BG_PANEL"])
        self.title_frame.pack(side="left")

        self.title_label = tk.Label(
            self.title_frame, text="LIBBY  V6",
            bg=T["BG_PANEL"], fg=T["ACCENT"],
            font=("Courier New", 22, "bold")
        )
        self.title_label.pack(anchor="w")

        self.subtitle_label = tk.Label(
            self.title_frame,
            text="Enterprise Intelligence  •  Smart Report Generation  •  Offline",
            bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
            font=("Courier New", 9)
        )
        self.subtitle_label.pack(anchor="w")

        self.right_frame = tk.Frame(self.header, bg=T["BG_PANEL"])
        self.right_frame.pack(side="right", padx=16)

        self.theme_btn = tk.Button(
            self.right_frame, text="🌙  Dark Mode",
            bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
            activebackground=T["BG_INPUT"],
            font=("Courier New", 9), relief="flat",
            cursor="hand2", command=self._toggle_theme
        )
        self.theme_btn.pack(anchor="e", pady=(0, 4))

        self.status_label = tk.Label(
            self.right_frame, text="⏳  Loading...",
            bg=T["BG_PANEL"], fg=T["WARNING"],
            font=("Courier New", 9)
        )
        self.status_label.pack(anchor="e")

        self.top_divider = tk.Frame(self.root, bg=T["BORDER"], height=1)
        self.top_divider.pack(fill="x")

        # ── Body ────────────────────────────
        self.body = tk.Frame(self.root, bg=T["BG_DARK"])
        self.body.pack(fill="both", expand=True)

        # ── Sidebar ─────────────────────────
        self.sidebar = tk.Frame(self.body, bg=T["BG_SIDEBAR"], width=240)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)
        self._build_sidebar_contents()

        self.side_divider = tk.Frame(self.body, bg=T["BORDER"], width=1)
        self.side_divider.pack(side="left", fill="y")

        # ── Chat ────────────────────────────
        self.chat_outer = tk.Frame(self.body, bg=T["BG_PANEL"])
        self.chat_outer.pack(side="left", fill="both", expand=True)

        self.chat_canvas = tk.Canvas(
            self.chat_outer, bg=T["BG_PANEL"],
            highlightthickness=0, bd=0
        )
        scrollbar = tk.Scrollbar(self.chat_outer, command=self.chat_canvas.yview)
        self.chat_canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.chat_canvas.pack(side="left", fill="both", expand=True)

        self.chat_inner = tk.Frame(self.chat_canvas, bg=T["BG_PANEL"])
        self.chat_window = self.chat_canvas.create_window(
            (0, 0), window=self.chat_inner, anchor="nw"
        )
        self.chat_inner.bind("<Configure>", self._on_frame_configure)
        self.chat_canvas.bind("<Configure>", self._on_canvas_configure)

        # Only scroll chat when mouse is over chat area
        self.chat_canvas.bind("<Enter>", lambda e: self.chat_canvas.bind_all("<MouseWheel>", self._on_mousewheel))
        self.chat_canvas.bind("<Leave>", lambda e: self.chat_canvas.unbind_all("<MouseWheel>"))

        self.bottom_divider = tk.Frame(self.root, bg=T["BORDER"], height=1)
        self.bottom_divider.pack(fill="x")

        # ── Input ────────────────────────────
        self.input_frame = tk.Frame(self.root, bg=T["BG_PANEL"], pady=10)
        self.input_frame.pack(fill="x", padx=12)

        self.input_box = tk.Text(
            self.input_frame, height=3,
            bg=T["BG_INPUT"], fg=T["TEXT_MAIN"],
            insertbackground=T["ACCENT"],
            font=("Courier New", 11),
            relief="flat", bd=0,
            padx=12, pady=10, wrap="word"
        )
        self.input_box.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self.input_box.bind("<Return>", self._on_enter)
        self.input_box.insert("1.0", "Ask a question or type 'generate report ...'")
        self.input_box.config(fg=T["TEXT_DIM"])
        self.input_box.bind("<FocusIn>", self._clear_placeholder)
        self.input_box.bind("<FocusOut>", self._restore_placeholder)

        self.btn_frame = tk.Frame(self.input_frame, bg=T["BG_PANEL"])
        self.btn_frame.pack(side="right")

        self.send_btn = tk.Button(
            self.btn_frame, text="QUERY",
            bg=T["ACCENT_DIM"], fg=T["BG_PANEL"],
            activebackground=T["ACCENT"],
            font=("Courier New", 11, "bold"),
            relief="flat", padx=16, pady=4,
            cursor="hand2", command=self._send_question
        )
        self.send_btn.pack(pady=(0, 4))

        self.clear_btn = tk.Button(
            self.btn_frame, text="🗑  Clear",
            bg=T["BTN_CLEAR"], fg=T["BTN_CLEAR_FG"],
            activebackground=T["BTN_CLEAR"],
            font=("Courier New", 9),
            relief="flat", padx=10, pady=3,
            cursor="hand2", command=self._clear_chat
        )
        self.clear_btn.pack()

        # ── Footer ───────────────────────────
        self.footer = tk.Frame(self.root, bg=T["BG_PANEL"], pady=5)
        self.footer.pack(fill="x")
        self.footer_label = tk.Label(
            self.footer,
            text="🔒  Air-gapped  •  Type 'generate report ...' to export a formatted Excel report",
            bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
            font=("Courier New", 8)
        )
        self.footer_label.pack()

        self._add_libby_message(
            "LIBBY V6 — Smart Report Generation.\n\n"
            "Ask business questions as normal, or type:\n"
            "'generate report showing sales by region highest to lowest'\n"
            "'generate report of inventory items lowest to highest stock'\n"
            "'generate report of active customers by outstanding balance'\n"
            "'generate report of pending purchase orders by supplier'\n\n"
            "Libby will build and export a formatted Excel report automatically.",
            source=None, file_type=None, is_calc=False
        )

    # ─────────────────────────────────────────
    # SIDEBAR
    # ─────────────────────────────────────────
    def _build_sidebar_contents(self):
        T = self.T
        self.sidebar_title = tk.Label(
            self.sidebar, text="📁  DATA SOURCES",
            bg=T["BG_SIDEBAR"], fg=T["ACCENT"],
            font=("Courier New", 9, "bold"), pady=12
        )
        self.sidebar_title.pack(anchor="w", padx=12)
        tk.Frame(self.sidebar, bg=T["BORDER"], height=1).pack(fill="x", padx=8)

        self.pandas_label = tk.Label(
            self.sidebar, text="📊  CALCULATION ENGINE",
            bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"],
            font=("Courier New", 7, "bold"), pady=6
        )
        self.pandas_label.pack(anchor="w", padx=12)

        self.pandas_status = tk.Label(
            self.sidebar, text="⏳  Loading...",
            bg=T["BG_SIDEBAR"], fg=T["WARNING"],
            font=("Courier New", 7)
        )
        self.pandas_status.pack(anchor="w", padx=16)

        tk.Frame(self.sidebar, bg=T["BORDER"], height=1).pack(fill="x", padx=8, pady=4)

        # ── Scrollable file list ─────────────────
        list_container = tk.Frame(self.sidebar, bg=T["BG_SIDEBAR"])
        list_container.pack(fill="both", expand=True, padx=0, pady=4)

        self.sidebar_canvas = tk.Canvas(
            list_container, bg=T["BG_SIDEBAR"],
            highlightthickness=0, bd=0
        )
        sidebar_scrollbar = tk.Scrollbar(
            list_container, orient="vertical",
            command=self.sidebar_canvas.yview
        )
        self.sidebar_canvas.configure(yscrollcommand=sidebar_scrollbar.set)
        sidebar_scrollbar.pack(side="right", fill="y")
        self.sidebar_canvas.pack(side="left", fill="both", expand=True)

        self.sidebar_list = tk.Frame(self.sidebar_canvas, bg=T["BG_SIDEBAR"])
        self.sidebar_list_window = self.sidebar_canvas.create_window(
            (0, 0), window=self.sidebar_list, anchor="nw"
        )

        # Update scroll region when content changes
        self.sidebar_list.bind("<Configure>", lambda e: self.sidebar_canvas.configure(
            scrollregion=self.sidebar_canvas.bbox("all")
        ))
        self.sidebar_canvas.bind("<Configure>", lambda e: self.sidebar_canvas.itemconfig(
            self.sidebar_list_window, width=e.width
        ))

        # Mouse wheel scrolling — only when hovering over sidebar
        self.sidebar_canvas.bind("<Enter>", lambda e: self.sidebar_canvas.bind_all(
            "<MouseWheel>", lambda ev: self.sidebar_canvas.yview_scroll(
                int(-1 * (ev.delta / 120)), "units"
            )
        ))
        self.sidebar_canvas.bind("<Leave>", lambda e: self.sidebar_canvas.unbind_all("<MouseWheel>"))
        self.sidebar_list.bind("<Enter>", lambda e: self.sidebar_canvas.bind_all(
            "<MouseWheel>", lambda ev: self.sidebar_canvas.yview_scroll(
                int(-1 * (ev.delta / 120)), "units"
            )
        ))
        self.sidebar_list.bind("<Leave>", lambda e: self.sidebar_canvas.unbind_all("<MouseWheel>"))

        tk.Frame(self.sidebar, bg=T["BORDER"], height=1).pack(fill="x", padx=8, pady=4)
        tk.Label(
            self.sidebar, text="💡  REPORT TRIGGER",
            bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"],
            font=("Courier New", 7, "bold"), pady=4
        ).pack(anchor="w", padx=12)
        tk.Label(
            self.sidebar,
            text="Start with:\n\"generate report...\"\nto export Excel.",
            bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"],
            font=("Courier New", 7), justify="left"
        ).pack(anchor="w", padx=16, pady=(0, 8))

        self.sidebar_footer_label = tk.Label(
            self.sidebar,
            text="📄 TXT   📊 XLSX ⚡   📕 PDF",
            bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"],
            font=("Courier New", 7), pady=8
        )
        self.sidebar_footer_label.pack(side="bottom")

    def _refresh_sidebar(self):
        T = self.T
        for widget in self.sidebar_list.winfo_children():
            widget.destroy()
        xlsx_count = len(dataframes)
        self.pandas_status.config(
            text=f"✅  {xlsx_count} file(s) calc-ready",
            fg=T["SUCCESS"]
        )
        for entry in known_files:
            filename  = entry["filename"]
            folder    = os.path.basename(os.path.dirname(entry["source"]))
            file_type = entry.get("file_type", "txt")
            icon  = "📊" if file_type == "xlsx" else ("📕" if file_type == "pdf" else "📄")
            badge = " ⚡" if file_type == "xlsx" else ""
            row = tk.Frame(self.sidebar_list, bg=T["BG_SIDEBAR"])
            row.pack(anchor="w", fill="x", pady=2, padx=8)

            lbl1 = tk.Label(row, text=f"{icon} {filename}{badge}",
                     bg=T["BG_SIDEBAR"], fg=T["TEXT_MAIN"],
                     font=("Courier New", 8), wraplength=190,
                     justify="left", anchor="w")
            lbl1.pack(anchor="w")

            lbl2 = tk.Label(row, text=f"   📁 {folder}",
                     bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"],
                     font=("Courier New", 7), anchor="w")
            lbl2.pack(anchor="w", pady=(0, 4))

            # Bind scroll to labels and row so entire area scrolls on hover
            for widget in [row, lbl1, lbl2]:
                widget.bind("<Enter>", lambda e: self.sidebar_canvas.bind_all(
                    "<MouseWheel>", lambda ev: self.sidebar_canvas.yview_scroll(
                        int(-1 * (ev.delta / 120)), "units"
                    )
                ))
                widget.bind("<Leave>", lambda e: self.sidebar_canvas.unbind_all("<MouseWheel>"))

    # ─────────────────────────────────────────
    # AVATAR
    # ─────────────────────────────────────────
    def _draw_avatar(self):
        c = self.avatar_canvas
        T = self.T
        c.delete("all")
        c.create_rectangle(4, 4, 48, 48, fill=T["BG_PANEL"], outline=T["ACCENT"], width=2)
        c.create_rectangle(8, 8, 44, 26, fill=T["ACCENT_DIM"], outline="")
        c.create_rectangle(10, 32, 17, 44, fill=T["ACCENT"], outline="")
        c.create_rectangle(20, 26, 27, 44, fill=T["ACCENT"], outline="")
        c.create_rectangle(30, 30, 37, 44, fill=T["ACCENT"], outline="")
        c.create_rectangle(40, 22, 44, 44, fill=T["SUCCESS"], outline="")
        tk.Label(c, text="V6", bg=T["BG_PANEL"], fg=T["TEXT_MAIN"],
                 font=("Courier New", 7, "bold")).place(x=11, y=10)

    # ─────────────────────────────────────────
    # THEME
    # ─────────────────────────────────────────
    def _toggle_theme(self):
        self.current_theme = "light" if self.current_theme == "dark" else "dark"
        self.T = THEMES[self.current_theme]
        self.theme_btn.config(
            text="☀️  Light Mode" if self.current_theme == "dark" else "🌙  Dark Mode"
        )
        self._apply_theme()

    def _apply_theme(self):
        T = self.T
        self.root.configure(bg=T["BG_DARK"])
        self.header.configure(bg=T["BG_PANEL"])
        self.avatar_canvas.configure(bg=T["BG_PANEL"])
        self._draw_avatar()
        self.title_frame.configure(bg=T["BG_PANEL"])
        self.title_label.configure(bg=T["BG_PANEL"], fg=T["ACCENT"])
        self.subtitle_label.configure(bg=T["BG_PANEL"], fg=T["TEXT_DIM"])
        self.right_frame.configure(bg=T["BG_PANEL"])
        self.theme_btn.configure(bg=T["BG_PANEL"], fg=T["TEXT_DIM"])
        self.status_label.configure(bg=T["BG_PANEL"])
        self.top_divider.configure(bg=T["BORDER"])
        self.body.configure(bg=T["BG_DARK"])
        self.sidebar.configure(bg=T["BG_SIDEBAR"])
        self.sidebar_title.configure(bg=T["BG_SIDEBAR"], fg=T["ACCENT"])
        self.pandas_label.configure(bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"])
        self.pandas_status.configure(bg=T["BG_SIDEBAR"])
        self.sidebar_canvas.configure(bg=T["BG_SIDEBAR"])
        self.sidebar_list.configure(bg=T["BG_SIDEBAR"])
        self.sidebar_footer_label.configure(bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"])
        self.side_divider.configure(bg=T["BORDER"])
        self.chat_outer.configure(bg=T["BG_PANEL"])
        self.chat_canvas.configure(bg=T["BG_PANEL"])
        self.chat_inner.configure(bg=T["BG_PANEL"])
        self.bottom_divider.configure(bg=T["BORDER"])
        self.input_frame.configure(bg=T["BG_PANEL"])
        self.btn_frame.configure(bg=T["BG_PANEL"])
        self.input_box.configure(bg=T["BG_INPUT"], fg=T["TEXT_MAIN"],
                                 insertbackground=T["ACCENT"])
        self.send_btn.configure(bg=T["ACCENT_DIM"], fg=T["BG_PANEL"])
        self.clear_btn.configure(bg=T["BTN_CLEAR"], fg=T["BTN_CLEAR_FG"])
        self.footer.configure(bg=T["BG_PANEL"])
        self.footer_label.configure(bg=T["BG_PANEL"], fg=T["TEXT_DIM"])
        self.status_label.configure(
            fg=T["SUCCESS"] if self.current_theme == "light" else T["WARNING"]
        )
        self._refresh_sidebar()
        self._refresh_chat_bubbles()

    # ─────────────────────────────────────────
    # CHAT BUBBLES
    # ─────────────────────────────────────────
    def _add_user_message(self, text):
        T     = self.T
        outer = tk.Frame(self.chat_inner, bg=T["BG_PANEL"], pady=6)
        outer.pack(fill="x", padx=16)
        self._chat_widgets.append(("user", text, None, None, False))
        tk.Label(outer, text="YOU", bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
                 font=("Courier New", 8, "bold")).pack(anchor="e", padx=4)
        bubble = tk.Frame(outer, bg=T["BG_BUBBLE_U"], padx=16, pady=10)
        bubble.pack(anchor="e")
        tk.Label(bubble, text=text, bg=T["BG_BUBBLE_U"], fg=T["TEXT_BUBBLE_U"],
                 font=("Courier New", 11), wraplength=700, justify="left").pack(anchor="w")

    def _add_libby_message(self, text, source=None, file_type=None, is_calc=False):
        T     = self.T
        outer = tk.Frame(self.chat_inner, bg=T["BG_PANEL"], pady=6)
        outer.pack(fill="x", padx=16)
        self._chat_widgets.append(("libby", text, source, file_type, is_calc))

        label = "LIBBY V6  ⚡" if is_calc else "LIBBY V6"
        tk.Label(outer, text=label, bg=T["BG_PANEL"], fg=T["ACCENT"],
                 font=("Courier New", 8, "bold")).pack(anchor="w", padx=4)

        bubble_bg  = T["BG_CALC"] if is_calc else T["BG_BUBBLE_L"]
        text_color = T["TEXT_CALC"] if is_calc else T["TEXT_BUBBLE_L"]

        bubble = tk.Frame(outer, bg=bubble_bg, padx=16, pady=10)
        bubble.pack(anchor="w")
        tk.Label(bubble, text=text, bg=bubble_bg, fg=text_color,
                 font=("Courier New", 11), wraplength=700, justify="left").pack(anchor="w")

        if source:
            icon = "📊" if file_type == "xlsx" else ("📕" if file_type == "pdf" else "📄")
            tk.Label(bubble, text=f"{icon}  {source}",
                     bg=bubble_bg, fg=T["TEXT_SOURCE"],
                     font=("Courier New", 8), wraplength=700,
                     justify="left").pack(anchor="w", pady=(6, 0))

    def _refresh_chat_bubbles(self):
        saved = list(self._chat_widgets)
        for widget in self.chat_inner.winfo_children():
            widget.destroy()
        self._chat_widgets = []
        for kind, text, source, file_type, is_calc in saved:
            if kind == "user":
                self._add_user_message(text)
            else:
                self._add_libby_message(text, source, file_type, is_calc)

    # ─────────────────────────────────────────
    # THINKING ANIMATION
    # ─────────────────────────────────────────
    def _add_thinking_frame(self, is_report=False):
        T     = self.T
        frame = tk.Frame(self.chat_inner, bg=T["BG_PANEL"], pady=6)
        frame.pack(fill="x", padx=16)
        tk.Label(frame, text="LIBBY V6", bg=T["BG_PANEL"], fg=T["ACCENT"],
                 font=("Courier New", 8, "bold")).pack(anchor="w", padx=4)
        self._anim_msg = "📊  Building report" if is_report else "⚡  Calculating"
        self.thinking_label = tk.Label(
            frame, text=self._anim_msg,
            bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
            font=("Courier New", 10, "italic")
        )
        self.thinking_label.pack(anchor="w", padx=10)
        self.animating = True
        self._animate_dots()
        self._scroll_to_bottom()
        return frame

    def _animate_dots(self):
        if not self.animating:
            return
        dots = "." * (self.dot_count % 4)
        self.thinking_label.config(text=f"{self._anim_msg}{dots}")
        self.dot_count += 1
        self.root.after(300, self._animate_dots)

    def _stop_animation(self, frame):
        self.animating = False
        frame.destroy()

    # ─────────────────────────────────────────
    # CLEAR
    # ─────────────────────────────────────────
    def _clear_chat(self):
        for widget in self.chat_inner.winfo_children():
            widget.destroy()
        self._chat_widgets = []
        self._history      = []   # Reset conversation context
        self._add_libby_message(
            "Session cleared. Ready for queries.",
            source=None, file_type=None, is_calc=False
        )

    # ─────────────────────────────────────────
    # SCROLL
    # ─────────────────────────────────────────
    def _on_frame_configure(self, event):
        self.chat_canvas.configure(scrollregion=self.chat_canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.chat_canvas.itemconfig(self.chat_window, width=event.width)

    def _on_mousewheel(self, event):
        self.chat_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _scroll_to_bottom(self):
        self.root.after(50, lambda: self.chat_canvas.yview_moveto(1.0))

    # ─────────────────────────────────────────
    # INPUT
    # ─────────────────────────────────────────
    def _clear_placeholder(self, event):
        placeholder = "Ask a question or type 'generate report ...'"
        if self.input_box.get("1.0", "end-1c") == placeholder:
            self.input_box.delete("1.0", "end")
            self.input_box.config(fg=self.T["TEXT_MAIN"])

    def _restore_placeholder(self, event):
        if not self.input_box.get("1.0", "end-1c").strip():
            self.input_box.insert("1.0", "Ask a question or type 'generate report ...'")
            self.input_box.config(fg=self.T["TEXT_DIM"])

    def _on_enter(self, event):
        if not event.state & 0x1:
            self._send_question()
            return "break"

    # ─────────────────────────────────────────
    # SEND & QUERY
    # ─────────────────────────────────────────
    def _send_question(self):
        if not self.ready:
            return
        placeholder = "Ask a question or type 'generate report ...'"
        question = self.input_box.get("1.0", "end-1c").strip()
        if not question or question == placeholder:
            return

        self.input_box.delete("1.0", "end")
        self.input_box.config(fg=self.T["TEXT_MAIN"])
        self._add_user_message(question)

        is_report = "generate report" in question.lower()
        thinking_frame = self._add_thinking_frame(is_report=is_report)
        self.send_btn.config(state="disabled", bg=self.T["BG_INPUT"])
        self.input_box.config(state="disabled")

        if is_report:
            self.root.after(0, lambda: self._prompt_save_report(question, thinking_frame))
        else:
            def run_query():
                # ── Enrich question with context ─────────
                # If question uses pronouns or vague references
                # ("that", "it", "those", "the same"), prepend
                # recent history so retrieval finds the right data
                vague_words = ["that", "it", "those", "the same", "its", "their", "this", "them"]
                q_lower = question.lower()
                search_question = question
                if any(w in q_lower.split() for w in vague_words) and self._history:
                    # Build enriched query from last user question + current
                    last_user_msgs = [m for role, m in self._history if role == "User"]
                    if last_user_msgs:
                        search_question = f"{last_user_msgs[-1]} {question}"

                chunks       = retrieve(search_question)
                calc_results = get_calculations(search_question)
                answer, sources = ask_ollama(question, chunks, calc_results, self._history)
                source_str   = ", ".join(sources) if sources else None
                file_type    = chunks[0]["file_type"] if chunks else None
                is_calc      = len(calc_results) > 0
                # Update conversation history
                self._history.append(("User", question))
                self._history.append(("Libby", answer))
                # Keep history to last 10 messages to avoid prompt bloat
                if len(self._history) > 10:
                    self._history = self._history[-10:]
                self.root.after(0, lambda: self._show_answer(
                    answer, source_str, file_type, is_calc, thinking_frame
                ))
            threading.Thread(target=run_query, daemon=True).start()

    def _prompt_save_report(self, question, thinking_frame):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath  = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"Libby_Report_{timestamp}.xlsx",
            title="Save Report As"
        )
        if not filepath:
            self._stop_animation(thinking_frame)
            self._add_libby_message("Report cancelled.", source=None, file_type=None, is_calc=False)
            self._re_enable_input()
            return

        def do_generate():
            summary, success = generate_excel_report(question, filepath)
            self.root.after(0, lambda: self._show_report_result(
                summary, success, filepath, thinking_frame
            ))
        threading.Thread(target=do_generate, daemon=True).start()

    def _show_report_result(self, summary, success, filepath, thinking_frame):
        self._stop_animation(thinking_frame)
        if success:
            self._add_libby_message(
                f"{summary}\n📁  Saved: {os.path.basename(filepath)}",
                source=filepath, file_type="xlsx", is_calc=True
            )
        else:
            self._add_libby_message(summary, source=None, file_type=None, is_calc=False)
        self._scroll_to_bottom()
        self._re_enable_input()

    def _re_enable_input(self):
        self.send_btn.config(state="normal", bg=self.T["ACCENT_DIM"])
        self.input_box.config(state="normal")
        self.input_box.focus()
        self.root.update_idletasks()
        self.root.update()

    def _show_answer(self, answer, source, file_type, is_calc, thinking_frame):
        self._stop_animation(thinking_frame)
        self._add_libby_message(answer, source=source, file_type=file_type, is_calc=is_calc)
        self._scroll_to_bottom()
        self._re_enable_input()

    # ─────────────────────────────────────────
    # BACKEND
    # ─────────────────────────────────────────
    def _start_backend(self):
        def load():
            load_backend()
            self.ready = True
            count = collection.count()
            self.root.after(0, lambda: self.status_label.config(
                text=f"✅  {count} chunks  •  phi3:mini  •  Offline",
                fg=self.T["SUCCESS"]
            ))
            self.root.after(0, self._refresh_sidebar)
        threading.Thread(target=load, daemon=True).start()


# ─────────────────────────────────────────────
# LAUNCH
# ─────────────────────────────────────────────
if __name__ == "__main__":
    root = tk.Tk()
    app  = LibbyApp(root)
    root.mainloop()