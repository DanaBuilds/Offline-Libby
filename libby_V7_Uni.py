import tkinter as tk
from tkinter import filedialog, messagebox
import threading
import os
import json
import requests
import chromadb
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import fitz
import pandas as pd
from sentence_transformers import SentenceTransformer
from datetime import datetime
import re

# ─────────────────────────────────────────────
# CONFIG FILE — saves user settings locally
# ─────────────────────────────────────────────
SCRIPT_DIR      = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE     = os.path.join(SCRIPT_DIR, "libby_config.json")
DEFAULT_CONFIG  = {
    "base_path":    SCRIPT_DIR,
    "theme":        "dark",
    "mode":         "knowledge",
    "company_name": "Libby Intelligence",
    "ollama_model": "phi3:mini",
}

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r") as f:
                cfg = json.load(f)
            # Fill in any missing keys with defaults
            for k, v in DEFAULT_CONFIG.items():
                cfg.setdefault(k, v)
            return cfg
        except Exception:
            pass
    return DEFAULT_CONFIG.copy()

def save_config(cfg):
    with open(CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)

# ─────────────────────────────────────────────
# THEMES — Charcoal & Rose Gold (dark)
#          Warm Linen (light)
# ─────────────────────────────────────────────
THEMES = {
    "dark": {
        "BG_DARK":       "#1e1e1e",   # True charcoal
        "BG_PANEL":      "#252525",   # Slightly lighter charcoal
        "BG_INPUT":      "#1a1a1a",   # Near black input
        "BG_BUBBLE_U":   "#2e2420",   # Warm dark bubble (user)
        "BG_BUBBLE_L":   "#222222",   # Charcoal libby bubble
        "BG_SIDEBAR":    "#1a1a1a",   # Dark sidebar
        "BG_SETTINGS":   "#2a2a2a",   # Settings panel bg
        "ACCENT":        "#c9956c",   # Rose gold
        "ACCENT_DIM":    "#8b5e3c",   # Darker rose gold
        "ACCENT_BRIGHT": "#e8b49a",   # Light rose gold highlight
        "TEXT_MAIN":     "#f0ece8",   # Warm off-white
        "TEXT_DIM":      "#6b6560",   # Warm gray
        "TEXT_SOURCE":   "#c9956c",   # Rose gold
        "SUCCESS":       "#c9956c",   # Rose gold success
        "WARNING":       "#e8c547",   # Warm yellow warning
        "BORDER":        "#3a3530",   # Warm dark border
        "BTN_CLEAR":     "#2a1a1a",
        "BTN_CLEAR_FG":  "#e05555",
        "BG_CALC":       "#2a2018",   # Warm dark for calc bubbles
        "TEXT_CALC":     "#e8b49a",   # Light rose gold
        "TEXT_BUBBLE_U": "#f0ece8",
        "TEXT_BUBBLE_L": "#f0ece8",
        "BTN_EXPORT":    "#1a2a1a",
        "BTN_EXPORT_FG": "#6dbe6d",
        "MODE_KB":       "#1a2030",   # Knowledge mode indicator
        "MODE_EBI":      "#2a1a30",   # EBI mode indicator
    },
    "light": {
        "BG_DARK":       "#f5f0eb",   # Warm linen
        "BG_PANEL":      "#ffffff",   # White panel
        "BG_INPUT":      "#ede8e2",   # Warm input
        "BG_BUBBLE_U":   "#e8ddd4",   # Warm beige user bubble
        "BG_BUBBLE_L":   "#f5f0eb",   # Linen libby bubble
        "BG_SIDEBAR":    "#ede8e2",   # Warm sidebar
        "BG_SETTINGS":   "#f0ebe5",   # Settings panel bg
        "ACCENT":        "#9b5e3a",   # Deep rose gold
        "ACCENT_DIM":    "#7a4a2c",   # Darker rose gold
        "ACCENT_BRIGHT": "#c9956c",   # Rose gold
        "TEXT_MAIN":     "#2a2018",   # Warm dark brown
        "TEXT_DIM":      "#8a7a6a",   # Warm gray-brown
        "TEXT_SOURCE":   "#9b5e3a",   # Deep rose gold
        "SUCCESS":       "#3a7a3a",   # Green
        "WARNING":       "#8a6a00",   # Dark amber
        "BORDER":        "#d5c8bc",   # Warm border
        "BTN_CLEAR":     "#f0ddd8",
        "BTN_CLEAR_FG":  "#8a2020",
        "BG_CALC":       "#e8ddd4",
        "TEXT_CALC":     "#2a2018",
        "TEXT_BUBBLE_U": "#2a2018",
        "TEXT_BUBBLE_L": "#2a2018",
        "BTN_EXPORT":    "#d8ecd8",
        "BTN_EXPORT_FG": "#1a5a1a",
        "MODE_KB":       "#d8e8f0",
        "MODE_EBI":      "#e8d8f0",
    }
}

# ─────────────────────────────────────────────
# PROMPTS — Knowledge mode vs EBI mode
# ─────────────────────────────────────────────
PROMPT_KNOWLEDGE = """You are Libby, an offline knowledge assistant.
Answer questions using ONLY the provided context.
- Be clear, helpful and practical.
- Use bullet points "•" for lists.
- Number steps when applicable: 1. 2. 3.
- If not found in context: "Not found in knowledge base."
- Never guess or use outside knowledge."""

PROMPT_EBI = """You are Libby, an ultra-concise enterprise data retrieval system.
- Return numbers and dollar amounts ONLY. No narrative.
- Maximum 3 lines per answer.
- Show formula if calculating: e.g. $6,750 + $4,800 = $11,550
- Format currency: $1,234.56
- Format percentages: 12.5%
- If not found: "Not found."
- NEVER use markdown. Plain text only."""

# ─────────────────────────────────────────────
# PANDAS DATA STORE
# ─────────────────────────────────────────────
dataframes = {}

def load_excel_as_dataframe(filepath):
    filename = os.path.basename(filepath)
    try:
        xl = pd.ExcelFile(filepath)
        sheets = {}
        for sheet in xl.sheet_names:
            df = pd.read_excel(filepath, sheet_name=sheet)
            df.columns = [str(c).strip() for c in df.columns]
            sheets[sheet] = df
        dataframes[filename] = sheets
    except Exception as e:
        print(f"  → Could not load {filename}: {e}")

def find_column(df, candidates):
    cols_lower = {c.lower(): c for c in df.columns}
    for candidate in candidates:
        c = candidate.lower()
        if c in cols_lower:
            return cols_lower[c]
        for col_lower, col_actual in cols_lower.items():
            if c in col_lower or col_lower in c:
                return col_actual
    return None

def lookup_full_row(question, history):
    if not history:
        return ""
    last_answer = ""
    last_user_q = ""
    for role, msg in reversed(history):
        if role == "Libby" and not last_answer:
            last_answer = msg
        if role == "User" and not last_user_q:
            last_user_q = msg
        if last_answer and last_user_q:
            break
    if not last_answer:
        return ""
    result_lines = []
    amounts = re.findall(r'\$?[\d,]+\.?\d*', last_answer)
    for filename, sheets in dataframes.items():
        for sheet_name, df in sheets.items():
            for _, row in df.iterrows():
                row_str = " ".join([str(v) for v in row.values]).lower()
                matches = sum(1 for a in amounts if a.replace('$','').replace(',','') in row_str.replace(',',''))
                if matches > 0:
                    result_lines.append(f"\n[Full record from {filename} — {sheet_name}]")
                    for col, val in row.items():
                        result_lines.append(f"{col}: {val}")
                    break
            if result_lines:
                break
    return "\n".join(result_lines)

def perform_calculation(question, filename, sheet_name, df):
    q = question.lower()
    result_lines = []
    try:
        numeric_cols = df.select_dtypes(include="number").columns.tolist()
        if not numeric_cols:
            return None

        def is_currency(col):
            return any(w in col.lower() for w in ["price","cost","amount","revenue","sales","value","total","balance","limit","variance","quote","ytd","target"])
        def fmt(val, col): return f"${val:,.2f}" if is_currency(col) else f"{val:,.2f}"
        def get_id(df, idx):
            id_cols = [c for c in df.columns if any(w in c.lower() for w in ["name","sku","id","product","item","description","number","customer","rep","month"])]
            return f" ({df[id_cols[0]].iloc[idx]})" if id_cols else ""

        if any(w in q for w in ["total","sum","how much"]):
            for col in numeric_cols:
                total = df[col].sum()
                if total != 0: result_lines.append(f"{col}: {fmt(total, col)}")
        if any(w in q for w in ["average","avg","mean"]):
            for col in numeric_cols:
                result_lines.append(f"Avg {col}: {fmt(df[col].mean(), col)}")
        if any(w in q for w in ["highest","maximum","max","most","top"]):
            for col in numeric_cols:
                idx = df[col].idxmax()
                result_lines.append(f"Max {col}: {fmt(df[col].max(), col)}{get_id(df, idx)}")
        if any(w in q for w in ["lowest","minimum","min","least"]):
            for col in numeric_cols:
                idx = df[col].idxmin()
                result_lines.append(f"Min {col}: {fmt(df[col].min(), col)}{get_id(df, idx)}")
        if any(w in q for w in ["how many","count","number of"]):
            result_lines.append(f"Count: {len(df):,}")
    except Exception as e:
        print(f"Calc error: {e}")
    if result_lines:
        return f"[{filename} — {sheet_name}]\n" + "\n".join(result_lines)
    return None

def get_calculations(question):
    results = []
    for filename, sheets in dataframes.items():
        for sheet_name, df in sheets.items():
            calc = perform_calculation(question, filename, sheet_name, df)
            if calc:
                results.append(calc)
    return results

# ─────────────────────────────────────────────
# FILE READERS
# ─────────────────────────────────────────────
def read_excel_text(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    all_text = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_lines = [f"Sheet: {sheet_name}"]
        for row in ws.iter_rows(values_only=True):
            cleaned = [str(cell).strip() if cell is not None else "" for cell in row]
            if any(cleaned):
                sheet_lines.append(" | ".join(cleaned))
        if len(sheet_lines) > 1:
            all_text.append("\n".join(sheet_lines))
    wb.close()
    return "\n\n".join(all_text)

def read_pdf(filepath):
    doc = fitz.open(filepath)
    all_text = []
    for page_num in range(len(doc)):
        text = doc[page_num].get_text().strip()
        if text:
            all_text.append(f"[Page {page_num + 1}]\n{text}")
    doc.close()
    return "\n\n".join(all_text)

def chunk_text(text, chunk_size=500, overlap=100):
    chunks = []
    start = 0
    while start < len(text):
        chunks.append(text[start:start + chunk_size].strip())
        start += chunk_size - overlap
    return [c for c in chunks if len(c) > 50]

# ─────────────────────────────────────────────
# RAG BACKEND
# ─────────────────────────────────────────────
rag_model   = None
collection  = None
known_files = []

def load_backend(base_path):
    global rag_model, collection, known_files
    known_files = []
    dataframes.clear()

    db_path = os.path.join(base_path, "libby_db")
    rag_model  = SentenceTransformer("all-MiniLM-L6-v2")
    client     = chromadb.PersistentClient(path=db_path)
    collection = client.get_or_create_collection(name="libby_knowledge")

    existing_ids = set(collection.get()["ids"])
    new_chunks, new_embeddings, new_ids, new_metadata = [], [], [], []

    for root, dirs, files in os.walk(base_path):
        dirs[:] = [d for d in dirs if d not in ["libby_db", "__pycache__"]]
        for file in files:
            if file.startswith("libby") or file in ["libby_config.json"]:
                continue
            filepath = os.path.join(root, file)
            if file.endswith(".txt"):
                try:
                    with open(filepath, "r", encoding="utf-8") as f:
                        content = f.read()
                    file_type = "txt"
                except Exception as e:
                    print(f"Error: {e}"); continue
            elif file.endswith(".xlsx"):
                try:
                    content = read_excel_text(filepath)
                    file_type = "xlsx"
                    load_excel_as_dataframe(filepath)
                except Exception as e:
                    print(f"Error: {e}"); continue
            elif file.endswith(".pdf"):
                try:
                    content = read_pdf(filepath)
                    file_type = "pdf"
                    if not content.strip(): continue
                except Exception as e:
                    print(f"Error: {e}"); continue
            else:
                continue

            chunks = chunk_text(content)
            for i, chunk in enumerate(chunks):
                chunk_id = f"{filepath}::chunk{i}"
                if chunk_id in existing_ids: continue
                new_chunks.append(chunk)
                new_ids.append(chunk_id)
                new_metadata.append({
                    "source": filepath, "chunk_index": i,
                    "filename": file, "file_type": file_type
                })

    if new_chunks:
        new_embeddings = rag_model.encode(new_chunks).tolist()
        collection.add(documents=new_chunks, embeddings=new_embeddings,
                       ids=new_ids, metadatas=new_metadata)

    all_meta = collection.get(include=["metadatas"])["metadatas"]
    seen = set()
    for m in all_meta:
        src = m.get("source", "")
        if src and src not in seen:
            seen.add(src)
            known_files.append({
                "source": src,
                "filename": m.get("filename", os.path.basename(src)),
                "file_type": m.get("file_type", "txt")
            })

def retrieve(question):
    if not collection:
        return []
    query_embedding = rag_model.encode([question]).tolist()
    results = collection.query(
        query_embeddings=query_embedding, n_results=5,
        include=["documents", "metadatas", "distances"]
    )
    filtered = []
    for doc, meta, dist in zip(results["documents"][0],
                               results["metadatas"][0],
                               results["distances"][0]):
        if dist <= 1.4:
            filtered.append({
                "text": doc.strip(),
                "source": meta.get("source", "Unknown"),
                "filename": meta.get("filename", "Unknown"),
                "file_type": meta.get("file_type", "txt")
            })
    return filtered

def ask_ollama(question, chunks, calc_results, history, mode, ollama_model):
    system_prompt = PROMPT_EBI if mode == "ebi" else PROMPT_KNOWLEDGE

    if not chunks and not calc_results:
        return ("Not found." if mode == "ebi" else
                "I couldn't find anything relevant in the knowledge base. Try rephrasing."), []

    context_parts = []
    if calc_results:
        context_parts.append("=== CALCULATED DATA ===")
        context_parts.extend(calc_results)
    if chunks:
        context_parts.append("=== SOURCE DOCUMENTS ===")
        for c in chunks:
            context_parts.append(f"[{c['filename']}]\n{c['text']}")

    vague_words = ["that", "it", "those", "the same", "its", "their", "this", "them"]
    if any(w in question.lower().split() for w in vague_words) and history:
        full_row = lookup_full_row(question, history)
        if full_row:
            context_parts.append("=== FULL RECORD (follow-up context) ===")
            context_parts.append(full_row)

    context = "\n\n".join(context_parts)
    sources  = list(set([c["source"] for c in chunks])) if chunks else []

    history_block = ""
    if history:
        history_block = "\n--- RECENT CONVERSATION ---\n"
        for role, msg in history[-6:]:
            history_block += f"{role}: {msg}\n"
        history_block += "--- END CONVERSATION ---\n"

    prompt = f"""{system_prompt}
{history_block}
--- DATA ---
{context}
--- END ---

Question: {question}
Answer:"""

    try:
        response = requests.post(
            "http://localhost:11434/api/generate",
            json={"model": ollama_model, "prompt": prompt, "stream": False,
                  "options": {"temperature": 0.1 if mode == "knowledge" else 0.0,
                              "top_p": 0.9, "num_predict": 300 if mode == "knowledge" else 150}},
            timeout=120
        )
        response.raise_for_status()
        raw = response.json().get("response", "").strip()
        raw = raw.replace("```plaintext","").replace("```python","").replace("```","").strip()
        return raw, sources
    except requests.exceptions.ConnectionError:
        return "⚠️  Ollama not running. Open CMD and type: ollama serve", []
    except Exception as e:
        return f"⚠️  Error: {str(e)}", []

# ─────────────────────────────────────────────
# REPORT ENGINE (EBI mode)
# ─────────────────────────────────────────────
def find_best_dataframe(request):
    r = request.lower()
    dataset_hints = {
        "sales":     ["sales","revenue","monthly","rep","region","variance","target","units"],
        "inventory": ["inventory","stock","reorder","on hand","sku","category"],
        "customer":  ["customer","account","balance","credit","outstanding","ytd"],
        "purchase":  ["purchase","po","supplier","order","received"],
        "quotes":    ["quote","quoting"],
        "pricing":   ["price","pricing","margin"],
    }
    best_file, best_score = None, 0
    for filename, sheets in dataframes.items():
        fn_lower = filename.lower()
        score = 0
        for hint_key, keywords in dataset_hints.items():
            if hint_key in fn_lower:
                score += sum(2 for k in keywords if k in r)
            score += sum(1 for k in keywords if k in r and k in fn_lower)
        if score > best_score:
            best_score, best_file = score, filename
    if not best_file and dataframes:
        best_file = list(dataframes.keys())[0]
    if best_file:
        sheet_name = list(dataframes[best_file].keys())[0]
        return best_file, sheet_name, dataframes[best_file][sheet_name].copy()
    return None, None, None

def generate_excel_report(request, filepath, company_name):
    filename, sheet_name, df = find_best_dataframe(request)
    if df is None:
        return "No data available.", False

    r = request.lower()
    # Group detection
    group_map = {"region": ["Region"], "rep": ["Sales Rep"], "supplier": ["Supplier"],
                 "customer": ["Customer Name","Customer"], "category": ["Category"],
                 "month": ["Month"], "status": ["Status"], "product": ["Product Name","Product"]}
    group_col = None
    for kw, candidates in group_map.items():
        if kw in r:
            group_col = find_column(df, candidates)
            if group_col: break

    numeric_cols = df.select_dtypes(include="number").columns.tolist()
    if group_col and numeric_cols:
        df = df.groupby(group_col)[numeric_cols].sum().reset_index()
        numeric_cols = df.select_dtypes(include="number").columns.tolist()

    sort_asc = any(w in r for w in ["lowest to highest","ascending","smallest"])
    if numeric_cols:
        priority = ["Revenue","Total","Value","Amount","Cost","On Hand","Balance"]
        sort_col = next((find_column(df,[p]) for p in priority if find_column(df,[p]) in numeric_cols), numeric_cols[0])
        df = df.sort_values(by=sort_col, ascending=sort_asc).reset_index(drop=True)

    # Build Excel
    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = "Report"
    CHARCOAL = "252525"
    ROSE     = "c9956c"
    WHITE    = "ffffff"
    LIGHT    = "faf7f4"
    TOTAL_BG = "f5e8de"

    def tb():
        s = Side(style="thin", color="d5c8bc")
        return Border(left=s, right=s, top=s, bottom=s)

    nc = len(df.columns)
    lc = get_column_letter(nc)

    ws.merge_cells(f"A1:{lc}1")
    ws["A1"] = company_name
    ws["A1"].font      = Font(name="Calibri", bold=True, size=16, color=WHITE)
    ws["A1"].fill      = PatternFill("solid", fgColor=CHARCOAL)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    title = f"Report — {datetime.now().strftime('%B %d, %Y  %H:%M')}    Source: {filename}"
    ws.merge_cells(f"A2:{lc}2")
    ws["A2"] = title
    ws["A2"].font      = Font(name="Calibri", size=10, italic=True, color=ROSE)
    ws["A2"].fill      = PatternFill("solid", fgColor=CHARCOAL)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 6

    for ci, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=4, column=ci, value=col_name)
        cell.font      = Font(name="Calibri", bold=True, size=11, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=ROSE.replace("c9","8b").replace("956c","5e3c"))
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = tb()
    ws.row_dimensions[4].height = 22

    def is_cur(col): return any(w in col.lower() for w in ["price","cost","amount","revenue","sales","value","total","balance","limit","variance","quote","ytd","target"])
    num_col_names = df.select_dtypes(include="number").columns.tolist()
    last_row = 4

    for ri, (_, row) in enumerate(df.iterrows(), 5):
        bg = LIGHT if ri % 2 == 0 else WHITE
        for ci, (col_name, val) in enumerate(row.items(), 1):
            cell = ws.cell(row=ri, column=ci)
            if col_name in num_col_names and pd.notna(val):
                cell.value = val
                cell.number_format = '$#,##0.00' if is_cur(col_name) else '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.value = val
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.font   = Font(name="Calibri", size=10)
            cell.fill   = PatternFill("solid", fgColor=bg)
            cell.border = tb()
        ws.row_dimensions[ri].height = 18
        last_row = ri

    # Totals row
    tr = last_row + 1
    for ci, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=tr, column=ci)
        if col_name in num_col_names:
            cell.value = df[col_name].sum()
            cell.number_format = '$#,##0.00' if is_cur(col_name) else '#,##0.00'
            cell.alignment = Alignment(horizontal="right")
        else:
            cell.value = "TOTAL" if ci == 1 else ""
        cell.font = Font(name="Calibri", bold=True, size=11, color=CHARCOAL)
        cell.fill = PatternFill("solid", fgColor=TOTAL_BG)
        s = Side(style="medium", color=CHARCOAL)
        cell.border = Border(left=s, right=s, top=s, bottom=s)
    ws.row_dimensions[tr].height = 22

    # Footer
    fr = tr + 2
    ws.merge_cells(f"A{fr}:{lc}{fr}")
    fc = ws.cell(row=fr, column=1, value=f"🔒  Confidential  •  {company_name}  •  {datetime.now().strftime('%Y-%m-%d')}  •  Air-gapped")
    fc.font = Font(name="Calibri", size=9, italic=True, color=ROSE)
    fc.fill = PatternFill("solid", fgColor=CHARCOAL)
    fc.alignment = Alignment(horizontal="center")
    ws.row_dimensions[fr].height = 16

    for ci, col_name in enumerate(df.columns, 1):
        try:
            max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max() if not df.empty else 0)
        except Exception:
            max_len = len(str(col_name))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len + 4, 12), 40)

    wb.save(filepath)
    rows     = len(df)
    sort_dir = "lowest to highest" if sort_asc else "highest to lowest"
    summary  = f"✅  Report generated\n• {rows:,} rows\n• Sorted {sort_dir}"
    if group_col: summary += f"\n• Grouped by: {group_col}"
    return summary, True


# ─────────────────────────────────────────────
# MAIN APP
# ─────────────────────────────────────────────
class LibbyApp:
    def __init__(self, root):
        self.root          = root
        self.cfg           = load_config()
        self.T             = THEMES[self.cfg["theme"]]
        self.current_theme = self.cfg["theme"]
        self.mode          = self.cfg["mode"]   # "knowledge" or "ebi"
        self.ready         = False
        self.dot_count     = 0
        self.animating     = False
        self._anim_msg     = "⚡  Thinking"
        self._chat_widgets = []
        self._history      = []
        self._settings_open = False

        self.root.title("Libby V7 — Universal Knowledge System")
        self.root.geometry("1200x800")
        self.root.minsize(800, 600)
        self.root.configure(bg=self.T["BG_DARK"])

        self._build_ui()
        self._start_backend()

    # ─────────────────────────────────────────
    # BUILD UI
    # ─────────────────────────────────────────
    def _build_ui(self):
        T = self.T

        # ── Header ──────────────────────────
        self.header = tk.Frame(self.root, bg=T["BG_PANEL"], pady=12)
        self.header.pack(fill="x")

        left_h = tk.Frame(self.header, bg=T["BG_PANEL"])
        left_h.pack(side="left", padx=16)

        self.avatar_canvas = tk.Canvas(
            left_h, width=52, height=52,
            bg=T["BG_PANEL"], highlightthickness=0
        )
        self.avatar_canvas.pack(side="left", padx=(0, 12))
        self._draw_avatar()

        self.title_frame = tk.Frame(left_h, bg=T["BG_PANEL"])
        self.title_frame.pack(side="left")

        self.title_label = tk.Label(
            self.title_frame, text="LIBBY  V7",
            bg=T["BG_PANEL"], fg=T["ACCENT"],
            font=("Courier New", 22, "bold")
        )
        self.title_label.pack(anchor="w")

        self.company_label = tk.Label(
            self.title_frame, text=self.cfg["company_name"],
            bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
            font=("Courier New", 9)
        )
        self.company_label.pack(anchor="w")

        # Right controls
        self.right_frame = tk.Frame(self.header, bg=T["BG_PANEL"])
        self.right_frame.pack(side="right", padx=16)

        self.settings_btn = tk.Button(
            self.right_frame, text="⚙️  Settings",
            bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
            activebackground=T["BG_INPUT"],
            font=("Courier New", 9), relief="flat",
            cursor="hand2", command=self._toggle_settings
        )
        self.settings_btn.pack(anchor="e", pady=(0, 4))

        self.status_label = tk.Label(
            self.right_frame, text="⏳  Loading...",
            bg=T["BG_PANEL"], fg=T["WARNING"],
            font=("Courier New", 9)
        )
        self.status_label.pack(anchor="e")

        self.top_divider = tk.Frame(self.root, bg=T["BORDER"], height=1)
        self.top_divider.pack(fill="x")

        # ── Mode bar ────────────────────────
        self.mode_bar = tk.Frame(self.root, bg=T["BG_PANEL"], pady=6)
        self.mode_bar.pack(fill="x")

        tk.Label(
            self.mode_bar, text="MODE:",
            bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
            font=("Courier New", 8, "bold")
        ).pack(side="left", padx=(16, 8))

        self.kb_btn = tk.Button(
            self.mode_bar,
            text="📚  Knowledge Assistant",
            bg=T["ACCENT"] if self.mode == "knowledge" else T["BG_INPUT"],
            fg=T["BG_PANEL"] if self.mode == "knowledge" else T["TEXT_DIM"],
            font=("Courier New", 9, "bold" if self.mode == "knowledge" else "normal"),
            relief="flat", padx=12, pady=4,
            cursor="hand2", command=lambda: self._switch_mode("knowledge")
        )
        self.kb_btn.pack(side="left", padx=(0, 4))

        self.ebi_btn = tk.Button(
            self.mode_bar,
            text="📊  Enterprise BI",
            bg=T["ACCENT"] if self.mode == "ebi" else T["BG_INPUT"],
            fg=T["BG_PANEL"] if self.mode == "ebi" else T["TEXT_DIM"],
            font=("Courier New", 9, "bold" if self.mode == "ebi" else "normal"),
            relief="flat", padx=12, pady=4,
            cursor="hand2", command=lambda: self._switch_mode("ebi")
        )
        self.ebi_btn.pack(side="left")

        self.mode_hint = tk.Label(
            self.mode_bar,
            text=self._mode_hint(),
            bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
            font=("Courier New", 8)
        )
        self.mode_hint.pack(side="left", padx=16)

        self.mode_divider = tk.Frame(self.root, bg=T["BORDER"], height=1)
        self.mode_divider.pack(fill="x")

        # ── Body ────────────────────────────
        self.body = tk.Frame(self.root, bg=T["BG_DARK"])
        self.body.pack(fill="both", expand=True)

        # ── Sidebar ─────────────────────────
        self.sidebar = tk.Frame(self.body, bg=T["BG_SIDEBAR"], width=230)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)
        self._build_sidebar()

        self.side_divider = tk.Frame(self.body, bg=T["BORDER"], width=1)
        self.side_divider.pack(side="left", fill="y")

        # ── Chat ────────────────────────────
        self.chat_outer = tk.Frame(self.body, bg=T["BG_PANEL"])
        self.chat_outer.pack(side="left", fill="both", expand=True)

        self.chat_canvas = tk.Canvas(
            self.chat_outer, bg=T["BG_PANEL"],
            highlightthickness=0, bd=0
        )
        scrollbar = tk.Scrollbar(self.chat_outer, command=self.chat_canvas.yview)
        self.chat_canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.chat_canvas.pack(side="left", fill="both", expand=True)

        self.chat_inner = tk.Frame(self.chat_canvas, bg=T["BG_PANEL"])
        self.chat_window = self.chat_canvas.create_window(
            (0, 0), window=self.chat_inner, anchor="nw"
        )
        self.chat_inner.bind("<Configure>", self._on_frame_configure)
        self.chat_canvas.bind("<Configure>", self._on_canvas_configure)
        self.chat_canvas.bind("<Enter>", lambda e: self.chat_canvas.bind_all("<MouseWheel>", self._on_mousewheel))
        self.chat_canvas.bind("<Leave>", lambda e: self.chat_canvas.unbind_all("<MouseWheel>"))

        self.bottom_divider = tk.Frame(self.root, bg=T["BORDER"], height=1)
        self.bottom_divider.pack(fill="x")

        # ── Input ────────────────────────────
        self.input_frame = tk.Frame(self.root, bg=T["BG_PANEL"], pady=10)
        self.input_frame.pack(fill="x", padx=12)

        self.input_box = tk.Text(
            self.input_frame, height=3,
            bg=T["BG_INPUT"], fg=T["TEXT_MAIN"],
            insertbackground=T["ACCENT"],
            font=("Courier New", 11),
            relief="flat", bd=0,
            padx=12, pady=10, wrap="word"
        )
        self.input_box.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self.input_box.bind("<Return>", self._on_enter)
        self._set_placeholder()
        self.input_box.bind("<FocusIn>",  self._clear_placeholder)
        self.input_box.bind("<FocusOut>", self._restore_placeholder)

        self.btn_frame = tk.Frame(self.input_frame, bg=T["BG_PANEL"])
        self.btn_frame.pack(side="right")

        self.send_btn = tk.Button(
            self.btn_frame, text="ASK",
            bg=T["ACCENT_DIM"], fg=T["TEXT_MAIN"],
            activebackground=T["ACCENT"],
            font=("Courier New", 11, "bold"),
            relief="flat", padx=16, pady=4,
            cursor="hand2", command=self._send_question
        )
        self.send_btn.pack(pady=(0, 4))

        self.clear_btn = tk.Button(
            self.btn_frame, text="🗑  Clear",
            bg=T["BTN_CLEAR"], fg=T["BTN_CLEAR_FG"],
            activebackground=T["BTN_CLEAR"],
            font=("Courier New", 9),
            relief="flat", padx=10, pady=3,
            cursor="hand2", command=self._clear_chat
        )
        self.clear_btn.pack()

        # ── Footer ───────────────────────────
        self.footer = tk.Frame(self.root, bg=T["BG_PANEL"], pady=4)
        self.footer.pack(fill="x")
        self.footer_label = tk.Label(
            self.footer,
            text="🔒  Air-gapped  •  All answers from your local knowledge base",
            bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
            font=("Courier New", 8)
        )
        self.footer_label.pack()

        self._add_libby_message(self._welcome_message(), source=None, file_type=None)

    # ─────────────────────────────────────────
    # SIDEBAR
    # ─────────────────────────────────────────
    def _build_sidebar(self):
        T = self.T
        self.sidebar_title = tk.Label(
            self.sidebar, text="📁  KNOWLEDGE BASE",
            bg=T["BG_SIDEBAR"], fg=T["ACCENT"],
            font=("Courier New", 9, "bold"), pady=10
        )
        self.sidebar_title.pack(anchor="w", padx=12)

        tk.Frame(self.sidebar, bg=T["BORDER"], height=1).pack(fill="x", padx=8)

        self.path_label = tk.Label(
            self.sidebar,
            text=f"📂  {os.path.basename(self.cfg['base_path'])}",
            bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"],
            font=("Courier New", 7), pady=4,
            wraplength=200, justify="left"
        )
        self.path_label.pack(anchor="w", padx=12)

        tk.Frame(self.sidebar, bg=T["BORDER"], height=1).pack(fill="x", padx=8, pady=2)

        # Scrollable file list
        list_container = tk.Frame(self.sidebar, bg=T["BG_SIDEBAR"])
        list_container.pack(fill="both", expand=True)

        self.sidebar_canvas = tk.Canvas(
            list_container, bg=T["BG_SIDEBAR"],
            highlightthickness=0, bd=0
        )
        sb_scroll = tk.Scrollbar(list_container, orient="vertical",
                                 command=self.sidebar_canvas.yview)
        self.sidebar_canvas.configure(yscrollcommand=sb_scroll.set)
        sb_scroll.pack(side="right", fill="y")
        self.sidebar_canvas.pack(side="left", fill="both", expand=True)

        self.sidebar_list = tk.Frame(self.sidebar_canvas, bg=T["BG_SIDEBAR"])
        self.sidebar_list_window = self.sidebar_canvas.create_window(
            (0, 0), window=self.sidebar_list, anchor="nw"
        )
        self.sidebar_list.bind("<Configure>", lambda e: self.sidebar_canvas.configure(
            scrollregion=self.sidebar_canvas.bbox("all")
        ))
        self.sidebar_canvas.bind("<Configure>", lambda e: self.sidebar_canvas.itemconfig(
            self.sidebar_list_window, width=e.width
        ))
        self.sidebar_canvas.bind("<Enter>", lambda e: self.sidebar_canvas.bind_all(
            "<MouseWheel>", lambda ev: self.sidebar_canvas.yview_scroll(int(-1*(ev.delta/120)),"units")
        ))
        self.sidebar_canvas.bind("<Leave>", lambda e: self.sidebar_canvas.unbind_all("<MouseWheel>"))

        self.sidebar_footer_label = tk.Label(
            self.sidebar, text="📄 TXT   📊 XLSX   📕 PDF",
            bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"],
            font=("Courier New", 7), pady=6
        )
        self.sidebar_footer_label.pack(side="bottom")

    def _refresh_sidebar(self):
        T = self.T
        for w in self.sidebar_list.winfo_children():
            w.destroy()
        for entry in known_files:
            filename  = entry["filename"]
            folder    = os.path.basename(os.path.dirname(entry["source"]))
            file_type = entry.get("file_type", "txt")
            icon  = "📊" if file_type == "xlsx" else ("📕" if file_type == "pdf" else "📄")
            row = tk.Frame(self.sidebar_list, bg=T["BG_SIDEBAR"])
            row.pack(anchor="w", fill="x", pady=2, padx=8)
            lbl1 = tk.Label(row, text=f"{icon} {filename}",
                            bg=T["BG_SIDEBAR"], fg=T["TEXT_MAIN"],
                            font=("Courier New", 8), wraplength=190,
                            justify="left", anchor="w")
            lbl1.pack(anchor="w")
            lbl2 = tk.Label(row, text=f"   📁 {folder}",
                            bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"],
                            font=("Courier New", 7), anchor="w")
            lbl2.pack(anchor="w", pady=(0, 3))
            for widget in [row, lbl1, lbl2]:
                widget.bind("<Enter>", lambda e: self.sidebar_canvas.bind_all(
                    "<MouseWheel>", lambda ev: self.sidebar_canvas.yview_scroll(int(-1*(ev.delta/120)),"units")
                ))
                widget.bind("<Leave>", lambda e: self.sidebar_canvas.unbind_all("<MouseWheel>"))

    # ─────────────────────────────────────────
    # SETTINGS PANEL
    # ─────────────────────────────────────────
    def _toggle_settings(self):
        if self._settings_open:
            self._close_settings()
        else:
            self._open_settings()

    def _open_settings(self):
        T = self.T
        self._settings_open = True
        self.settings_btn.config(text="✕  Close Settings")

        self.settings_panel = tk.Frame(
            self.root, bg=T["BG_SETTINGS"],
            padx=20, pady=16,
            highlightthickness=1,
            highlightbackground=T["ACCENT"]
        )
        self.settings_panel.place(relx=1.0, rely=0.0, anchor="ne", x=-16, y=60)

        tk.Label(
            self.settings_panel, text="⚙️  SETTINGS",
            bg=T["BG_SETTINGS"], fg=T["ACCENT"],
            font=("Courier New", 11, "bold")
        ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 12))

        # ── Knowledge folder ────────────────
        tk.Label(
            self.settings_panel, text="Knowledge Folder:",
            bg=T["BG_SETTINGS"], fg=T["TEXT_MAIN"],
            font=("Courier New", 9)
        ).grid(row=1, column=0, sticky="w", pady=4)

        folder_frame = tk.Frame(self.settings_panel, bg=T["BG_SETTINGS"])
        folder_frame.grid(row=1, column=1, sticky="w", padx=(8, 0), pady=4)

        self.folder_var = tk.StringVar(value=self.cfg["base_path"])
        self.folder_entry = tk.Entry(
            folder_frame, textvariable=self.folder_var,
            bg=T["BG_INPUT"], fg=T["TEXT_MAIN"],
            insertbackground=T["ACCENT"],
            font=("Courier New", 8),
            relief="flat", width=28
        )
        self.folder_entry.pack(side="left", padx=(0, 4))

        tk.Button(
            folder_frame, text="Browse",
            bg=T["ACCENT_DIM"], fg=T["TEXT_MAIN"],
            font=("Courier New", 8), relief="flat",
            padx=8, cursor="hand2",
            command=self._browse_folder
        ).pack(side="left")

        # ── Theme ───────────────────────────
        tk.Label(
            self.settings_panel, text="Default Theme:",
            bg=T["BG_SETTINGS"], fg=T["TEXT_MAIN"],
            font=("Courier New", 9)
        ).grid(row=2, column=0, sticky="w", pady=4)

        theme_frame = tk.Frame(self.settings_panel, bg=T["BG_SETTINGS"])
        theme_frame.grid(row=2, column=1, sticky="w", padx=(8, 0), pady=4)

        self.theme_var = tk.StringVar(value=self.cfg["theme"])
        for label, val in [("🌙 Dark", "dark"), ("☀️ Light", "light")]:
            tk.Radiobutton(
                theme_frame, text=label, variable=self.theme_var, value=val,
                bg=T["BG_SETTINGS"], fg=T["TEXT_MAIN"],
                selectcolor=T["BG_INPUT"],
                activebackground=T["BG_SETTINGS"],
                font=("Courier New", 9)
            ).pack(side="left", padx=(0, 12))

        # ── Company name ─────────────────────
        tk.Label(
            self.settings_panel, text="Company Name:",
            bg=T["BG_SETTINGS"], fg=T["TEXT_MAIN"],
            font=("Courier New", 9)
        ).grid(row=3, column=0, sticky="w", pady=4)

        self.company_var = tk.StringVar(value=self.cfg["company_name"])
        tk.Entry(
            self.settings_panel, textvariable=self.company_var,
            bg=T["BG_INPUT"], fg=T["TEXT_MAIN"],
            insertbackground=T["ACCENT"],
            font=("Courier New", 9), relief="flat", width=28
        ).grid(row=3, column=1, sticky="w", padx=(8, 0), pady=4)

        # ── Divider ─────────────────────────
        tk.Frame(self.settings_panel, bg=T["BORDER"], height=1).grid(
            row=4, column=0, columnspan=2, sticky="ew", pady=10
        )

        # ── Save button ─────────────────────
        btn_frame = tk.Frame(self.settings_panel, bg=T["BG_SETTINGS"])
        btn_frame.grid(row=5, column=0, columnspan=2, pady=(0, 4))

        tk.Button(
            btn_frame, text="💾  Save & Apply",
            bg=T["ACCENT_DIM"], fg=T["TEXT_MAIN"],
            activebackground=T["ACCENT"],
            font=("Courier New", 10, "bold"),
            relief="flat", padx=16, pady=6,
            cursor="hand2", command=self._save_settings
        ).pack(side="left", padx=(0, 8))

        tk.Button(
            btn_frame, text="Cancel",
            bg=T["BG_INPUT"], fg=T["TEXT_DIM"],
            font=("Courier New", 9),
            relief="flat", padx=10, pady=6,
            cursor="hand2", command=self._close_settings
        ).pack(side="left")

    def _close_settings(self):
        self._settings_open = False
        self.settings_btn.config(text="⚙️  Settings")
        if hasattr(self, "settings_panel"):
            self.settings_panel.destroy()

    def _browse_folder(self):
        folder = filedialog.askdirectory(
            title="Select Your Knowledge Base Folder",
            initialdir=self.cfg["base_path"]
        )
        if folder:
            self.folder_var.set(folder)

    def _save_settings(self):
        new_path    = self.folder_var.get().strip()
        new_theme   = self.theme_var.get()
        new_company = self.company_var.get().strip() or "Libby Intelligence"

        if not os.path.exists(new_path):
            messagebox.showerror("Invalid Folder",
                                 f"Folder not found:\n{new_path}\nPlease choose a valid folder.")
            return

        changed_path  = new_path    != self.cfg["base_path"]
        changed_theme = new_theme   != self.cfg["theme"]

        self.cfg["base_path"]    = new_path
        self.cfg["theme"]        = new_theme
        self.cfg["company_name"] = new_company
        save_config(self.cfg)

        self._close_settings()

        # Apply theme change immediately
        if changed_theme:
            self.current_theme = new_theme
            self.T = THEMES[new_theme]
            self._apply_theme()

        # Update company label
        self.company_label.config(text=new_company)
        self.path_label.config(text=f"📂  {os.path.basename(new_path)}")

        # Reload backend if folder changed
        if changed_path:
            self.ready = False
            self.status_label.config(text="⏳  Reloading...", fg=self.T["WARNING"])
            self._add_libby_message(
                f"Knowledge folder changed to:\n{new_path}\nReloading...",
                source=None, file_type=None
            )
            self._start_backend()
        else:
            messagebox.showinfo("Settings Saved", "✅  Settings saved successfully!")

    # ─────────────────────────────────────────
    # MODE SWITCHING
    # ─────────────────────────────────────────
    def _mode_hint(self):
        if self.mode == "knowledge":
            return "Ask questions from your documents"
        return "Type 'generate report...' to export Excel"

    def _switch_mode(self, mode):
        self.mode = mode
        self.cfg["mode"] = mode
        save_config(self.cfg)
        T = self.T
        self.kb_btn.config(
            bg=T["ACCENT"] if mode == "knowledge" else T["BG_INPUT"],
            fg=T["BG_PANEL"] if mode == "knowledge" else T["TEXT_DIM"],
            font=("Courier New", 9, "bold" if mode == "knowledge" else "normal")
        )
        self.ebi_btn.config(
            bg=T["ACCENT"] if mode == "ebi" else T["BG_INPUT"],
            fg=T["BG_PANEL"] if mode == "ebi" else T["TEXT_DIM"],
            font=("Courier New", 9, "bold" if mode == "ebi" else "normal")
        )
        self.mode_hint.config(text=self._mode_hint())
        self.send_btn.config(text="ASK" if mode == "knowledge" else "QUERY")
        self._set_placeholder()
        self._add_libby_message(
            f"Switched to {'📚 Knowledge Assistant' if mode == 'knowledge' else '📊 Enterprise BI'} mode.",
            source=None, file_type=None
        )

    def _welcome_message(self):
        if self.mode == "knowledge":
            return ("LIBBY V7 — Knowledge Assistant Mode.\n"
                    "Ask me anything from your loaded documents.\n"
                    "Switch to Enterprise BI mode for data calculations and reports.")
        return ("LIBBY V7 — Enterprise BI Mode.\n"
                "Ask data questions or type 'generate report...' to export Excel.\n"
                "Switch to Knowledge Assistant mode for document Q&A.")

    # ─────────────────────────────────────────
    # AVATAR
    # ─────────────────────────────────────────
    def _draw_avatar(self):
        c = self.avatar_canvas
        T = self.T
        c.delete("all")
        # Charcoal background with rose gold accents
        c.create_rectangle(4, 4, 48, 48, fill=T["BG_PANEL"],
                           outline=T["ACCENT"], width=2)
        # Book pages
        c.create_rectangle(10, 10, 30, 42, fill=T["BG_INPUT"], outline=T["BORDER"])
        c.create_rectangle(22, 10, 42, 42, fill=T["BG_INPUT"], outline=T["BORDER"])
        # Spine
        c.create_rectangle(20, 10, 24, 42, fill=T["ACCENT"], outline="")
        # Lines on pages
        for y in [16, 20, 24, 28, 32, 36]:
            c.create_line(12, y, 19, y, fill=T["ACCENT_DIM"], width=1)
            c.create_line(25, y, 40, y, fill=T["ACCENT_DIM"], width=1)
        tk.Label(c, text="V7", bg=T["BG_PANEL"], fg=T["ACCENT"],
                 font=("Courier New", 6, "bold")).place(x=30, y=5)

    # ─────────────────────────────────────────
    # THEME
    # ─────────────────────────────────────────
    def _apply_theme(self):
        T = self.T
        self.root.configure(bg=T["BG_DARK"])
        self.header.configure(bg=T["BG_PANEL"])
        self.avatar_canvas.configure(bg=T["BG_PANEL"])
        self._draw_avatar()
        self.title_frame.configure(bg=T["BG_PANEL"])
        self.title_label.configure(bg=T["BG_PANEL"], fg=T["ACCENT"])
        self.company_label.configure(bg=T["BG_PANEL"], fg=T["TEXT_DIM"])
        self.right_frame.configure(bg=T["BG_PANEL"])
        self.settings_btn.configure(bg=T["BG_PANEL"], fg=T["TEXT_DIM"])
        self.status_label.configure(bg=T["BG_PANEL"])
        self.top_divider.configure(bg=T["BORDER"])
        self.mode_bar.configure(bg=T["BG_PANEL"])
        self.mode_hint.configure(bg=T["BG_PANEL"], fg=T["TEXT_DIM"])
        self.mode_divider.configure(bg=T["BORDER"])
        self.body.configure(bg=T["BG_DARK"])
        self.sidebar.configure(bg=T["BG_SIDEBAR"])
        self.sidebar_title.configure(bg=T["BG_SIDEBAR"], fg=T["ACCENT"])
        self.path_label.configure(bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"])
        self.sidebar_canvas.configure(bg=T["BG_SIDEBAR"])
        self.sidebar_list.configure(bg=T["BG_SIDEBAR"])
        self.sidebar_footer_label.configure(bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"])
        self.side_divider.configure(bg=T["BORDER"])
        self.chat_outer.configure(bg=T["BG_PANEL"])
        self.chat_canvas.configure(bg=T["BG_PANEL"])
        self.chat_inner.configure(bg=T["BG_PANEL"])
        self.bottom_divider.configure(bg=T["BORDER"])
        self.input_frame.configure(bg=T["BG_PANEL"])
        self.btn_frame.configure(bg=T["BG_PANEL"])
        self.input_box.configure(bg=T["BG_INPUT"], fg=T["TEXT_MAIN"],
                                 insertbackground=T["ACCENT"])
        self.send_btn.configure(bg=T["ACCENT_DIM"], fg=T["TEXT_MAIN"])
        self.clear_btn.configure(bg=T["BTN_CLEAR"], fg=T["BTN_CLEAR_FG"])
        self.footer.configure(bg=T["BG_PANEL"])
        self.footer_label.configure(bg=T["BG_PANEL"], fg=T["TEXT_DIM"])
        self.status_label.configure(
            fg=T["SUCCESS"] if self.current_theme == "light" else T["WARNING"]
        )
        self._switch_mode(self.mode)
        self._refresh_sidebar()
        self._refresh_chat_bubbles()

    # ─────────────────────────────────────────
    # CHAT BUBBLES
    # ─────────────────────────────────────────
    def _add_user_message(self, text):
        T     = self.T
        outer = tk.Frame(self.chat_inner, bg=T["BG_PANEL"], pady=6)
        outer.pack(fill="x", padx=16)
        self._chat_widgets.append(("user", text, None, None))
        tk.Label(outer, text="YOU", bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
                 font=("Courier New", 8, "bold")).pack(anchor="e", padx=4)
        bubble = tk.Frame(outer, bg=T["BG_BUBBLE_U"], padx=16, pady=10)
        bubble.pack(anchor="e")
        tk.Label(bubble, text=text, bg=T["BG_BUBBLE_U"], fg=T["TEXT_BUBBLE_U"],
                 font=("Courier New", 11), wraplength=700, justify="left").pack(anchor="w")

    def _add_libby_message(self, text, source=None, file_type=None):
        T     = self.T
        outer = tk.Frame(self.chat_inner, bg=T["BG_PANEL"], pady=6)
        outer.pack(fill="x", padx=16)
        self._chat_widgets.append(("libby", text, source, file_type))
        tk.Label(outer, text="LIBBY V7", bg=T["BG_PANEL"], fg=T["ACCENT"],
                 font=("Courier New", 8, "bold")).pack(anchor="w", padx=4)
        bubble = tk.Frame(outer, bg=T["BG_BUBBLE_L"], padx=16, pady=10)
        bubble.pack(anchor="w")
        tk.Label(bubble, text=text, bg=T["BG_BUBBLE_L"], fg=T["TEXT_BUBBLE_L"],
                 font=("Courier New", 11), wraplength=700, justify="left").pack(anchor="w")
        if source:
            icon = "📊" if file_type == "xlsx" else ("📕" if file_type == "pdf" else "📄")
            tk.Label(bubble, text=f"{icon}  {source}",
                     bg=T["BG_BUBBLE_L"], fg=T["TEXT_SOURCE"],
                     font=("Courier New", 8), wraplength=700,
                     justify="left").pack(anchor="w", pady=(6, 0))

    def _refresh_chat_bubbles(self):
        saved = list(self._chat_widgets)
        for w in self.chat_inner.winfo_children():
            w.destroy()
        self._chat_widgets = []
        for kind, text, source, file_type in saved:
            if kind == "user":
                self._add_user_message(text)
            else:
                self._add_libby_message(text, source, file_type)

    # ─────────────────────────────────────────
    # THINKING ANIMATION
    # ─────────────────────────────────────────
    def _add_thinking_frame(self, is_report=False):
        T     = self.T
        frame = tk.Frame(self.chat_inner, bg=T["BG_PANEL"], pady=6)
        frame.pack(fill="x", padx=16)
        tk.Label(frame, text="LIBBY V7", bg=T["BG_PANEL"], fg=T["ACCENT"],
                 font=("Courier New", 8, "bold")).pack(anchor="w", padx=4)
        self._anim_msg = "📊  Building report" if is_report else "✨  Thinking"
        self.thinking_label = tk.Label(
            frame, text=self._anim_msg,
            bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
            font=("Courier New", 10, "italic")
        )
        self.thinking_label.pack(anchor="w", padx=10)
        self.animating = True
        self._animate_dots()
        self._scroll_to_bottom()
        return frame

    def _animate_dots(self):
        if not self.animating:
            return
        dots = "." * (self.dot_count % 4)
        self.thinking_label.config(text=f"{self._anim_msg}{dots}")
        self.dot_count += 1
        self.root.after(300, self._animate_dots)

    def _stop_animation(self, frame):
        self.animating = False
        frame.destroy()

    # ─────────────────────────────────────────
    # CLEAR
    # ─────────────────────────────────────────
    def _clear_chat(self):
        for w in self.chat_inner.winfo_children():
            w.destroy()
        self._chat_widgets = []
        self._history      = []
        self._add_libby_message("Session cleared. Ready for questions.",
                                source=None, file_type=None)

    # ─────────────────────────────────────────
    # SCROLL
    # ─────────────────────────────────────────
    def _on_frame_configure(self, event):
        self.chat_canvas.configure(scrollregion=self.chat_canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.chat_canvas.itemconfig(self.chat_window, width=event.width)

    def _on_mousewheel(self, event):
        self.chat_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _scroll_to_bottom(self):
        self.root.after(50, lambda: self.chat_canvas.yview_moveto(1.0))

    # ─────────────────────────────────────────
    # INPUT
    # ─────────────────────────────────────────
    def _placeholder_text(self):
        return ("Ask me anything from your knowledge base..."
                if self.mode == "knowledge"
                else "Ask a data question or type 'generate report ...'")

    def _set_placeholder(self):
        self.input_box.delete("1.0", "end")
        self.input_box.insert("1.0", self._placeholder_text())
        self.input_box.config(fg=self.T["TEXT_DIM"])

    def _clear_placeholder(self, event):
        if self.input_box.get("1.0", "end-1c") == self._placeholder_text():
            self.input_box.delete("1.0", "end")
            self.input_box.config(fg=self.T["TEXT_MAIN"])

    def _restore_placeholder(self, event):
        if not self.input_box.get("1.0", "end-1c").strip():
            self._set_placeholder()

    def _on_enter(self, event):
        if not event.state & 0x1:
            self._send_question()
            return "break"

    # ─────────────────────────────────────────
    # SEND & QUERY
    # ─────────────────────────────────────────
    def _send_question(self):
        if not self.ready:
            return
        placeholder = self._placeholder_text()
        question = self.input_box.get("1.0", "end-1c").strip()
        if not question or question == placeholder:
            return

        self.input_box.delete("1.0", "end")
        self.input_box.config(fg=self.T["TEXT_MAIN"])
        self._add_user_message(question)

        is_report = self.mode == "ebi" and "generate report" in question.lower()
        thinking_frame = self._add_thinking_frame(is_report=is_report)
        self.send_btn.config(state="disabled", bg=self.T["BG_INPUT"])
        self.input_box.config(state="disabled")

        if is_report:
            self.root.after(0, lambda: self._prompt_save_report(question, thinking_frame))
        else:
            def run_query():
                # Enrich vague follow-up questions with context
                vague = ["that","it","those","the same","its","their","this","them"]
                search_q = question
                if any(w in question.lower().split() for w in vague) and self._history:
                    last_user = [m for r, m in self._history if r == "User"]
                    if last_user:
                        search_q = f"{last_user[-1]} {question}"

                chunks       = retrieve(search_q)
                calc_results = get_calculations(search_q) if self.mode == "ebi" else []
                answer, sources = ask_ollama(
                    question, chunks, calc_results,
                    self._history, self.mode,
                    self.cfg["ollama_model"]
                )
                source_str = ", ".join(sources) if sources else None
                file_type  = chunks[0]["file_type"] if chunks else None
                self._history.append(("User", question))
                self._history.append(("Libby", answer))
                if len(self._history) > 10:
                    self._history = self._history[-10:]
                self.root.after(0, lambda: self._show_answer(
                    answer, source_str, file_type, thinking_frame
                ))
            threading.Thread(target=run_query, daemon=True).start()

    def _prompt_save_report(self, question, thinking_frame):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath  = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"Libby_Report_{timestamp}.xlsx",
            title="Save Report As"
        )
        if not filepath:
            self._stop_animation(thinking_frame)
            self._add_libby_message("Report cancelled.", source=None, file_type=None)
            self._re_enable_input()
            return

        def do_generate():
            summary, success = generate_excel_report(
                question, filepath, self.cfg["company_name"]
            )
            self.root.after(0, lambda: self._show_report_result(
                summary, success, filepath, thinking_frame
            ))
        threading.Thread(target=do_generate, daemon=True).start()

    def _show_report_result(self, summary, success, filepath, thinking_frame):
        self._stop_animation(thinking_frame)
        msg = f"{summary}\n📁  Saved: {os.path.basename(filepath)}" if success else summary
        self._add_libby_message(msg, source=filepath if success else None,
                                file_type="xlsx" if success else None)
        self._scroll_to_bottom()
        self._re_enable_input()

    def _show_answer(self, answer, source, file_type, thinking_frame):
        self._stop_animation(thinking_frame)
        self._add_libby_message(answer, source=source, file_type=file_type)
        self._scroll_to_bottom()
        self._re_enable_input()

    def _re_enable_input(self):
        self.send_btn.config(state="normal", bg=self.T["ACCENT_DIM"])
        self.input_box.config(state="normal")
        self.input_box.focus()
        self.root.update_idletasks()
        self.root.update()

    # ─────────────────────────────────────────
    # BACKEND
    # ─────────────────────────────────────────
    def _start_backend(self):
        def load():
            load_backend(self.cfg["base_path"])
            self.ready = True
            count = collection.count() if collection else 0
            self.root.after(0, lambda: self.status_label.config(
                text=f"✅  {count} chunks  •  {self.cfg['ollama_model']}  •  Offline",
                fg=self.T["SUCCESS"]
            ))
            self.root.after(0, self._refresh_sidebar)
        threading.Thread(target=load, daemon=True).start()


# ─────────────────────────────────────────────
# LAUNCH
# ─────────────────────────────────────────────
if __name__ == "__main__":
    root = tk.Tk()
    app  = LibbyApp(root)
    root.mainloop()