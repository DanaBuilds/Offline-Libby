import os
import chromadb
import openpyxl
from sentence_transformers import SentenceTransformer

BASE_PATH = "C:\\RAG Test"
DB_PATH   = "C:\\RAG Test\\rag_db"

print("Loading model...")
model = SentenceTransformer("all-MiniLM-L6-v2")

client     = chromadb.PersistentClient(path=DB_PATH)
collection = client.get_or_create_collection(name="knowledge")

for root, dirs, files in os.walk(BASE_PATH):
    dirs[:] = [d for d in dirs if "rag_db" not in d]
    for f in files:
        filepath = os.path.join(root, f)
        try:
            if f.endswith(".txt"):
                with open(filepath, "r", encoding="utf-8") as fh:
                    content = fh.read()
            elif f.endswith(".xlsx"):
                wb = openpyxl.load_workbook(filepath, data_only=True)
                lines = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        cleaned = [str(c) for c in row if c is not None]
                        if cleaned:
                            lines.append(" | ".join(cleaned))
                content = "\n".join(lines)
                wb.close()
            else:
                continue

            chunk_id  = filepath
            embedding = model.encode([content]).tolist()
            collection.add(
                documents=[content],
                embeddings=embedding,
                ids=[chunk_id],
                metadatas=[{"source": filepath, "filename": f}]
            )
            print(f"✅ Indexed: {f}")

        except Exception as e:
            print(f"❌ Failed on {f}: {e}")

print(f"\nTotal chunks in DB: {collection.count()}")
