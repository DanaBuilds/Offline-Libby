import tkinter as tk
from tkinter import filedialog, messagebox
import threading
import os
import json
import uuid
import requests
import chromadb
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import fitz
import pandas as pd
from sentence_transformers import SentenceTransformer
from datetime import datetime
import re
import shutil
import gc

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
SCRIPT_DIR     = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE    = os.path.join(SCRIPT_DIR, "libby_config.json")
DEFAULT_CONFIG = {
    "base_path":    SCRIPT_DIR,
    "theme":        "dark",
    "mode":         "knowledge",
    "company_name": "Libby Intelligence",
    "ollama_model": "llama3.2",
}

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r") as f:
                cfg = json.load(f)
            for k, v in DEFAULT_CONFIG.items():
                cfg.setdefault(k, v)
            # Always respect the code default for ollama_model
            if cfg.get("ollama_model") == "phi3:mini":
                cfg["ollama_model"] = DEFAULT_CONFIG["ollama_model"]
            return cfg
        except Exception:
            pass
    return DEFAULT_CONFIG.copy()

def save_config(cfg):
    with open(CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)

# ─────────────────────────────────────────────
# FONTS
# ─────────────────────────────────────────────
F_TITLE    = ("Segoe UI", 22, "bold")
F_SUBTITLE = ("Segoe UI", 9)
F_LABEL    = ("Segoe UI", 9)
F_LABEL_B  = ("Segoe UI", 9, "bold")
F_SMALL    = ("Segoe UI", 8)
F_TINY     = ("Segoe UI", 7)
F_CHAT     = ("Segoe UI", 11)
F_BTN      = ("Segoe UI", 10, "bold")
F_BTN_SM   = ("Segoe UI", 9)
F_INPUT    = ("Segoe UI", 11)

# ─────────────────────────────────────────────
# THEMES
# ─────────────────────────────────────────────
THEMES = {
    "dark": {
        "BG_DARK":        "#1e1e1e",
        "BG_PANEL":       "#252525",
        "BG_INPUT":       "#1a1a1a",
        "BG_BUBBLE_U":    "#2e2420",
        "BG_BUBBLE_L":    "#2a2a2a",
        "BG_SIDEBAR":     "#1a1a1a",
        "BG_SETTINGS":    "#2a2a2a",
        "BG_TAB_ACTIVE":  "#252525",
        "BG_TAB_IDLE":    "#1a1a1a",
        "ACCENT":         "#c9956c",
        "ACCENT_DIM":     "#8b5e3c",
        "ACCENT_BRIGHT":  "#e8b49a",
        "TEXT_MAIN":      "#f0ece8",
        "TEXT_DIM":       "#6b6560",
        "TEXT_BUBBLE_U":  "#f0ece8",
        "TEXT_BUBBLE_L":  "#f0ece8",
        "TEXT_SOURCE":    "#c9956c",
        "TEXT_TAB_ACT":   "#f0ece8",
        "TEXT_TAB_IDLE":  "#6b6560",
        "SUCCESS":        "#c9956c",
        "WARNING":        "#e8c547",
        "BORDER":         "#333333",
        "BORDER_TAB":     "#c9956c",
        "BTN_CLEAR":      "#2a1a1a",
        "BTN_CLEAR_FG":   "#e05555",
        "DOT_COLOR":      "#c9956c",
    },
    "light": {
        "BG_DARK":        "#f5f0eb",
        "BG_PANEL":       "#ffffff",
        "BG_INPUT":       "#ede8e2",
        "BG_BUBBLE_U":    "#e8ddd4",
        "BG_BUBBLE_L":    "#f5f0eb",
        "BG_SIDEBAR":     "#ede8e2",
        "BG_SETTINGS":    "#f0ebe5",
        "BG_TAB_ACTIVE":  "#ffffff",
        "BG_TAB_IDLE":    "#ede8e2",
        "ACCENT":         "#9b5e3a",
        "ACCENT_DIM":     "#7a4a2c",
        "ACCENT_BRIGHT":  "#c9956c",
        "TEXT_MAIN":      "#2a2018",
        "TEXT_DIM":       "#8a7a6a",
        "TEXT_BUBBLE_U":  "#2a2018",
        "TEXT_BUBBLE_L":  "#2a2018",
        "TEXT_SOURCE":    "#9b5e3a",
        "TEXT_TAB_ACT":   "#2a2018",
        "TEXT_TAB_IDLE":  "#8a7a6a",
        "SUCCESS":        "#3a7a3a",
        "WARNING":        "#8a6a00",
        "BORDER":         "#d5c8bc",
        "BORDER_TAB":     "#9b5e3a",
        "BTN_CLEAR":      "#f0ddd8",
        "BTN_CLEAR_FG":   "#8a2020",
        "DOT_COLOR":      "#9b5e3a",
    }
}

# ─────────────────────────────────────────────
# PROMPTS
# ─────────────────────────────────────────────
PROMPT_KNOWLEDGE = """You are Libby, an offline knowledge assistant.
Answer questions using ONLY the provided context.
- Always complete your sentences fully. Never stop mid-sentence.
- Give thorough, detailed answers using ALL relevant information from the context.
- Use bullet points with "•" for lists.
- Number steps when applicable: 1. 2. 3.
- Include specific numbers, times, quantities and instructions when present.
- If not found in context: "Not found in knowledge base."
- Never guess or use outside knowledge."""

PROMPT_EBI = """You are Libby, an ultra-concise enterprise data retrieval system.
- Return numbers and dollar amounts ONLY. No narrative.
- Maximum 3 lines per answer.
- Show formula if calculating: e.g. $6,750 + $4,800 = $11,550
- Format currency: $1,234.56
- Format percentages: 12.5%
- If not found: "Not found."
- NEVER use markdown. Plain text only."""

# ─────────────────────────────────────────────
# PANDAS
# ─────────────────────────────────────────────
dataframes = {}

def load_excel_as_dataframe(filepath):
    filename = os.path.basename(filepath)
    try:
        xl = pd.ExcelFile(filepath)
        sheets = {}
        for sheet in xl.sheet_names:
            df = pd.read_excel(filepath, sheet_name=sheet)
            df.columns = [str(c).strip() for c in df.columns]
            sheets[sheet] = df
        dataframes[filename] = sheets
    except Exception as e:
        print(f"  → Could not load {filename}: {e}")

def find_column(df, candidates):
    cols_lower = {c.lower(): c for c in df.columns}
    for candidate in candidates:
        c = candidate.lower()
        if c in cols_lower:
            return cols_lower[c]
        for col_lower, col_actual in cols_lower.items():
            if c in col_lower or col_lower in c:
                return col_actual
    return None

def lookup_full_row(history):
    if not history: return ""
    last_answer = next((m for r, m in reversed(history) if r == "Libby"), "")
    if not last_answer: return ""
    amounts = re.findall(r'\$?[\d,]+\.?\d*', last_answer)
    for filename, sheets in dataframes.items():
        for sheet_name, df in sheets.items():
            for _, row in df.iterrows():
                row_str = " ".join([str(v) for v in row.values]).lower()
                if sum(1 for a in amounts if a.replace('$','').replace(',','') in row_str.replace(',','')) > 0:
                    lines = [f"\n[Full record from {filename} — {sheet_name}]"]
                    for col, val in row.items():
                        lines.append(f"{col}: {val}")
                    return "\n".join(lines)
    return ""

def perform_calculation(question, filename, sheet_name, df):
    q = question.lower()
    result_lines = []
    try:
        numeric_cols = df.select_dtypes(include="number").columns.tolist()
        if not numeric_cols: return None
        def is_cur(col): return any(w in col.lower() for w in ["price","cost","amount","revenue","sales","value","total","balance","limit","variance","quote","ytd","target","salary","budget","spent","profit","margin","remaining"])
        def fmt(val, col): return f"${val:,.2f}" if is_cur(col) else f"{val:,.2f}"
        def get_id(df, idx):
            id_cols = [c for c in df.columns if any(w in c.lower() for w in ["name","sku","id","product","item","description","number","customer","rep","month","project","employee"])]
            return f" ({df[id_cols[0]].iloc[idx]})" if id_cols else ""
        if any(w in q for w in ["total","sum","how much"]):
            for col in numeric_cols:
                total = df[col].sum()
                if total != 0: result_lines.append(f"{col}: {fmt(total, col)}")
        if any(w in q for w in ["average","avg","mean"]):
            for col in numeric_cols:
                result_lines.append(f"Avg {col}: {fmt(df[col].mean(), col)}")
        if any(w in q for w in ["highest","maximum","max","most","top"]):
            for col in numeric_cols:
                idx = df[col].idxmax()
                result_lines.append(f"Max {col}: {fmt(df[col].max(), col)}{get_id(df, idx)}")
        if any(w in q for w in ["lowest","minimum","min","least"]):
            for col in numeric_cols:
                idx = df[col].idxmin()
                result_lines.append(f"Min {col}: {fmt(df[col].min(), col)}{get_id(df, idx)}")
        if any(w in q for w in ["how many","count","number of"]):
            result_lines.append(f"Count: {len(df):,}")
    except Exception as e:
        print(f"Calc error: {e}")
    return (f"[{filename} — {sheet_name}]\n" + "\n".join(result_lines)) if result_lines else None

def get_calculations(question):
    results = []
    for filename, sheets in dataframes.items():
        for sheet_name, df in sheets.items():
            calc = perform_calculation(question, filename, sheet_name, df)
            if calc: results.append(calc)
    return results

# ─────────────────────────────────────────────
# FILE READERS
# ─────────────────────────────────────────────
def read_excel_text(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    all_text = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines = [f"Sheet: {sheet_name}"]
        for row in ws.iter_rows(values_only=True):
            cleaned = [str(c).strip() if c is not None else "" for c in row]
            if any(cleaned): lines.append(" | ".join(cleaned))
        if len(lines) > 1: all_text.append("\n".join(lines))
    wb.close()
    return "\n\n".join(all_text)

def read_pdf(filepath):
    doc = fitz.open(filepath)
    all_text = []
    for i in range(len(doc)):
        text = doc[i].get_text().strip()
        if text: all_text.append(f"[Page {i+1}]\n{text}")
    doc.close()
    return "\n\n".join(all_text)

def chunk_text(text, chunk_size=1200, overlap=150):
    chunks, start = [], 0
    while start < len(text):
        chunks.append(text[start:start+chunk_size].strip())
        start += chunk_size - overlap
    return [c for c in chunks if len(c) > 50]

# ─────────────────────────────────────────────
# RAG BACKEND
# ─────────────────────────────────────────────
rag_model   = None
collection  = None
known_files = []

def load_backend(base_path, fresh=False):
    global rag_model, collection, known_files
    known_files = []
    dataframes.clear()
    db_path = os.path.join(base_path, "libby_db")
    if fresh and os.path.exists(db_path):
        collection = None
        gc.collect()
        try:
            shutil.rmtree(db_path)
        except Exception as e:
            print(f"  → Could not delete old DB: {e}")
    rag_model  = SentenceTransformer("all-MiniLM-L6-v2")
    client     = chromadb.PersistentClient(path=db_path)
    folder_id  = os.path.basename(base_path).replace(" ","_").lower()[:20]
    collection = client.get_or_create_collection(name=f"libby_{folder_id}")
    existing_ids = set(collection.get()["ids"])
    new_chunks, new_embeddings, new_ids, new_metadata = [], [], [], []
    for root, dirs, files in os.walk(base_path):
        dirs[:] = [d for d in dirs if d not in ["libby_db","__pycache__"]]
        for file in files:
            if file.startswith("libby") or file == "libby_config.json": continue
            filepath = os.path.join(root, file)
            if file.endswith(".txt"):
                try:
                    with open(filepath, "r", encoding="utf-8") as f: content = f.read()
                    file_type = "txt"
                except: continue
            elif file.endswith(".xlsx"):
                try:
                    content = read_excel_text(filepath)
                    file_type = "xlsx"
                    load_excel_as_dataframe(filepath)
                except: continue
            elif file.endswith(".pdf"):
                try:
                    content = read_pdf(filepath)
                    file_type = "pdf"
                    if not content.strip(): continue
                except: continue
            else: continue
            for i, chunk in enumerate(chunk_text(content)):
                chunk_id = f"{filepath}::chunk{i}"
                if chunk_id in existing_ids: continue
                new_chunks.append(chunk)
                new_ids.append(chunk_id)
                new_metadata.append({"source": filepath, "chunk_index": i,
                                     "filename": file, "file_type": file_type})
    if new_chunks:
        new_embeddings = rag_model.encode(new_chunks).tolist()
        collection.add(documents=new_chunks, embeddings=new_embeddings,
                       ids=new_ids, metadatas=new_metadata)
    for m in collection.get(include=["metadatas"])["metadatas"]:
        src = m.get("source","")
        if src and src not in [f["source"] for f in known_files]:
            known_files.append({"source": src,
                                "filename": m.get("filename", os.path.basename(src)),
                                "file_type": m.get("file_type","txt")})

def retrieve(question):
    if not collection: return []
    qe = rag_model.encode([question]).tolist()
    results = collection.query(query_embeddings=qe, n_results=5,
                               include=["documents","metadatas","distances"])
    return [{"text": doc.strip(), "source": meta.get("source","Unknown"),
              "filename": meta.get("filename","Unknown"),
              "file_type": meta.get("file_type","txt")}
            for doc, meta, dist in zip(results["documents"][0],
                                       results["metadatas"][0],
                                       results["distances"][0]) if dist <= 1.4]

def ask_ollama(question, chunks, calc_results, history, mode, model, is_followup=False):
    system_prompt = PROMPT_EBI if mode == "ebi" else PROMPT_KNOWLEDGE

    # ── V9: If follow-up but no chunks found, retry with last topic ──
    if is_followup and not chunks and not calc_results and history:
        last_user_msgs = [m for r, m in history if r == "User"]
        if last_user_msgs:
            fallback_q = last_user_msgs[-1]
            chunks = retrieve(fallback_q)

    if not chunks and not calc_results:
        return ("Not found." if mode == "ebi" else
                "I couldn't find anything relevant. Try rephrasing."), []

    context_parts = []
    if calc_results: context_parts.extend(["=== CALCULATED DATA ==="] + calc_results)
    if chunks:
        context_parts.append("=== SOURCE DOCUMENTS ===")
        for c in chunks: context_parts.append(f"[{c['filename']}]\n{c['text']}")

    # Full row lookup for EBI follow-ups
    vague = ["that","it","those","the same","its","their","this","them"]
    if any(w in question.lower().split() for w in vague) and history:
        full_row = lookup_full_row(history)
        if full_row: context_parts.extend(["=== FULL RECORD (follow-up) ===", full_row])

    context = "\n\n".join(context_parts)
    sources = list(set([c["source"] for c in chunks])) if chunks else []

    # ── V9: Compact conversation history — last 4 exchanges only to save tokens ──
    history_block = ""
    if history:
        history_block = "\n--- RECENT CONVERSATION ---\n"
        for role, msg in history[-4:]:
            short_msg = msg[:200] if len(msg) > 200 else msg
            history_block += f"{role}: {short_msg}\n"
        history_block += "--- END ---\n"

    # ── V9: Explicit follow-up instruction injected into prompt ──
    followup_instruction = ""
    if is_followup:
        last_user_msgs  = [m for r, m in history if r == "User"]
        last_libby_msgs = [m for r, m in history if r == "Libby"]
        prev_q = last_user_msgs[-1]  if last_user_msgs  else ""
        prev_a = last_libby_msgs[-1] if last_libby_msgs else ""
        followup_instruction = (
            f"\n--- FOLLOW-UP CONTEXT ---\n"
            f"Previous question: {prev_q}\n"
            f"Previous answer: {prev_a[:200]}\n"
            f'Current request "{question}" continues the above topic.\n'
            f"Expand or continue from that answer using the documents below.\n"
            f"--- END FOLLOW-UP CONTEXT ---\n"
        )

    prompt = (
        f"{system_prompt}\n"
        f"{history_block}"
        f"{followup_instruction}"
        f"\n--- DATA ---\n"
        f"{context}\n"
        f"--- END ---\n\n"
        f"Question: {question}\n"
        f"Answer:"
    )

    try:
        max_tokens = 1500 if mode == "knowledge" else 800
        response = requests.post(
            "http://localhost:11434/api/generate",
            json={
                "model": model,
                "prompt": prompt,
                "stream": False,
                "options": {
                    "temperature": 0.1 if mode == "knowledge" else 0.0,
                    "top_p": 0.9,
                    "num_predict": max_tokens,
                    "num_ctx": 4096,
                    "stop": []
                }
            },
            timeout=180
        )
        response.raise_for_status()
        raw = response.json().get("response", "").strip()
        return raw.replace("```plaintext","").replace("```python","").replace("```","").strip(), sources
    except requests.exceptions.ConnectionError:
        return "⚠  Ollama not running. Open CMD and type: ollama serve", []
    except Exception as e:
        return f"⚠  Error: {str(e)}", []

# ─────────────────────────────────────────────
# REPORT ENGINE
# ─────────────────────────────────────────────
def find_best_dataframe(request):
    r = request.lower()
    hints = {
        "sales":     ["sales","revenue","monthly","rep","region","variance","target","units"],
        "inventory": ["inventory","stock","reorder","on hand","sku","category"],
        "customer":  ["customer","account","balance","credit","outstanding","ytd"],
        "purchase":  ["purchase","po","supplier","order","received"],
        "employee":  ["employee","salary","staff","department","hr"],
        "project":   ["project","budget","spent","timeline"],
        "product":   ["product","performance","margin","profit","sku"],
    }
    best_file, best_score = None, 0
    for filename, sheets in dataframes.items():
        fn = filename.lower()
        score = sum(2 for hk, kws in hints.items() if hk in fn for k in kws if k in r)
        score += sum(1 for hk, kws in hints.items() for k in kws if k in r and k in fn)
        if score > best_score: best_score, best_file = score, filename
    if not best_file and dataframes: best_file = list(dataframes.keys())[0]
    if best_file:
        sn = list(dataframes[best_file].keys())[0]
        return best_file, sn, dataframes[best_file][sn].copy()
    return None, None, None

def generate_excel_report(request, filepath, company_name):
    filename, sheet_name, df = find_best_dataframe(request)
    if df is None: return "No data available.", False
    r = request.lower()
    group_map = {"region":["Region"],"rep":["Sales Rep"],"supplier":["Supplier"],
                 "customer":["Customer Name","Customer"],"category":["Category"],
                 "month":["Month"],"status":["Status"],"product":["Product Name","Product"],
                 "department":["Department"],"lead":["Lead"]}
    group_col = next((find_column(df, cands) for kw, cands in group_map.items()
                      if kw in r and find_column(df, cands)), None)
    numeric_cols = df.select_dtypes(include="number").columns.tolist()
    if group_col and numeric_cols:
        df = df.groupby(group_col)[numeric_cols].sum().reset_index()
        numeric_cols = df.select_dtypes(include="number").columns.tolist()
    sort_asc = any(w in r for w in ["lowest to highest","ascending","smallest"])
    if numeric_cols:
        priority = ["Revenue","Total","Value","Amount","Cost","Balance","Salary","Budget","Profit"]
        sort_col = next((find_column(df,[p]) for p in priority
                         if find_column(df,[p]) in numeric_cols), numeric_cols[0])
        df = df.sort_values(by=sort_col, ascending=sort_asc).reset_index(drop=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    CHARCOAL, ROSE, WHITE, LIGHT, TOTAL_BG = "252525","c9956c","ffffff","faf7f4","f5e8de"
    def tb():
        s = Side(style="thin", color="d5c8bc")
        return Border(left=s, right=s, top=s, bottom=s)
    nc, lc = len(df.columns), get_column_letter(len(df.columns))
    ws.merge_cells(f"A1:{lc}1")
    ws["A1"] = company_name
    ws["A1"].font = Font(name="Segoe UI", bold=True, size=16, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=CHARCOAL)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36
    ws.merge_cells(f"A2:{lc}2")
    ws["A2"] = f"Report  •  {datetime.now().strftime('%B %d, %Y  %H:%M')}  •  {filename}"
    ws["A2"].font = Font(name="Segoe UI", size=10, italic=True, color=ROSE)
    ws["A2"].fill = PatternFill("solid", fgColor=CHARCOAL)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 6
    def is_cur(col): return any(w in col.lower() for w in ["price","cost","amount","revenue","sales","value","total","balance","limit","variance","salary","budget","spent","profit","margin","remaining"])
    for ci, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=4, column=ci, value=col_name)
        cell.font = Font(name="Segoe UI", bold=True, size=11, color=WHITE)
        cell.fill = PatternFill("solid", fgColor="8b5e3c")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = tb()
    ws.row_dimensions[4].height = 22
    num_col_names = df.select_dtypes(include="number").columns.tolist()
    last_row = 4
    for ri, (_, row) in enumerate(df.iterrows(), 5):
        bg = LIGHT if ri % 2 == 0 else WHITE
        for ci, (col_name, val) in enumerate(row.items(), 1):
            cell = ws.cell(row=ri, column=ci)
            if col_name in num_col_names and pd.notna(val):
                cell.value = val
                cell.number_format = '$#,##0.00' if is_cur(col_name) else '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.value = val
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.font = Font(name="Segoe UI", size=10)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = tb()
        ws.row_dimensions[ri].height = 18
        last_row = ri
    tr = last_row + 1
    for ci, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=tr, column=ci)
        if col_name in num_col_names:
            cell.value = df[col_name].sum()
            cell.number_format = '$#,##0.00' if is_cur(col_name) else '#,##0.00'
            cell.alignment = Alignment(horizontal="right")
        else:
            cell.value = "TOTAL" if ci == 1 else ""
        cell.font = Font(name="Segoe UI", bold=True, size=11, color=CHARCOAL)
        cell.fill = PatternFill("solid", fgColor=TOTAL_BG)
        s = Side(style="medium", color=CHARCOAL)
        cell.border = Border(left=s, right=s, top=s, bottom=s)
    ws.row_dimensions[tr].height = 22
    fr = tr + 2
    ws.merge_cells(f"A{fr}:{lc}{fr}")
    fc = ws.cell(row=fr, column=1,
                 value=f"Confidential  •  {company_name}  •  {datetime.now().strftime('%Y-%m-%d')}  •  Air-gapped")
    fc.font = Font(name="Segoe UI", size=9, italic=True, color=ROSE)
    fc.fill = PatternFill("solid", fgColor=CHARCOAL)
    fc.alignment = Alignment(horizontal="center")
    ws.row_dimensions[fr].height = 16
    for ci, col_name in enumerate(df.columns, 1):
        try: max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max() if not df.empty else 0)
        except: max_len = len(str(col_name))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len+4, 12), 40)
    wb.save(filepath)
    sd = "lowest to highest" if sort_asc else "highest to lowest"
    summary = f"✅  Report generated\n• {len(df):,} rows  •  Sorted {sd}"
    if group_col: summary += f"\n• Grouped by: {group_col}"
    return summary, True


# ─────────────────────────────────────────────
# MAIN APP
# ─────────────────────────────────────────────
class LibbyApp:
    def __init__(self, root):
        self.root           = root
        self.cfg            = load_config()
        self.T              = THEMES[self.cfg["theme"]]
        self.current_theme  = self.cfg["theme"]
        self.active_tab     = "knowledge"   # currently visible tab
        self.ready          = False
        self.dot_count      = 0
        self.animating      = False
        self._anim_msg      = "Thinking"
        self._settings_open = False

        # ── Each tab has its own independent state ──
        self.tabs = {
            "knowledge": {
                "widgets":  [],       # Chat bubble data
                "history":  [],       # Ollama conversation history
                "has_msgs": False,    # Dot indicator flag
            },
            "ebi": {
                "widgets":  [],
                "history":  [],
                "has_msgs": False,
            }
        }

        self.root.title("Libby V9 — Universal Knowledge System")
        self.root.geometry("1240x820")
        self.root.minsize(900, 640)
        self.root.configure(bg=self.T["BG_DARK"])

        self._build_ui()
        self._start_backend()

    # ─────────────────────────────────────────
    # BUILD UI
    # ─────────────────────────────────────────
    def _build_ui(self):
        T = self.T

        # ── Header ──────────────────────────
        self.header = tk.Frame(self.root, bg=T["BG_PANEL"], pady=14)
        self.header.pack(fill="x")

        self.left_h = tk.Frame(self.header, bg=T["BG_PANEL"])
        self.left_h.pack(side="left", padx=20)

        self.avatar_canvas = tk.Canvas(
            self.left_h, width=44, height=50,
            bg=T["BG_PANEL"], highlightthickness=0
        )
        self.avatar_canvas.pack(side="left", padx=(0, 14))
        self._draw_avatar()

        self.title_frame = tk.Frame(self.left_h, bg=T["BG_PANEL"])
        self.title_frame.pack(side="left")

        self.title_label = tk.Label(
            self.title_frame, text="Libby",
            bg=T["BG_PANEL"], fg=T["ACCENT"],
            font=("Segoe UI", 26, "bold")
        )
        self.title_label.pack(anchor="w")

        self.company_label = tk.Label(
            self.title_frame, text=self.cfg["company_name"],
            bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
            font=F_SUBTITLE
        )
        self.company_label.pack(anchor="w")

        self.right_frame = tk.Frame(self.header, bg=T["BG_PANEL"])
        self.right_frame.pack(side="right", padx=20)

        self.settings_btn = tk.Button(
            self.right_frame, text="⚙   Settings",
            bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
            activebackground=T["BG_INPUT"],
            font=F_BTN_SM, relief="flat",
            cursor="hand2", command=self._toggle_settings
        )
        self.settings_btn.pack(anchor="e", pady=(0, 6))

        self.status_label = tk.Label(
            self.right_frame, text="⏳  Loading...",
            bg=T["BG_PANEL"], fg=T["WARNING"],
            font=F_SMALL
        )
        self.status_label.pack(anchor="e")

        self.top_divider = tk.Frame(self.root, bg=T["BORDER"], height=1)
        self.top_divider.pack(fill="x")

        # ── Tab bar ─────────────────────────
        self.tab_bar = tk.Frame(self.root, bg=T["BG_DARK"], pady=0)
        self.tab_bar.pack(fill="x")
        self._build_tab_bar()

        # ── Body ────────────────────────────
        self.body = tk.Frame(self.root, bg=T["BG_DARK"])
        self.body.pack(fill="both", expand=True)

        # ── Sidebar ─────────────────────────
        self.sidebar = tk.Frame(self.body, bg=T["BG_SIDEBAR"], width=240)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)
        self._build_sidebar()

        self.side_divider = tk.Frame(self.body, bg=T["BORDER"], width=1)
        self.side_divider.pack(side="left", fill="y")

        # ── Chat area ───────────────────────
        self.chat_outer = tk.Frame(self.body, bg=T["BG_PANEL"])
        self.chat_outer.pack(side="left", fill="both", expand=True)

        self.chat_canvas = tk.Canvas(
            self.chat_outer, bg=T["BG_PANEL"],
            highlightthickness=0, bd=0
        )
        scrollbar = tk.Scrollbar(self.chat_outer, command=self.chat_canvas.yview)
        self.chat_canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.chat_canvas.pack(side="left", fill="both", expand=True)

        self.chat_inner = tk.Frame(self.chat_canvas, bg=T["BG_PANEL"])
        self.chat_window = self.chat_canvas.create_window(
            (0, 0), window=self.chat_inner, anchor="nw"
        )
        self.chat_inner.bind("<Configure>", self._on_frame_configure)
        self.chat_canvas.bind("<Configure>", self._on_canvas_configure)
        self.chat_canvas.bind("<Enter>", lambda e: self.chat_canvas.bind_all(
            "<MouseWheel>", self._on_mousewheel))
        self.chat_canvas.bind("<Leave>", lambda e: self.chat_canvas.unbind_all("<MouseWheel>"))

        self.bottom_divider = tk.Frame(self.root, bg=T["BORDER"], height=1)
        self.bottom_divider.pack(fill="x")

        # ── Input ────────────────────────────
        self.input_frame = tk.Frame(self.root, bg=T["BG_PANEL"], pady=12)
        self.input_frame.pack(fill="x", padx=16)

        self.input_box = tk.Text(
            self.input_frame, height=3,
            bg=T["BG_INPUT"], fg=T["TEXT_MAIN"],
            insertbackground=T["ACCENT"],
            font=F_INPUT, relief="flat", bd=0,
            padx=14, pady=10, wrap="word"
        )
        self.input_box.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.input_box.bind("<Return>", self._on_enter)
        self._set_placeholder()
        self.input_box.bind("<FocusIn>",  self._clear_placeholder)
        self.input_box.bind("<FocusOut>", self._restore_placeholder)

        self.btn_frame = tk.Frame(self.input_frame, bg=T["BG_PANEL"])
        self.btn_frame.pack(side="right")

        self.send_btn = tk.Button(
            self.btn_frame, text="Ask Libby",
            bg=T["ACCENT_DIM"], fg=T["TEXT_MAIN"],
            activebackground=T["ACCENT"],
            font=F_BTN, relief="flat", padx=18, pady=8,
            cursor="hand2", command=self._send_question
        )
        self.send_btn.pack(pady=(0, 6))

        self.clear_btn = tk.Button(
            self.btn_frame, text="Clear",
            bg=T["BTN_CLEAR"], fg=T["BTN_CLEAR_FG"],
            activebackground=T["BTN_CLEAR"],
            font=F_BTN_SM, relief="flat", padx=18, pady=5,
            cursor="hand2", command=self._clear_chat
        )
        self.clear_btn.pack()

        # ── Footer ───────────────────────────
        self.footer = tk.Frame(self.root, bg=T["BG_PANEL"], pady=5)
        self.footer.pack(fill="x")
        self.footer_label = tk.Label(
            self.footer,
            text="🔒  Air-gapped  •  All answers from your local knowledge base  •  No internet required",
            bg=T["BG_PANEL"], fg=T["TEXT_DIM"], font=F_TINY
        )
        self.footer_label.pack()

        # Show welcome message in active tab
        self._add_libby_message(self._welcome_message(), source=None, file_type=None)

    # ─────────────────────────────────────────
    # TAB BAR
    # ─────────────────────────────────────────
    def _build_tab_bar(self):
        T = self.T
        for w in self.tab_bar.winfo_children():
            w.destroy()

        tabs_def = [
            ("knowledge", "📚  Knowledge Assistant"),
            ("ebi",       "📊  Enterprise BI"),
        ]

        for tab_id, label in tabs_def:
            is_active = tab_id == self.active_tab
            has_msgs  = self.tabs[tab_id]["has_msgs"]

            tab_frame = tk.Frame(
                self.tab_bar,
                bg=T["BG_TAB_ACTIVE"] if is_active else T["BG_TAB_IDLE"],
                padx=0, pady=0
            )
            tab_frame.pack(side="left")

            # Active tab gets a rose gold top border line
            if is_active:
                tk.Frame(tab_frame, bg=T["BORDER_TAB"], height=3).pack(fill="x")
            else:
                tk.Frame(tab_frame, bg=T["BG_TAB_IDLE"], height=3).pack(fill="x")

            inner = tk.Frame(
                tab_frame,
                bg=T["BG_TAB_ACTIVE"] if is_active else T["BG_TAB_IDLE"],
                padx=16, pady=10
            )
            inner.pack()

            btn = tk.Button(
                inner,
                text=label,
                bg=T["BG_TAB_ACTIVE"] if is_active else T["BG_TAB_IDLE"],
                fg=T["TEXT_TAB_ACT"] if is_active else T["TEXT_TAB_IDLE"],
                font=("Segoe UI", 10, "bold" if is_active else "normal"),
                relief="flat", cursor="hand2",
                command=lambda tid=tab_id: self._switch_tab(tid)
            )
            btn.pack(side="left")

            # Dot indicator — small colored circle if tab has messages
            if has_msgs and not is_active:
                dot = tk.Label(
                    inner, text="●",
                    bg=T["BG_TAB_IDLE"],
                    fg=T["DOT_COLOR"],
                    font=("Segoe UI", 7)
                )
                dot.pack(side="left", padx=(6, 0))

        # Bottom border of tab bar
        tk.Frame(self.root, bg=T["BORDER"], height=1).pack(fill="x")

    def _switch_tab(self, tab_id):
        if tab_id == self.active_tab:
            return

        # Save current tab's chat widgets to its state
        self.tabs[self.active_tab]["widgets"] = list(self._get_current_widgets())

        # Switch
        self.active_tab = tab_id
        self._build_tab_bar()

        # Update send button label
        self.send_btn.config(
            text="Ask Libby" if tab_id == "knowledge" else "Query"
        )
        self._set_placeholder()

        # Reload the other tab's chat
        self._reload_chat()
        self._scroll_to_bottom()

    def _get_current_widgets(self):
        """Return stored widget data for current tab."""
        return self.tabs[self.active_tab]["widgets"]

    def _reload_chat(self):
        """Clear chat area and repopulate from active tab's stored widgets."""
        for w in self.chat_inner.winfo_children():
            w.destroy()

        stored = self.tabs[self.active_tab]["widgets"]
        if stored:
            for kind, text, source, file_type in stored:
                if kind == "user":
                    self._draw_user_bubble(text)
                else:
                    self._draw_libby_bubble(text, source, file_type)
        else:
            # Fresh tab — show welcome
            self._draw_libby_bubble(self._welcome_message(), None, None)

    # ─────────────────────────────────────────
    # SIDEBAR
    # ─────────────────────────────────────────
    def _build_sidebar(self):
        T = self.T
        self.sb_header = tk.Frame(self.sidebar, bg=T["BG_SIDEBAR"], pady=12)
        self.sb_header.pack(fill="x", padx=12)

        self.sidebar_title = tk.Label(
            self.sb_header, text="Knowledge Base",
            bg=T["BG_SIDEBAR"], fg=T["ACCENT"], font=F_LABEL_B
        )
        self.sidebar_title.pack(anchor="w")

        self.path_label = tk.Label(
            self.sb_header,
            text=f"📂  {os.path.basename(self.cfg['base_path'])}",
            bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"],
            font=F_TINY, wraplength=200, justify="left"
        )
        self.path_label.pack(anchor="w", pady=(2, 0))

        tk.Frame(self.sidebar, bg=T["BORDER"], height=1).pack(fill="x", padx=8)

        list_container = tk.Frame(self.sidebar, bg=T["BG_SIDEBAR"])
        list_container.pack(fill="both", expand=True)

        self.sidebar_canvas = tk.Canvas(
            list_container, bg=T["BG_SIDEBAR"],
            highlightthickness=0, bd=0
        )
        sb_scroll = tk.Scrollbar(list_container, orient="vertical",
                                 command=self.sidebar_canvas.yview)
        self.sidebar_canvas.configure(yscrollcommand=sb_scroll.set)
        sb_scroll.pack(side="right", fill="y")
        self.sidebar_canvas.pack(side="left", fill="both", expand=True)

        self.sidebar_list = tk.Frame(self.sidebar_canvas, bg=T["BG_SIDEBAR"])
        self.sidebar_list_window = self.sidebar_canvas.create_window(
            (0, 0), window=self.sidebar_list, anchor="nw"
        )
        self.sidebar_list.bind("<Configure>", lambda e: self.sidebar_canvas.configure(
            scrollregion=self.sidebar_canvas.bbox("all")))
        self.sidebar_canvas.bind("<Configure>", lambda e: self.sidebar_canvas.itemconfig(
            self.sidebar_list_window, width=e.width))
        self.sidebar_canvas.bind("<Enter>", lambda e: self.sidebar_canvas.bind_all(
            "<MouseWheel>", lambda ev: self.sidebar_canvas.yview_scroll(
                int(-1*(ev.delta/120)),"units")))
        self.sidebar_canvas.bind("<Leave>", lambda e: self.sidebar_canvas.unbind_all("<MouseWheel>"))

        tk.Frame(self.sidebar, bg=T["BORDER"], height=1).pack(fill="x", padx=8)
        self.sidebar_footer_label = tk.Label(
            self.sidebar, text="📄 txt   📊 xlsx   📕 pdf",
            bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"],
            font=F_TINY, pady=6
        )
        self.sidebar_footer_label.pack(side="bottom")

    def _refresh_sidebar(self):
        T = self.T
        for w in self.sidebar_list.winfo_children():
            w.destroy()
        for entry in known_files:
            filename  = entry["filename"]
            folder    = os.path.basename(os.path.dirname(entry["source"]))
            file_type = entry.get("file_type","txt")
            icon      = "📊" if file_type=="xlsx" else ("📕" if file_type=="pdf" else "📄")
            row = tk.Frame(self.sidebar_list, bg=T["BG_SIDEBAR"])
            row.pack(anchor="w", fill="x", pady=1, padx=10)
            lbl1 = tk.Label(row, text=f"{icon}  {filename}",
                            bg=T["BG_SIDEBAR"], fg=T["TEXT_MAIN"],
                            font=F_SMALL, wraplength=200, justify="left", anchor="w")
            lbl1.pack(anchor="w")
            lbl2 = tk.Label(row, text=f"     {folder}",
                            bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"],
                            font=F_TINY, anchor="w")
            lbl2.pack(anchor="w", pady=(0, 3))
            for widget in [row, lbl1, lbl2]:
                widget.bind("<Enter>", lambda e: self.sidebar_canvas.bind_all(
                    "<MouseWheel>", lambda ev: self.sidebar_canvas.yview_scroll(
                        int(-1*(ev.delta/120)),"units")))
                widget.bind("<Leave>", lambda e: self.sidebar_canvas.unbind_all("<MouseWheel>"))

    # ─────────────────────────────────────────
    # SETTINGS
    # ─────────────────────────────────────────
    def _toggle_settings(self):
        if self._settings_open: self._close_settings()
        else: self._open_settings()

    def _open_settings(self):
        T = self.T
        self._settings_open = True
        self.settings_btn.config(text="✕  Close")

        self.settings_win = tk.Toplevel(self.root)
        self.settings_win.title("Libby Settings")
        self.settings_win.resizable(False, False)
        self.settings_win.configure(bg=T["BG_SETTINGS"])
        self.settings_win.grab_set()
        self.settings_win.protocol("WM_DELETE_WINDOW", self._close_settings)

        self.root.update_idletasks()
        rx = self.root.winfo_x() + (self.root.winfo_width()  // 2) - 220
        ry = self.root.winfo_y() + (self.root.winfo_height() // 2) - 180
        self.settings_win.geometry(f"440x440+{rx}+{ry}")

        p = tk.Frame(self.settings_win, bg=T["BG_SETTINGS"])
        p.pack(fill="both", expand=True, padx=24, pady=20)

        tk.Label(p, text="Settings", bg=T["BG_SETTINGS"], fg=T["ACCENT"],
                 font=("Segoe UI", 14, "bold")).grid(
                     row=0, column=0, columnspan=3, sticky="w", pady=(0, 16))

        tk.Label(p, text="Knowledge Folder", bg=T["BG_SETTINGS"],
                 fg=T["TEXT_MAIN"], font=F_LABEL_B).grid(
                     row=1, column=0, columnspan=3, sticky="w", pady=(0, 4))
        self.folder_var = tk.StringVar(value=self.cfg["base_path"])
        tk.Entry(p, textvariable=self.folder_var, bg=T["BG_INPUT"],
                 fg=T["TEXT_MAIN"], insertbackground=T["ACCENT"],
                 font=F_SMALL, relief="flat", width=30).grid(
                     row=2, column=0, columnspan=2, sticky="w", ipady=4, pady=(0, 14))
        tk.Button(p, text="Browse", bg=T["ACCENT_DIM"], fg=T["TEXT_MAIN"],
                  font=F_BTN_SM, relief="flat", padx=8, cursor="hand2",
                  command=self._browse_folder).grid(
                      row=2, column=2, sticky="w", padx=(6,0), pady=(0,14))

        tk.Label(p, text="Theme", bg=T["BG_SETTINGS"],
                 fg=T["TEXT_MAIN"], font=F_LABEL_B).grid(
                     row=3, column=0, columnspan=3, sticky="w", pady=(0, 4))
        self.theme_var = tk.StringVar(value=self.cfg["theme"])
        tk.Radiobutton(p, text="🌙  Dark", variable=self.theme_var, value="dark",
                       bg=T["BG_SETTINGS"], fg=T["TEXT_MAIN"],
                       selectcolor=T["BG_INPUT"], activebackground=T["BG_SETTINGS"],
                       font=F_LABEL).grid(row=4, column=0, sticky="w", pady=(0,14))
        tk.Radiobutton(p, text="☀️  Light", variable=self.theme_var, value="light",
                       bg=T["BG_SETTINGS"], fg=T["TEXT_MAIN"],
                       selectcolor=T["BG_INPUT"], activebackground=T["BG_SETTINGS"],
                       font=F_LABEL).grid(row=4, column=1, sticky="w", pady=(0,14))

        tk.Label(p, text="Company Name", bg=T["BG_SETTINGS"],
                 fg=T["TEXT_MAIN"], font=F_LABEL_B).grid(
                     row=5, column=0, columnspan=3, sticky="w", pady=(0,4))
        self.company_var = tk.StringVar(value=self.cfg["company_name"])
        tk.Entry(p, textvariable=self.company_var, bg=T["BG_INPUT"],
                 fg=T["TEXT_MAIN"], insertbackground=T["ACCENT"],
                 font=F_LABEL, relief="flat", width=32).grid(
                     row=6, column=0, columnspan=3, sticky="w", ipady=4, pady=(0,20))

        tk.Frame(p, bg=T["BORDER"], height=1).grid(
            row=7, column=0, columnspan=3, sticky="ew", pady=(0,14))

        tk.Label(p, text="AI Model", bg=T["BG_SETTINGS"],
                 fg=T["TEXT_MAIN"], font=F_LABEL_B).grid(
                     row=8, column=0, columnspan=3, sticky="w", pady=(0,4))
        self.model_var = tk.StringVar(value=self.cfg.get("ollama_model","llama3.2"))
        model_options = ["llama3.2", "phi3:mini", "mistral", "llama3.1:8b"]
        if self.model_var.get() not in model_options:
            model_options.insert(0, self.model_var.get())
        tk.OptionMenu(p, self.model_var, *model_options).grid(
            row=9, column=0, columnspan=2, sticky="w", pady=(0,20))

        tk.Frame(p, bg=T["BORDER"], height=1).grid(
            row=10, column=0, columnspan=3, sticky="ew", pady=(0,14))

        tk.Button(p, text="Save & Apply", bg=T["ACCENT_DIM"], fg=T["TEXT_MAIN"],
                  activebackground=T["ACCENT"], font=F_BTN, relief="flat",
                  padx=16, pady=6, cursor="hand2",
                  command=self._save_settings).grid(row=11, column=0, sticky="w")
        tk.Button(p, text="Cancel", bg=T["BG_INPUT"], fg=T["TEXT_DIM"],
                  font=F_BTN_SM, relief="flat", padx=12, pady=6, cursor="hand2",
                  command=self._close_settings).grid(row=11, column=1, sticky="w", padx=(8,0))

    def _close_settings(self):
        self._settings_open = False
        self.settings_btn.config(text="⚙   Settings")
        if hasattr(self, "settings_win") and self.settings_win.winfo_exists():
            self.settings_win.destroy()

    def _browse_folder(self):
        folder = filedialog.askdirectory(title="Select Knowledge Base Folder",
                                         initialdir=self.cfg["base_path"])
        if folder: self.folder_var.set(folder)

    def _save_settings(self):
        new_path    = self.folder_var.get().strip()
        new_theme   = self.theme_var.get()
        new_company = self.company_var.get().strip() or "Libby Intelligence"
        new_model   = self.model_var.get().strip() or "llama3.2"
        if not os.path.exists(new_path):
            messagebox.showerror("Invalid Folder", f"Folder not found:\n{new_path}")
            return
        changed_path  = new_path  != self.cfg["base_path"]
        changed_theme = new_theme != self.cfg["theme"]
        self.cfg.update({"base_path": new_path, "theme": new_theme,
                         "company_name": new_company, "ollama_model": new_model})
        save_config(self.cfg)
        self._close_settings()
        if changed_theme:
            self.current_theme = new_theme
            self.T = THEMES[new_theme]
            self._apply_theme()
        self.company_label.config(text=new_company)
        self.path_label.config(text=f"📂  {os.path.basename(new_path)}")
        if changed_path:
            self.ready = False
            for w in self.sidebar_list.winfo_children():
                w.destroy()
            self.status_label.config(text="⏳  Reloading...", fg=self.T["WARNING"])
            self._add_libby_message(f"Knowledge folder updated.\nLoading: {new_path}",
                                    source=None, file_type=None)
            self._start_backend(fresh=True)
        else:
            count = collection.count() if collection else 0
            self.status_label.config(
                text=f"✅  {count} chunks  •  {new_model}  •  Offline",
                fg=self.T["SUCCESS"]
            )
            messagebox.showinfo("Saved", f"✅  Settings saved!\nModel: {new_model}")

    # ─────────────────────────────────────────
    # AVATAR
    # ─────────────────────────────────────────
    def _draw_avatar(self):
        c = self.avatar_canvas
        T = self.T
        c.delete("all")
        bw, bh, ox, oy = 32, 46, 6, 2
        c.create_polygon(ox+2, oy+2, ox+bw+2, oy+2, ox+bw+2, oy+bh+2,
                         ox+bw//2+2, oy+bh-10+2, ox+2, oy+bh+2,
                         fill=T["BORDER"], outline="")
        c.create_polygon(ox, oy, ox+bw, oy, ox+bw, oy+bh,
                         ox+bw//2, oy+bh-10, ox, oy+bh,
                         fill=T["ACCENT_DIM"], outline="")
        c.create_rectangle(ox, oy, ox+bw, oy+6, fill=T["ACCENT"], outline="")
        for y in [oy+14, oy+20, oy+26, oy+32]:
            lw = bw-8 if y < oy+30 else bw-14
            c.create_line(ox+4, y, ox+4+lw, y, fill=T["ACCENT_BRIGHT"], width=1.5)

    # ─────────────────────────────────────────
    # THEME
    # ─────────────────────────────────────────
    def _apply_theme(self):
        T = self.T
        self.root.configure(bg=T["BG_DARK"])
        self.header.configure(bg=T["BG_PANEL"])
        self.left_h.configure(bg=T["BG_PANEL"])
        self.avatar_canvas.configure(bg=T["BG_PANEL"])
        self._draw_avatar()
        self.title_frame.configure(bg=T["BG_PANEL"])
        self.title_label.configure(bg=T["BG_PANEL"], fg=T["ACCENT"])
        self.company_label.configure(bg=T["BG_PANEL"], fg=T["TEXT_DIM"])
        self.right_frame.configure(bg=T["BG_PANEL"])
        self.settings_btn.configure(bg=T["BG_PANEL"], fg=T["TEXT_DIM"])
        self.status_label.configure(bg=T["BG_PANEL"])
        self.top_divider.configure(bg=T["BORDER"])
        self.tab_bar.configure(bg=T["BG_DARK"])
        self._build_tab_bar()
        self.body.configure(bg=T["BG_DARK"])
        self.sidebar.configure(bg=T["BG_SIDEBAR"])
        self.sb_header.configure(bg=T["BG_SIDEBAR"])
        for w in self.sidebar.winfo_children():
            try: w.configure(bg=T["BG_SIDEBAR"])
            except: pass
        self.sidebar_title.configure(bg=T["BG_SIDEBAR"], fg=T["ACCENT"])
        self.path_label.configure(bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"])
        self.sidebar_canvas.configure(bg=T["BG_SIDEBAR"])
        self.sidebar_list.configure(bg=T["BG_SIDEBAR"])
        self.sidebar_footer_label.configure(bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"])
        self.side_divider.configure(bg=T["BORDER"])
        self.chat_outer.configure(bg=T["BG_PANEL"])
        self.chat_canvas.configure(bg=T["BG_PANEL"])
        self.chat_inner.configure(bg=T["BG_PANEL"])
        self.bottom_divider.configure(bg=T["BORDER"])
        self.input_frame.configure(bg=T["BG_PANEL"])
        self.btn_frame.configure(bg=T["BG_PANEL"])
        self.input_box.configure(bg=T["BG_INPUT"], fg=T["TEXT_MAIN"],
                                 insertbackground=T["ACCENT"])
        self.send_btn.configure(bg=T["ACCENT_DIM"], fg=T["TEXT_MAIN"])
        self.clear_btn.configure(bg=T["BTN_CLEAR"], fg=T["BTN_CLEAR_FG"])
        self.footer.configure(bg=T["BG_PANEL"])
        self.footer_label.configure(bg=T["BG_PANEL"], fg=T["TEXT_DIM"])
        self.status_label.configure(
            fg=T["SUCCESS"] if self.current_theme=="light" else T["WARNING"])
        self._refresh_sidebar()
        self._reload_chat()

    # ─────────────────────────────────────────
    # CHAT BUBBLES
    # ─────────────────────────────────────────
    def _draw_user_bubble(self, text):
        T     = self.T
        outer = tk.Frame(self.chat_inner, bg=T["BG_PANEL"], pady=8)
        outer.pack(fill="x", padx=20)
        tk.Label(outer, text="You", bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
                 font=F_SMALL).pack(anchor="e", padx=4, pady=(0, 2))
        bubble = tk.Frame(outer, bg=T["BG_BUBBLE_U"], padx=18, pady=12)
        bubble.pack(anchor="e")
        txt = tk.Text(bubble, height=3, bg=T["BG_BUBBLE_U"], fg=T["TEXT_BUBBLE_U"],
                     font=F_CHAT, wrap="word", relief="flat", bd=0, padx=0, pady=0)
        txt.insert("1.0", text)
        txt.update_idletasks()
        line_count = int(txt.index("end-1c").split(".")[0])
        txt.config(height=max(2, line_count), state="disabled", cursor="arrow")
        txt.pack(anchor="w", fill="both", expand=True)

    def _draw_libby_bubble(self, text, source=None, file_type=None, entry_id=None):
        T     = self.T
        outer = tk.Frame(self.chat_inner, bg=T["BG_PANEL"], pady=8)
        outer.pack(fill="x", padx=20)
        tk.Label(outer, text="Libby", bg=T["BG_PANEL"], fg=T["ACCENT"],
                 font=F_LABEL_B).pack(anchor="w", padx=4, pady=(0, 2))
        bubble = tk.Frame(outer, bg=T["BG_BUBBLE_L"], padx=18, pady=12)
        bubble.pack(anchor="w")
        txt = tk.Text(bubble, height=3, bg=T["BG_BUBBLE_L"], fg=T["TEXT_BUBBLE_L"],
                     font=F_CHAT, wrap="word", relief="flat", bd=0, padx=0, pady=0)
        txt.insert("1.0", text)
        txt.update_idletasks()
        line_count = int(txt.index("end-1c").split(".")[0])
        txt.config(height=max(2, line_count), state="disabled", cursor="arrow")
        txt.pack(anchor="w", fill="both", expand=True)
        if source:
            icon = "📊" if file_type=="xlsx" else ("📕" if file_type=="pdf" else "📄")
            tk.Label(bubble, text=f"{icon}  {os.path.basename(source)}",
                     bg=T["BG_BUBBLE_L"], fg=T["TEXT_SOURCE"],
                     font=F_SMALL, wraplength=720,
                     justify="left").pack(anchor="w", pady=(8, 0))
        # ── Feedback buttons ──
        if entry_id:
            fb_frame = tk.Frame(bubble, bg=T["BG_BUBBLE_L"])
            fb_frame.pack(anchor="w", pady=(8, 0))
            thumb_up = tk.Button(
                fb_frame, text="👍",
                bg=T["BG_BUBBLE_L"], fg=T["TEXT_DIM"],
                activebackground=T["BG_INPUT"],
                font=F_TINY, relief="flat", cursor="hand2",
                command=lambda eid=entry_id: self._rate_answer(eid, "positive", thumb_up, thumb_dn)
            )
            thumb_up.pack(side="left", padx=(0, 4))
            thumb_dn = tk.Button(
                fb_frame, text="👎",
                bg=T["BG_BUBBLE_L"], fg=T["TEXT_DIM"],
                activebackground=T["BG_INPUT"],
                font=F_TINY, relief="flat", cursor="hand2",
                command=lambda eid=entry_id: self._rate_answer(eid, "negative", thumb_up, thumb_dn)
            )
            thumb_dn.pack(side="left")

    def _rate_answer(self, entry_id, rating, thumb_up, thumb_dn):
        self._log_feedback(entry_id, rating)
        T = self.T
        if rating == "positive":
            thumb_up.config(fg=T["ACCENT"])
            thumb_dn.config(fg=T["TEXT_DIM"])
        else:
            thumb_dn.config(fg=T["BTN_CLEAR_FG"])
            thumb_up.config(fg=T["TEXT_DIM"])

    def _add_user_message(self, text):
        self.tabs[self.active_tab]["widgets"].append(("user", text, None, None))
        self.tabs[self.active_tab]["has_msgs"] = True
        self._draw_user_bubble(text)
        self._build_tab_bar()   # Refresh dot indicator

    def _add_libby_message(self, text, source=None, file_type=None, entry_id=None):
        self.tabs[self.active_tab]["widgets"].append(("libby", text, source, file_type))
        self._draw_libby_bubble(text, source, file_type, entry_id=entry_id)

    # ─────────────────────────────────────────
    # AUDIT LOG
    # ─────────────────────────────────────────
    def _write_log_entry(self, tab_id, question, answer, sources):
        log_file = os.path.join(SCRIPT_DIR, "feedback_log.json")
        entry_id = str(uuid.uuid4())[:6]
        entry = {
            "id":        entry_id,
            "timestamp": datetime.now().isoformat(),
            "tab":       tab_id,
            "question":  question,
            "answer":    answer,
            "sources":   sources,
            "rating":    None
        }
        log = []
        if os.path.exists(log_file):
            try:
                with open(log_file, "r") as f:
                    log = json.load(f)
            except (json.JSONDecodeError, Exception):
                log = []
        log.append(entry)
        try:
            with open(log_file, "w") as f:
                json.dump(log, f, indent=2)
            print(f"Log entry written: {entry_id}")
        except Exception as e:
            print(f"Log write error: {e}")
        return entry_id

    def _log_feedback(self, entry_id, rating):
        if not entry_id:
            return
        log_file = os.path.join(SCRIPT_DIR, "feedback_log.json")
        if not os.path.exists(log_file):
            print("No log file found")
            return
        try:
            with open(log_file, "r") as f:
                log = json.load(f)
            for entry in log:
                if entry.get("id") == entry_id:
                    entry["rating"] = rating
                    break
            with open(log_file, "w") as f:
                json.dump(log, f, indent=2)
            print(f"Rating updated: {entry_id} → {rating}")
        except Exception as e:
            print(f"Log feedback error: {e}")

    # ─────────────────────────────────────────
    # THINKING ANIMATION
    # ─────────────────────────────────────────
    def _add_thinking_frame(self, is_report=False):
        T     = self.T
        frame = tk.Frame(self.chat_inner, bg=T["BG_PANEL"], pady=8)
        frame.pack(fill="x", padx=20)
        tk.Label(frame, text="Libby", bg=T["BG_PANEL"], fg=T["ACCENT"],
                 font=F_LABEL_B).pack(anchor="w", padx=4, pady=(0, 2))
        self._anim_msg = "Building report" if is_report else "Thinking"
        self.thinking_label = tk.Label(
            frame, text=self._anim_msg,
            bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
            font=("Segoe UI", 10, "italic")
        )
        self.thinking_label.pack(anchor="w", padx=8)
        self.animating = True
        self._animate_dots()
        self._scroll_to_bottom()
        return frame

    def _animate_dots(self):
        if not self.animating: return
        dots = "." * (self.dot_count % 4)
        self.thinking_label.config(text=f"{self._anim_msg}{dots}")
        self.dot_count += 1
        self.root.after(300, self._animate_dots)

    def _stop_animation(self, frame):
        self.animating = False
        frame.destroy()

    # ─────────────────────────────────────────
    # CLEAR — clears only the active tab
    # ─────────────────────────────────────────
    def _clear_chat(self):
        for w in self.chat_inner.winfo_children():
            w.destroy()
        self.tabs[self.active_tab]["widgets"]  = []
        self.tabs[self.active_tab]["history"]  = []
        self.tabs[self.active_tab]["has_msgs"] = False
        self._build_tab_bar()
        self._draw_libby_bubble("Session cleared. Ready for questions.",
                                None, None)

    # ─────────────────────────────────────────
    # SCROLL
    # ─────────────────────────────────────────
    def _on_frame_configure(self, event):
        self.chat_canvas.configure(scrollregion=self.chat_canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.chat_canvas.itemconfig(self.chat_window, width=event.width)

    def _on_mousewheel(self, event):
        self.chat_canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    def _scroll_to_bottom(self):
        self.root.after(50, lambda: self.chat_canvas.yview_moveto(1.0))

    # ─────────────────────────────────────────
    # INPUT
    # ─────────────────────────────────────────
    def _placeholder_text(self):
        return ("Ask me anything from your knowledge base..."
                if self.active_tab == "knowledge"
                else "Ask a data question or type 'generate report ...'")

    def _set_placeholder(self):
        self.input_box.delete("1.0", "end")
        self.input_box.insert("1.0", self._placeholder_text())
        self.input_box.config(fg=self.T["TEXT_DIM"])

    def _clear_placeholder(self, event):
        if self.input_box.get("1.0", "end-1c") == self._placeholder_text():
            self.input_box.delete("1.0", "end")
            self.input_box.config(fg=self.T["TEXT_MAIN"])

    def _restore_placeholder(self, event):
        if not self.input_box.get("1.0", "end-1c").strip():
            self._set_placeholder()

    def _on_enter(self, event):
        if not event.state & 0x1:
            self._send_question()
            return "break"

    # ─────────────────────────────────────────
    # SEND & QUERY
    # ─────────────────────────────────────────
    def _send_question(self):
        if not self.ready: return
        question = self.input_box.get("1.0", "end-1c").strip()
        if not question or question == self._placeholder_text(): return

        self.input_box.delete("1.0", "end")
        self.input_box.config(fg=self.T["TEXT_MAIN"])
        self._add_user_message(question)

        is_report = self.active_tab == "ebi" and "generate report" in question.lower()
        thinking_frame = self._add_thinking_frame(is_report=is_report)
        self.send_btn.config(state="disabled", bg=self.T["BG_INPUT"])
        self.input_box.config(state="disabled")

        if is_report:
            self.root.after(0, lambda: self._prompt_save_report(question, thinking_frame))
        else:
            tab_history = self.tabs[self.active_tab]["history"]
            active_tab  = self.active_tab

            def run_query():
                # ── V9 Conversational Context Engine ──────────────────
                # Step 1: Detect if this is a follow-up question
                FOLLOWUP_TRIGGERS = [
                    # Vague pronouns
                    "that","it","those","the same","its","their","this","them","these",
                    # Short commands that imply continuation
                    "provide","explain","elaborate","expand","continue","more","detail",
                    "details","steps","list","show","give","tell","describe","summarize",
                    "summarise","how","why","what about","and","also","further","go on",
                    "keep going","next","then what","after that"
                ]
                q_lower   = question.lower().strip()
                q_words   = q_lower.split()
                is_short  = len(q_words) <= 5
                has_trigger = any(
                    q_lower == t or q_lower.startswith(t + " ") or (" " + t + " ") in q_lower
                    for t in FOLLOWUP_TRIGGERS
                )
                is_followup = (is_short or has_trigger) and len(tab_history) > 0

                # Step 2: Build enriched search query using prior context
                search_q = question
                if is_followup and tab_history:
                    last_user_msgs  = [m for r, m in tab_history if r == "User"]
                    last_libby_msgs = [m for r, m in tab_history if r == "Libby"]
                    last_topic      = last_user_msgs[-1]  if last_user_msgs  else ""
                    last_answer     = last_libby_msgs[-1] if last_libby_msgs else ""
                    # Extract key nouns from last answer for richer retrieval
                    topic_words = [
                        w for w in last_topic.split()
                        if len(w) > 3 and w.lower() not in
                        {"what","where","when","which","how","does","the","and","for",
                         "with","from","that","this","are","can","you","tell","me","about"}
                    ]
                    topic_hint = " ".join(topic_words[:6])
                    search_q   = f"{topic_hint} {question}".strip()
                    print(f"[V9] Follow-up detected. Enriched query: '{search_q}'")

                # Step 3: Retrieve and calculate as normal
                chunks       = retrieve(search_q)
                calc_results = get_calculations(search_q) if active_tab == "ebi" else []

                # Step 4: Pass full enriched context + history to Ollama
                answer, sources = ask_ollama(
                    question, chunks, calc_results,
                    tab_history, active_tab, self.cfg["ollama_model"],
                    is_followup=is_followup
                )
                source_str = ", ".join(sources) if sources else None
                file_type  = chunks[0]["file_type"] if chunks else None
                tab_history.append(("User", question))
                tab_history.append(("Libby", answer))
                if len(tab_history) > 20:
                    self.tabs[active_tab]["history"] = tab_history[-20:]
                self.root.after(0, lambda: self._show_answer(
                    answer, source_str, file_type, thinking_frame,
                    question, active_tab, sources
                ))
            threading.Thread(target=run_query, daemon=True).start()

    def _prompt_save_report(self, question, thinking_frame):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath  = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files","*.xlsx")],
            initialfile=f"Libby_Report_{timestamp}.xlsx",
            title="Save Report As"
        )
        if not filepath:
            self._stop_animation(thinking_frame)
            self._add_libby_message("Report cancelled.", source=None, file_type=None)
            self._re_enable_input()
            return
        def do_generate():
            summary, success = generate_excel_report(
                question, filepath, self.cfg["company_name"])
            self.root.after(0, lambda: self._show_report_result(
                summary, success, filepath, thinking_frame))
        threading.Thread(target=do_generate, daemon=True).start()

    def _show_report_result(self, summary, success, filepath, thinking_frame):
        self._stop_animation(thinking_frame)
        msg = f"{summary}\n📁  Saved: {os.path.basename(filepath)}" if success else summary
        self._add_libby_message(msg, source=filepath if success else None,
                                file_type="xlsx" if success else None)
        self._scroll_to_bottom()
        self._re_enable_input()

    def _show_answer(self, answer, source, file_type, thinking_frame, question, tab_id, sources):
        self._stop_animation(thinking_frame)
        entry_id = self._write_log_entry(
            tab_id=tab_id,
            question=question,
            answer=answer,
            sources=sources if sources else []
        )
        self._add_libby_message(answer, source=source, file_type=file_type, entry_id=entry_id)
        self._scroll_to_bottom()
        self._re_enable_input()

    def _re_enable_input(self):
        self.send_btn.config(state="normal", bg=self.T["ACCENT_DIM"])
        self.input_box.config(state="normal")
        self.input_box.focus()
        self.root.update_idletasks()
        self.root.update()

    def _welcome_message(self):
        if self.active_tab == "knowledge":
            return ("Welcome to Libby V9  —  Knowledge Assistant\n\n"
                    "Ask me anything from your loaded documents.\n"
                    "Switch to the Enterprise BI tab for data and reports.")
        return ("Welcome to Libby V9  —  Enterprise BI\n\n"
                "Ask data questions or type 'generate report...' to export Excel.\n"
                "Switch to the Knowledge Assistant tab for document Q&A.")

    # ─────────────────────────────────────────
    # BACKEND
    # ─────────────────────────────────────────
    def _start_backend(self, fresh=False):
        def load():
            load_backend(self.cfg["base_path"], fresh=fresh)
            self.ready = True
            count = collection.count() if collection else 0
            self.root.after(0, lambda: self.status_label.config(
                text=f"✅  {count} chunks  •  {self.cfg['ollama_model']}  •  Offline",
                fg=self.T["SUCCESS"]
            ))
            self.root.after(0, self._refresh_sidebar)
        threading.Thread(target=load, daemon=True).start()


# ─────────────────────────────────────────────
# LAUNCH
# ─────────────────────────────────────────────
if __name__ == "__main__":
    root = tk.Tk()
    app  = LibbyApp(root)
    root.mainloop()