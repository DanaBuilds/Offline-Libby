import tkinter as tk
import threading
import os
import requests
import chromadb
import openpyxl
from sentence_transformers import SentenceTransformer

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
BASE_PATH       = "C:\\RAG Test"
DB_PATH         = "C:\\RAG Test\\rag_db"
COLLECTION_NAME = "knowledge"
TOP_K           = 3
MIN_SCORE       = 1.2
OLLAMA_URL      = "http://localhost:11434/api/generate"
OLLAMA_MODEL    = "mistral"

# ─────────────────────────────────────────────
# THEMES
# ─────────────────────────────────────────────
THEMES = {
    "dark": {
        "BG_DARK":      "#0f1117",
        "BG_PANEL":     "#1a1d27",
        "BG_INPUT":     "#12151f",
        "BG_BUBBLE_U":  "#1e3a5f",
        "BG_BUBBLE_L":  "#1a2535",
        "BG_SIDEBAR":   "#13161f",
        "ACCENT":       "#4a9eff",
        "ACCENT_DIM":   "#2a5a99",
        "TEXT_MAIN":    "#e8eaf0",
        "TEXT_DIM":     "#6b7280",
        "TEXT_SOURCE":  "#4a9eff",
        "SUCCESS":      "#34d399",
        "WARNING":      "#fbbf24",
        "BORDER":       "#2a2d3e",
        "BTN_CLEAR":    "#3a1f1f",
        "BTN_CLEAR_FG": "#ff6b6b",
    },
    "light": {
        "BG_DARK":      "#f0f2f8",
        "BG_PANEL":     "#ffffff",
        "BG_INPUT":     "#e8eaf0",
        "BG_BUBBLE_U":  "#d0e8ff",
        "BG_BUBBLE_L":  "#f0f4ff",
        "BG_SIDEBAR":   "#e4e7f0",
        "ACCENT":       "#2563eb",
        "ACCENT_DIM":   "#1d4ed8",
        "TEXT_MAIN":    "#1a1d27",
        "TEXT_DIM":     "#6b7280",
        "TEXT_SOURCE":  "#2563eb",
        "SUCCESS":      "#059669",
        "WARNING":      "#d97706",
        "BORDER":       "#cbd5e1",
        "BTN_CLEAR":    "#fee2e2",
        "BTN_CLEAR_FG": "#dc2626",
    }
}

# ─────────────────────────────────────────────
# RAG BACKEND
# ─────────────────────────────────────────────
rag_model   = None
collection  = None
known_files = []

def chunk_text(text, chunk_size=400, overlap=80):
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunks.append(text[start:end].strip())
        start += chunk_size - overlap
    return [c for c in chunks if len(c) > 50]

def read_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    all_text = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_lines = [f"Sheet: {sheet_name}"]
        for row in ws.iter_rows(values_only=True):
            cleaned = [str(cell).strip() if cell is not None else "" for cell in row]
            if any(cleaned):
                sheet_lines.append(" | ".join(cleaned))
        if len(sheet_lines) > 1:
            all_text.append("\n".join(sheet_lines))
    wb.close()
    return "\n\n".join(all_text)

def load_backend():
    global rag_model, collection, known_files

    print(f"Loading from BASE_PATH: {BASE_PATH}")
    print(f"Database path: {DB_PATH}")

    rag_model  = SentenceTransformer("all-MiniLM-L6-v2")
    client     = chromadb.PersistentClient(path=DB_PATH)
    collection = client.get_or_create_collection(name=COLLECTION_NAME)

    existing_ids = set(collection.get()["ids"])
    new_chunks, new_embeddings, new_ids, new_metadata = [], [], [], []

    for root, dirs, files in os.walk(BASE_PATH):
        dirs[:] = [d for d in dirs if "rag_db" not in d]
        for file in files:
            filepath = os.path.join(root, file)
            print(f"Found file: {filepath}")

            if file.endswith(".txt"):
                try:
                    with open(filepath, "r", encoding="utf-8") as f:
                        content = f.read()
                    file_type = "txt"
                except Exception as e:
                    print(f"Error reading {file}: {e}")
                    continue

            elif file.endswith(".xlsx"):
                try:
                    content = read_excel(filepath)
                    file_type = "xlsx"
                except Exception as e:
                    print(f"Error reading {file}: {e}")
                    continue
            else:
                continue

            chunks = chunk_text(content)
            print(f"  → {len(chunks)} chunks from {file}")

            for i, chunk in enumerate(chunks):
                chunk_id = f"{filepath}::chunk{i}"
                if chunk_id in existing_ids:
                    continue
                new_chunks.append(chunk)
                new_ids.append(chunk_id)
                new_metadata.append({
                    "source":      filepath,
                    "chunk_index": i,
                    "filename":    file,
                    "file_type":   file_type
                })

    print(f"Total new chunks to add: {len(new_chunks)}")

    if new_chunks:
        print("Embedding chunks...")
        new_embeddings = rag_model.encode(new_chunks).tolist()
        collection.add(
            documents=new_chunks,
            embeddings=new_embeddings,
            ids=new_ids,
            metadatas=new_metadata
        )
        print(f"Successfully added {len(new_chunks)} chunks!")
    else:
        print("No new chunks to add.")

    # Build sidebar list
    all_meta = collection.get(include=["metadatas"])["metadatas"]
    seen = set()
    for m in all_meta:
        src = m.get("source", "")
        if src and src not in seen:
            seen.add(src)
            known_files.append({
                "source":    src,
                "filename":  m.get("filename", os.path.basename(src)),
                "file_type": m.get("file_type", "txt")
            })

def retrieve(question):
    query_embedding = rag_model.encode([question]).tolist()
    results = collection.query(
        query_embeddings=query_embedding,
        n_results=TOP_K,
        include=["documents", "metadatas", "distances"]
    )
    filtered = []
    for doc, meta, dist in zip(
        results["documents"][0],
        results["metadatas"][0],
        results["distances"][0]
    ):
        if dist <= MIN_SCORE:
            filtered.append({
                "text":      doc.strip(),
                "source":    meta.get("source", "Unknown"),
                "filename":  meta.get("filename", "Unknown"),
                "file_type": meta.get("file_type", "txt")
            })
    return filtered

def ask_ollama(question, chunks):
    if not chunks:
        return (
            "I couldn't find anything relevant in my knowledge base. "
            "Try rephrasing your question or check that related documents are loaded."
        ), []

    context = "\n\n".join(
        [f"[Source: {c['filename']} ({c['file_type'].upper()})]\n{c['text']}" for c in chunks]
    )
    sources = list(set([c["source"] for c in chunks]))

    prompt = f"""You are Libby, an offline survival knowledge assistant.
Your job is to answer the user's question using ONLY the information provided below.
The information may come from text files or Excel spreadsheets.
Do not use any outside knowledge. Do not guess or make things up.
If the provided information does not contain a clear answer, say so honestly.
Keep your answer clear, practical, and helpful.

--- KNOWLEDGE BASE CONTEXT ---
{context}
--- END OF CONTEXT ---

User question: {question}

Answer:"""

    try:
        response = requests.post(
            OLLAMA_URL,
            json={"model": OLLAMA_MODEL, "prompt": prompt, "stream": False},
            timeout=120
        )
        response.raise_for_status()
        return response.json().get("response", "").strip(), sources
    except requests.exceptions.ConnectionError:
        return "⚠️  Ollama is not running. Open a new CMD window and type: ollama serve", []
    except Exception as e:
        return f"⚠️  Error: {str(e)}", []


# ─────────────────────────────────────────────
# MAIN APP
# ─────────────────────────────────────────────
class LibbyApp:
    def __init__(self, root):
        self.root          = root
        self.ready         = False
        self.current_theme = "dark"
        self.T             = THEMES["dark"]
        self.dot_count     = 0
        self.animating     = False
        self.chat_widgets  = []

        self.root.title("Libby — Offline Knowledge Assistant")
        self.root.geometry("1050x720")
        self.root.minsize(700, 500)
        self.root.configure(bg=self.T["BG_DARK"])

        self._build_ui()
        self._start_backend()

    def _build_ui(self):
        T = self.T

        # ── Header ──────────────────────────
        self.header = tk.Frame(self.root, bg=T["BG_PANEL"], pady=12)
        self.header.pack(fill="x")

        self.avatar_canvas = tk.Canvas(
            self.header, width=48, height=48,
            bg=T["BG_PANEL"], highlightthickness=0
        )
        self.avatar_canvas.pack(side="left", padx=(16, 8))
        self._draw_avatar()

        title_frame = tk.Frame(self.header, bg=T["BG_PANEL"])
        title_frame.pack(side="left")

        self.title_label = tk.Label(
            title_frame, text="LIBBY",
            bg=T["BG_PANEL"], fg=T["ACCENT"],
            font=("Courier New", 20, "bold")
        )
        self.title_label.pack(anchor="w")

        self.subtitle_label = tk.Label(
            title_frame, text="Offline Knowledge Assistant",
            bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
            font=("Courier New", 9)
        )
        self.subtitle_label.pack(anchor="w")

        right_frame = tk.Frame(self.header, bg=T["BG_PANEL"])
        right_frame.pack(side="right", padx=16)

        self.theme_btn = tk.Button(
            right_frame, text="☀️  Light Mode",
            bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
            activebackground=T["BG_INPUT"],
            font=("Courier New", 9),
            relief="flat", cursor="hand2",
            command=self._toggle_theme
        )
        self.theme_btn.pack(anchor="e", pady=(0, 4))

        self.status_label = tk.Label(
            right_frame, text="⏳  Loading...",
            bg=T["BG_PANEL"], fg=T["WARNING"],
            font=("Courier New", 9)
        )
        self.status_label.pack(anchor="e")

        self.top_divider = tk.Frame(self.root, bg=T["BORDER"], height=1)
        self.top_divider.pack(fill="x")

        self.body = tk.Frame(self.root, bg=T["BG_DARK"])
        self.body.pack(fill="both", expand=True)

        # ── Sidebar ─────────────────────────
        self.sidebar = tk.Frame(self.body, bg=T["BG_SIDEBAR"], width=220)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)

        self.sidebar_title = tk.Label(
            self.sidebar, text="📁  KNOWLEDGE BASE",
            bg=T["BG_SIDEBAR"], fg=T["ACCENT"],
            font=("Courier New", 9, "bold"), pady=12
        )
        self.sidebar_title.pack(anchor="w", padx=12)

        self.sidebar_divider = tk.Frame(self.sidebar, bg=T["BORDER"], height=1)
        self.sidebar_divider.pack(fill="x", padx=8)

        self.sidebar_list = tk.Frame(self.sidebar, bg=T["BG_SIDEBAR"])
        self.sidebar_list.pack(fill="both", expand=True, padx=8, pady=8)

        self.sidebar_footer = tk.Label(
            self.sidebar,
            text="📄 = Text   📊 = Excel",
            bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"],
            font=("Courier New", 7), pady=8
        )
        self.sidebar_footer.pack(side="bottom")

        self.side_divider = tk.Frame(self.body, bg=T["BORDER"], width=1)
        self.side_divider.pack(side="left", fill="y")

        # ── Chat Area ───────────────────────
        chat_outer = tk.Frame(self.body, bg=T["BG_PANEL"])
        chat_outer.pack(side="left", fill="both", expand=True)

        self.chat_canvas = tk.Canvas(
            chat_outer, bg=T["BG_PANEL"],
            highlightthickness=0, bd=0
        )
        scrollbar = tk.Scrollbar(chat_outer, command=self.chat_canvas.yview)
        self.chat_canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.chat_canvas.pack(side="left", fill="both", expand=True)

        self.chat_inner = tk.Frame(self.chat_canvas, bg=T["BG_PANEL"])
        self.chat_window = self.chat_canvas.create_window(
            (0, 0), window=self.chat_inner, anchor="nw"
        )
        self.chat_inner.bind("<Configure>", self._on_frame_configure)
        self.chat_canvas.bind("<Configure>", self._on_canvas_configure)
        self.chat_canvas.bind_all("<MouseWheel>", self._on_mousewheel)

        self.bottom_divider = tk.Frame(self.root, bg=T["BORDER"], height=1)
        self.bottom_divider.pack(fill="x")

        # ── Input Area ──────────────────────
        self.input_frame = tk.Frame(self.root, bg=T["BG_DARK"], pady=10)
        self.input_frame.pack(fill="x", padx=12)

        self.input_box = tk.Text(
            self.input_frame, height=3,
            bg=T["BG_INPUT"], fg=T["TEXT_MAIN"],
            insertbackground=T["ACCENT"],
            font=("Courier New", 11),
            relief="flat", bd=0,
            padx=12, pady=10, wrap="word"
        )
        self.input_box.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self.input_box.bind("<Return>", self._on_enter)
        self.input_box.insert("1.0", "Ask Libby a question...")
        self.input_box.config(fg=T["TEXT_DIM"])
        self.input_box.bind("<FocusIn>", self._clear_placeholder)
        self.input_box.bind("<FocusOut>", self._restore_placeholder)

        btn_frame = tk.Frame(self.input_frame, bg=T["BG_DARK"])
        btn_frame.pack(side="right")

        self.send_btn = tk.Button(
            btn_frame, text="ASK",
            bg=T["ACCENT_DIM"], fg=T["TEXT_MAIN"],
            activebackground=T["ACCENT"],
            font=("Courier New", 11, "bold"),
            relief="flat", padx=16, pady=6,
            cursor="hand2", command=self._send_question
        )
        self.send_btn.pack(pady=(0, 4))

        self.clear_btn = tk.Button(
            btn_frame, text="🗑  Clear",
            bg=T["BTN_CLEAR"], fg=T["BTN_CLEAR_FG"],
            activebackground=T["BTN_CLEAR"],
            font=("Courier New", 9),
            relief="flat", padx=16, pady=4,
            cursor="hand2", command=self._clear_chat
        )
        self.clear_btn.pack()

        self.footer = tk.Frame(self.root, bg=T["BG_DARK"], pady=5)
        self.footer.pack(fill="x")
        self.footer_label = tk.Label(
            self.footer,
            text="🔒  Air-gapped mode  •  Supports .txt and .xlsx files",
            bg=T["BG_DARK"], fg=T["TEXT_DIM"],
            font=("Courier New", 8)
        )
        self.footer_label.pack()

        self._add_libby_message(
            "Hello! I'm Libby, your offline knowledge assistant. "
            "I can read both text files and Excel spreadsheets. "
            "Ask me anything from your knowledge base!",
            source=None, file_type=None
        )

    def _draw_avatar(self):
        c = self.avatar_canvas
        T = self.T
        c.delete("all")
        c.create_rectangle(8, 6, 40, 42, fill=T["ACCENT"], outline="", width=0)
        c.create_rectangle(8, 6, 14, 42, fill=T["ACCENT_DIM"], outline="", width=0)
        c.create_rectangle(14, 10, 38, 38, fill=T["BG_PANEL"], outline="", width=0)
        for y in [16, 21, 26, 31]:
            c.create_line(17, y, 35, y, fill=T["BORDER"], width=1)
        c.create_polygon(32, 6, 38, 6, 38, 16, 35, 13, 32, 16, fill=T["WARNING"], outline="")

    def _toggle_theme(self):
        self.current_theme = "light" if self.current_theme == "dark" else "dark"
        self.T = THEMES[self.current_theme]
        self.theme_btn.config(
            text="🌙  Dark Mode" if self.current_theme == "light" else "☀️  Light Mode"
        )
        self._apply_theme()

    def _apply_theme(self):
        T = self.T
        self.root.configure(bg=T["BG_DARK"])
        self.header.configure(bg=T["BG_PANEL"])
        self.avatar_canvas.configure(bg=T["BG_PANEL"])
        self._draw_avatar()
        self.title_label.configure(bg=T["BG_PANEL"], fg=T["ACCENT"])
        self.subtitle_label.configure(bg=T["BG_PANEL"], fg=T["TEXT_DIM"])
        self.theme_btn.configure(bg=T["BG_PANEL"], fg=T["TEXT_DIM"])
        self.status_label.configure(bg=T["BG_PANEL"])
        self.top_divider.configure(bg=T["BORDER"])
        self.body.configure(bg=T["BG_DARK"])
        self.sidebar.configure(bg=T["BG_SIDEBAR"])
        self.sidebar_title.configure(bg=T["BG_SIDEBAR"], fg=T["ACCENT"])
        self.sidebar_divider.configure(bg=T["BORDER"])
        self.sidebar_list.configure(bg=T["BG_SIDEBAR"])
        self.sidebar_footer.configure(bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"])
        self.side_divider.configure(bg=T["BORDER"])
        self.chat_canvas.configure(bg=T["BG_PANEL"])
        self.chat_inner.configure(bg=T["BG_PANEL"])
        self.bottom_divider.configure(bg=T["BORDER"])
        self.input_frame.configure(bg=T["BG_DARK"])
        self.input_box.configure(bg=T["BG_INPUT"], fg=T["TEXT_MAIN"], insertbackground=T["ACCENT"])
        self.send_btn.configure(bg=T["ACCENT_DIM"], fg=T["TEXT_MAIN"])
        self.clear_btn.configure(bg=T["BTN_CLEAR"], fg=T["BTN_CLEAR_FG"])
        self.footer.configure(bg=T["BG_DARK"])
        self.footer_label.configure(bg=T["BG_DARK"], fg=T["TEXT_DIM"])
        self._refresh_sidebar()
        self._refresh_chat_bubbles()

    def _refresh_sidebar(self):
        T = self.T
        for widget in self.sidebar_list.winfo_children():
            widget.destroy()
        for entry in known_files:
            filename  = entry["filename"]
            folder    = os.path.basename(os.path.dirname(entry["source"]))
            file_type = entry.get("file_type", "txt")
            icon      = "📊" if file_type == "xlsx" else "📄"
            row = tk.Frame(self.sidebar_list, bg=T["BG_SIDEBAR"])
            row.pack(anchor="w", fill="x", pady=2)
            tk.Label(
                row, text=f"{icon} {filename}",
                bg=T["BG_SIDEBAR"], fg=T["TEXT_MAIN"],
                font=("Courier New", 8),
                wraplength=180, justify="left", anchor="w"
            ).pack(anchor="w")
            tk.Label(
                row, text=f"   📁 {folder}",
                bg=T["BG_SIDEBAR"], fg=T["TEXT_DIM"],
                font=("Courier New", 7), anchor="w"
            ).pack(anchor="w", pady=(0, 4))

    def _add_user_message(self, text):
        T     = self.T
        outer = tk.Frame(self.chat_inner, bg=T["BG_PANEL"], pady=6)
        outer.pack(fill="x", padx=10)
        self.chat_widgets.append(("user", text, None, None))
        tk.Label(outer, text="YOU", bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
                 font=("Courier New", 8, "bold")).pack(anchor="e", padx=6)
        bubble = tk.Frame(outer, bg=T["BG_BUBBLE_U"], padx=14, pady=10)
        bubble.pack(anchor="e", padx=6)
        tk.Label(bubble, text=text, bg=T["BG_BUBBLE_U"], fg=T["TEXT_MAIN"],
                 font=("Courier New", 11), wraplength=560, justify="left").pack(anchor="w")

    def _add_libby_message(self, text, source=None, file_type=None):
        T     = self.T
        outer = tk.Frame(self.chat_inner, bg=T["BG_PANEL"], pady=6)
        outer.pack(fill="x", padx=10)
        self.chat_widgets.append(("libby", text, source, file_type))
        tk.Label(outer, text="LIBBY", bg=T["BG_PANEL"], fg=T["ACCENT"],
                 font=("Courier New", 8, "bold")).pack(anchor="w", padx=6)
        bubble = tk.Frame(outer, bg=T["BG_BUBBLE_L"], padx=14, pady=10)
        bubble.pack(anchor="w", padx=6)
        tk.Label(bubble, text=text, bg=T["BG_BUBBLE_L"], fg=T["TEXT_MAIN"],
                 font=("Courier New", 11), wraplength=580, justify="left").pack(anchor="w")
        if source:
            icon = "📊" if file_type == "xlsx" else "📄"
            tk.Label(
                bubble, text=f"{icon}  Source: {source}",
                bg=T["BG_BUBBLE_L"], fg=T["TEXT_SOURCE"],
                font=("Courier New", 8), wraplength=580, justify="left"
            ).pack(anchor="w", pady=(6, 0))

    def _refresh_chat_bubbles(self):
        saved = list(self.chat_widgets)
        for widget in self.chat_inner.winfo_children():
            widget.destroy()
        self.chat_widgets = []
        for kind, text, source, file_type in saved:
            if kind == "user":
                self._add_user_message(text)
            else:
                self._add_libby_message(text, source, file_type)

    def _add_thinking_frame(self):
        T     = self.T
        frame = tk.Frame(self.chat_inner, bg=T["BG_PANEL"], pady=6)
        frame.pack(fill="x", padx=10)
        tk.Label(frame, text="LIBBY", bg=T["BG_PANEL"], fg=T["ACCENT"],
                 font=("Courier New", 8, "bold")).pack(anchor="w", padx=6)
        self.thinking_label = tk.Label(
            frame, text="⏳  Thinking",
            bg=T["BG_PANEL"], fg=T["TEXT_DIM"],
            font=("Courier New", 10, "italic")
        )
        self.thinking_label.pack(anchor="w", padx=10)
        self.animating = True
        self._animate_dots()
        self._scroll_to_bottom()
        return frame

    def _animate_dots(self):
        if not self.animating:
            return
        dots = "." * (self.dot_count % 4)
        self.thinking_label.config(text=f"⏳  Thinking{dots}")
        self.dot_count += 1
        self.root.after(400, self._animate_dots)

    def _stop_animation(self, frame):
        self.animating = False
        frame.destroy()

    def _clear_chat(self):
        for widget in self.chat_inner.winfo_children():
            widget.destroy()
        self.chat_widgets = []
        self._add_libby_message(
            "Chat cleared! Ask me anything.",
            source=None, file_type=None
        )

    def _on_frame_configure(self, event):
        self.chat_canvas.configure(scrollregion=self.chat_canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.chat_canvas.itemconfig(self.chat_window, width=event.width)

    def _on_mousewheel(self, event):
        self.chat_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _scroll_to_bottom(self):
        self.root.after(50, lambda: self.chat_canvas.yview_moveto(1.0))

    def _clear_placeholder(self, event):
        if self.input_box.get("1.0", "end-1c") == "Ask Libby a question...":
            self.input_box.delete("1.0", "end")
            self.input_box.config(fg=self.T["TEXT_MAIN"])

    def _restore_placeholder(self, event):
        if not self.input_box.get("1.0", "end-1c").strip():
            self.input_box.insert("1.0", "Ask Libby a question...")
            self.input_box.config(fg=self.T["TEXT_DIM"])

    def _on_enter(self, event):
        if not event.state & 0x1:
            self._send_question()
            return "break"

    def _send_question(self):
        if not self.ready:
            return
        question = self.input_box.get("1.0", "end-1c").strip()
        if not question or question == "Ask Libby a question...":
            return
        self.input_box.delete("1.0", "end")
        self.input_box.config(fg=self.T["TEXT_MAIN"])
        self._add_user_message(question)
        thinking_frame = self._add_thinking_frame()
        self.send_btn.config(state="disabled", bg=self.T["BG_INPUT"])
        self.input_box.config(state="disabled")

        def run_query():
            chunks          = retrieve(question)
            answer, sources = ask_ollama(question, chunks)
            source_str      = ", ".join(sources) if sources else None
            file_type       = chunks[0]["file_type"] if chunks else None
            self.root.after(0, lambda: self._show_answer(answer, source_str, file_type, thinking_frame))

        threading.Thread(target=run_query, daemon=True).start()

    def _show_answer(self, answer, source, file_type, thinking_frame):
        self._stop_animation(thinking_frame)
        self._add_libby_message(answer, source=source, file_type=file_type)
        self._scroll_to_bottom()
        self.send_btn.config(state="normal", bg=self.T["ACCENT_DIM"])
        self.input_box.config(state="normal")
        self.input_box.focus()

    def _start_backend(self):
        def load():
            load_backend()
            self.ready = True
            count = collection.count()
            self.root.after(0, lambda: self.status_label.config(
                text=f"✅  {count} chunks loaded  •  Offline",
                fg=self.T["SUCCESS"]
            ))
            self.root.after(0, self._refresh_sidebar)

        threading.Thread(target=load, daemon=True).start()


# ─────────────────────────────────────────────
# LAUNCH
# ─────────────────────────────────────────────
if __name__ == "__main__":
    root = tk.Tk()
    app  = LibbyApp(root)
    root.mainloop()